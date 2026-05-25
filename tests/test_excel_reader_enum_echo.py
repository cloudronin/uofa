"""Regression: enum-echo and template-placeholder-leak surfaced by OpenAI extract.

Bug context (May 25, 2026 — discovered during the post-Anthropic real-LLM
e2e validation with ``openai/gpt-4o-mini``): GPT-4o-mini extracted Morrison
COU1 evidence and emitted prompt enumerator-text verbatim into the Profile
cell instead of picking one option:

  Sheet 'Assessment Summary', cell D3: 'Minimal or Complete' is not a
  valid profile. Expected: Minimal, Complete

Five additional cells in the same xlsx showed template placeholder hint
text leaking through (COU Description, Assessor Name, Assessment Date,
Source Document, Criteria Set) because the writer skipped cells when the
LLM returned None for the field — leaving the template's editor-hint
text intact and indistinguishable from extracted data.

Two-axis fix (writer + reader):

1. ``excel_writer._fuzzy_match_dropdown`` enhanced to split on enumerator
   separators (" or ", " / ", ",") and pick the first valid token.
   "Minimal or Complete" → "Minimal".
2. ``excel_writer._write_summary_sheet`` / ``_write_decision_sheet`` now
   clear cells when extraction is None instead of leaving template text.
3. ``excel_reader._read_summary`` / ``_read_decision`` lenient mode
   (mirrors evidence_type pattern from 1a0e831): enum-echo values are
   normalized via the same enumerator-split logic and a per-field warning
   is appended instead of erroring.

This file pins the actual offending bytes from the real-LLM run so the
bug cannot regress silently. Closes #24 (template-placeholder leak).

Fixture provenance:
``tests/fixtures/regression/extract-enum-echo/morrison-cou1-openai-extracted.xlsx``
— openai/gpt-4o-mini output on morrison-evidence-cou1, May 25 2026.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ENUM_ECHO_XLSX = (Path(__file__).parent / "fixtures" / "regression"
                  / "extract-enum-echo"
                  / "morrison-cou1-openai-extracted.xlsx")

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")


@pytest.mark.skipif(not ENUM_ECHO_XLSX.exists(),
                    reason=f"missing fixture: {ENUM_ECHO_XLSX}")
class TestMorrisonCou1EnumEchoNormalizes:
    """The OpenAI fixture has Profile = 'Minimal or Complete' (verbatim
    enumerator-text echo from the prompt). Importing it must succeed via
    lenient mode, the Profile must normalize to a canonical value, and a
    warning must explain the substitution.
    """

    def test_import_succeeds(self):
        from uofa_cli.excel_reader import read_workbook
        data = read_workbook(ENUM_ECHO_XLSX, ["vv40"])
        # Strict mode would raise ImportError on the enum-echo. Reaching
        # this assertion at all proves lenient mode kicked in.
        assert data is not None
        assert "summary" in data

    def test_profile_normalized_to_canonical(self):
        from uofa_cli.excel_reader import read_workbook
        from uofa_cli.excel_constants import VALID_PROFILES
        data = read_workbook(ENUM_ECHO_XLSX, ["vv40"])
        profile = data["summary"]["profile"]
        assert profile in VALID_PROFILES, (
            f"profile {profile!r} should be in {VALID_PROFILES} after normalization"
        )
        # Enum-split picks the first token from "Minimal or Complete" → "Minimal".
        assert profile == "Minimal", (
            f"expected enum-split to pick first token 'Minimal'; got {profile!r}"
        )

    def test_profile_warning_emitted(self):
        from uofa_cli.excel_reader import read_workbook
        data = read_workbook(ENUM_ECHO_XLSX, ["vv40"])
        warnings = data.get("_warnings", [])
        profile_warnings = [w for w in warnings if "canonical profile" in w]
        assert len(profile_warnings) == 1, (
            f"expected 1 profile-normalization warning, got {len(profile_warnings)}: {warnings}"
        )
        # Warning should name the original LLM output for transparency
        assert "Minimal or Complete" in profile_warnings[0]
        # ... and the canonical it normalized to
        assert "'Minimal'" in profile_warnings[0]


@pytest.mark.skipif(not ENUM_ECHO_XLSX.exists(),
                    reason=f"missing fixture: {ENUM_ECHO_XLSX}")
class TestPlaceholderLeakClearedByWriter:
    """Closes #24 — verify the writer-side placeholder-clear fix.

    Re-extract the fixture's xlsx via the writer on a synthetic empty
    ExtractionResult: cells that have no extraction must be None (cleared),
    not template hint text. Validates the writer-side fix in isolation
    from the reader-side leniency.
    """

    def test_writer_clears_summary_cells_when_extraction_is_none(self, tmp_path):
        import shutil
        import openpyxl
        from uofa_cli.llm_extractor import ExtractionResult
        from uofa_cli.excel_writer import write_extraction
        from uofa_cli import paths

        # Use the vv40 template directly so the placeholder hints are
        # the same as the user-facing path.
        template = paths.pack_dir("vv40") / "templates" / "vv40-template.xlsx"
        if not template.exists():
            pytest.skip(f"missing template: {template}")

        # Synthetic ExtractionResult with empty fields → writer should
        # clear all summary cells (no template hint should survive).
        result = ExtractionResult(
            assessment_summary={},  # NO extractions
            model_and_data=[],
            validation_results=[],
            credibility_factors=[],
            decision={},
        )
        out = tmp_path / "empty-extract.xlsx"
        write_extraction(result, template, out, "vv40")

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Assessment Summary"]
        # Find the data row (typically row 3). Check every cell in the
        # field-map columns is None / empty, not template hint text.
        # Profile column is D (col 4). If the placeholder leak is still
        # present this would be 'Minimal or Complete'.
        profile_cell = ws.cell(row=3, column=4).value
        assert profile_cell is None or profile_cell == "", (
            f"Profile cell should be cleared on empty extraction, "
            f"got {profile_cell!r} (template-hint leak)"
        )
        # Assessor Name column is I (col 9). Template hint is
        # 'Person or organization responsible'.
        assessor_cell = ws.cell(row=3, column=9).value
        assert assessor_cell is None or assessor_cell == "", (
            f"Assessor Name cell should be cleared on empty extraction, "
            f"got {assessor_cell!r} (template-hint leak)"
        )


def test_fuzzy_match_dropdown_handles_enum_echo():
    """Unit test for the enum-split logic in _fuzzy_match_dropdown.

    Locks in the three real-world LLM-output patterns observed this
    session so they can't regress without a test flagging it.
    """
    from uofa_cli.excel_writer import _fuzzy_match_dropdown
    from uofa_cli.excel_constants import (
        VALID_PROFILES, VALID_ASSURANCE_LEVELS, VALID_DECISION_OUTCOMES,
    )

    # Exact match still passes through unchanged
    assert _fuzzy_match_dropdown("Minimal", VALID_PROFILES) == "Minimal"
    assert _fuzzy_match_dropdown("Complete", VALID_PROFILES) == "Complete"

    # GPT-4o-mini Morrison case: " or " separator
    assert _fuzzy_match_dropdown("Minimal or Complete", VALID_PROFILES) == "Minimal"

    # Generic enumerator cases: " / " separator
    assert _fuzzy_match_dropdown("Low / Medium / High", VALID_ASSURANCE_LEVELS) == "Low"

    # Decision outcome case (also " / " with multi-word values)
    assert _fuzzy_match_dropdown(
        "Accepted / Not accepted / Conditional",
        VALID_DECISION_OUTCOMES + ["Conditional"],
    ) == "Accepted"

    # Unrecognized value still falls through unchanged (caller decides)
    assert _fuzzy_match_dropdown("Totally Made Up", VALID_PROFILES) == "Totally Made Up"
