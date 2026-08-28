"""protocol-check — the scriptable subset of the encoding protocol.

Fixtures are built from the shipped nasa-7009b template rather than hand-authored, so a
template change surfaces here instead of silently making the checks vacuous.
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")

from uofa_cli import protocol_check  # noqa: E402
from uofa_cli.protocol_check import ANCHOR_HEADER, check_package, check_workbook  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
TEMPLATE = REPO / "packs" / "nasa-7009b" / "templates" / "nasa-7009b-template.xlsx"
REVIEWED = REPO / "dev" / "build" / "pilot-johnson" / "johnson-extracted.xlsx"


def _verdict(results, name_fragment):
    for r in results:
        if name_fragment in r.name:
            return r
    raise AssertionError(f"no check named like {name_fragment!r} in {[r.name for r in results]}")


def _add_anchor_column(path: Path, value: str = "p.1") -> None:
    """Give every populated data row an anchor, the state a reviewed workbook is in."""
    wb = openpyxl.load_workbook(path)
    for sheet, (first_header, offset) in protocol_check._SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        head = protocol_check._sheet_header_row(ws, first_header)
        if head is None:
            continue
        last = max(c for c in range(1, 40) if ws.cell(row=head, column=c).value) or 1
        ws.cell(row=head, column=last + 1).value = ANCHOR_HEADER
        for row in range(head + offset, ws.max_row + 1):
            if any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, last + 1)):
                ws.cell(row=row, column=last + 1).value = value
    wb.save(path)


def _blank_template(tmp_path: Path) -> Path:
    """The template with its description rows cleared, so only seeded defects remain."""
    out = tmp_path / "wb.xlsx"
    shutil.copy2(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    for sheet, (first_header, _) in protocol_check._SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        head = protocol_check._sheet_header_row(ws, first_header)
        if head is None:
            continue
        for col in range(1, 40):
            ws.cell(row=head + 1, column=col).value = None
    wb.save(out)
    return out


# ── workbook-side ────────────────────────────────────────────────────────────

def test_missing_anchor_column_fails(tmp_path):
    wb = _blank_template(tmp_path)
    assert not _verdict(check_workbook(wb, TEMPLATE), "anchor column present").passed


def test_anchor_column_present_passes(tmp_path):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    assert _verdict(check_workbook(wb, TEMPLATE), "anchor column present").passed


def test_unanchored_populated_row_fails(tmp_path):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    book = openpyxl.load_workbook(wb)
    ws = book["Credibility Factors"]
    acol = protocol_check._anchor_col(ws)
    ws.cell(row=5, column=acol).value = None  # a factor row loses its anchor
    book.save(wb)
    result = _verdict(check_workbook(wb, TEMPLATE), "anchor non-empty")
    assert not result.passed
    assert "Credibility Factors row 5" in result.detail


def test_template_placeholder_left_in_data_row_fails(tmp_path):
    """The exact defect the pilot's raw extract shipped six times (F-3d)."""
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    hints = protocol_check._template_placeholders(TEMPLATE)
    assert hints, "template carries no description-row hints to detect"
    book = openpyxl.load_workbook(wb)
    ws = book["Assessment Summary"]
    head = protocol_check._sheet_header_row(ws, "Project Name")
    ws.cell(row=head + 2, column=1).value = sorted(hints)[0]
    acol = protocol_check._anchor_col(ws)
    ws.cell(row=head + 2, column=acol).value = "p.1"
    book.save(wb)
    assert not _verdict(check_workbook(wb, TEMPLATE), "placeholder").passed


def test_placeholder_check_skipped_without_a_template(tmp_path):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    assert _verdict(check_workbook(wb, None), "placeholder").skipped


def _set_levels(path: Path, pairs: list[tuple[int, int]], criteria: str | None = None) -> None:
    book = openpyxl.load_workbook(path)
    ws = book["Credibility Factors"]
    head = protocol_check._sheet_header_row(ws, "Factor Type")
    for i, (req, ach) in enumerate(pairs):
        row = head + 2 + i
        ws.cell(row=row, column=3).value = req
        ws.cell(row=row, column=4).value = ach
        if criteria:
            ws.cell(row=row, column=5).value = criteria
    book.save(path)


def test_a_legacy_package_all_equal_warns_rather_than_refusing(tmp_path):
    """The contract change, stated as a test.

    This asserted a hard failure, on the reasoning that the extract prompt sets
    required equal to achieved by default, so all-equal means nobody looked.
    The reasoning is sound and the test was a proxy: **agreement writes
    nothing**, so a reviewer who read every level and agreed with all of them
    produces a byte-identical workbook. Shape-inference punishes honest
    agreement and misses nothing else.

    A package that cannot SAY whether the judgment happened is now warned rather
    than refused, so third parties scripting against exit codes keep their
    contract on the legacy path. A package that can say is asked -- see below.
    """
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert result.passed, "a legacy package was hard-refused on shape alone"
    assert result.skipped, "the warning is not carried as advisory"
    assert "cannot state whether the column was reviewed" in result.detail


def _declare_profile(wb, version="v0.8"):
    """Write the encoding-profile declaration a modern encoder ships.

    Separate from `_mark_levels` because the two are separable in reality, and
    the interesting failures are exactly the cases where they disagree.
    """
    import openpyxl
    from uofa_cli.excel_constants import WORKBOOK_PROFILE_HEADER
    from uofa_cli.protocol_check import _sheet_header_row
    book = openpyxl.load_workbook(wb)
    ws = book["Assessment Summary"]
    head = _sheet_header_row(ws, "Project Name")
    col = ws.max_column + 1
    ws.cell(row=head, column=col).value = WORKBOOK_PROFILE_HEADER
    ws.cell(row=head + 1, column=col).value = version
    book.save(wb)


def _mark_levels(wb, tokens, declare="v0.8"):
    """Write the review-provenance column a modern encoder ships."""
    import openpyxl
    from uofa_cli.protocol_check import LEVEL_PROVENANCE_HEADER
    if declare:
        _declare_profile(wb, declare)
    book = openpyxl.load_workbook(wb)
    ws = book["Credibility Factors"]
    head = 3
    col = ws.max_column + 1
    ws.cell(row=head, column=col).value = LEVEL_PROVENANCE_HEADER
    for i, token in enumerate(tokens):
        ws.cell(row=head + 2 + i, column=col).value = token
    book.save(wb)


def test_all_equal_passes_when_every_level_carries_a_review_act(tmp_path):
    """Honest agreement stops being punishable.

    A reviewer who reads seventeen required levels and agrees with all of them
    produces the same VALUES as one who never opened the sheet. The old check
    refused both. This one asks the package whether the judgment happened.
    """
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _mark_levels(wb, ["affirmed", "corrected", "affirmed"])
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert result.passed, result.detail
    assert "the package says so" in result.detail


def test_one_defaulted_level_refuses_and_names_the_cell(tmp_path):
    """The direction that catches run 25: defaults nobody went back to.

    The refusal must name WHICH cells and what discharges them -- a check that
    says only "unreviewed" sends the reader to re-read seventeen rows.
    """
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _mark_levels(wb, ["affirmed", "defaulted", "affirmed"])
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert not result.passed
    assert "1 of 3" in result.detail
    assert "weighed them against the achieved" in result.detail
    assert "Affirm each in the encoding tool" in result.detail


def test_a_waiver_still_releases_the_evidence_path(tmp_path):
    """A-7's recorded waiver is a judgment too, and outranks the defaults."""
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4)],
                criteria="Validation waived by Technical Authority")
    _mark_levels(wb, ["defaulted", "defaulted"])
    assert _verdict(check_workbook(wb, TEMPLATE), "required levels").passed


def test_one_differing_factor_passes(tmp_path):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (2, 4), (4, 4)])
    assert _verdict(check_workbook(wb, TEMPLATE), "required levels").passed


def test_recorded_waiver_releases_the_level_check(tmp_path):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4)],
                criteria="Validation waived by Technical Authority")
    assert _verdict(check_workbook(wb, TEMPLATE), "required levels").passed


@pytest.mark.skipif(not REVIEWED.exists(), reason="pilot workbook not present")
def test_reviewed_pilot_workbook_passes_every_workbook_check():
    for r in check_workbook(REVIEWED, TEMPLATE):
        assert r.passed or r.skipped, f"{r.name}: {r.detail}"


# ── package-side ─────────────────────────────────────────────────────────────

GOOD_RUN_LOG = (
    "# Run log\n\n"
    "model: anthropic/claude-sonnet-5\nbackend: anthropic\n"
    "site commit: 31cb466\nrepo HEAD: abc1234\nbase_uri: https://uofa.net\n"
)


def _package_dir(tmp_path: Path, *, ambiguity: str | None, run_log: str | None,
                 package_id: str | None = None) -> Path:
    pkg = tmp_path / "pkg.jsonld"
    pkg.write_text(json.dumps({"id": package_id} if package_id else {}), encoding="utf8")
    if ambiguity is not None:
        (tmp_path / "AMBIGUITY_LOG.md").write_text(ambiguity, encoding="utf8")
    if run_log is not None:
        (tmp_path / "RUN_LOG.md").write_text(run_log, encoding="utf8")
    return pkg


def test_missing_ambiguity_log_fails(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity=None, run_log=GOOD_RUN_LOG)
    assert not _verdict(check_package(pkg), "ambiguity log").passed


def test_empty_ambiguity_log_fails(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="   \n", run_log=GOOD_RUN_LOG)
    assert not _verdict(check_package(pkg), "ambiguity log").passed


def test_complete_run_log_passes(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...", run_log=GOOD_RUN_LOG)
    for r in check_package(pkg):
        assert r.passed or r.skipped, f"{r.name}: {r.detail}"


def test_run_log_missing_a_pin_fails(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...",
                       run_log=GOOD_RUN_LOG.replace("base_uri: https://uofa.net\n", ""))
    result = _verdict(check_package(pkg), "carries its pins")
    assert not result.passed
    assert "base_uri" in result.detail


def test_signing_in_a_pilot_run_log_fails(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...",
                       run_log=GOOD_RUN_LOG + "\npilot run\nuofa import x.xlsx --sign --key k\n")
    assert not _verdict(check_package(pkg), "no signing").passed


def test_prohibition_of_signing_is_not_a_use(tmp_path):
    """"no --sign" in a pilot log is the rule being stated, not broken."""
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...",
                       run_log=GOOD_RUN_LOG + "\npilot run\nImport target: no --sign anywhere.\n")
    assert _verdict(check_package(pkg), "no signing").passed


def test_signing_check_skipped_when_run_log_is_not_pilot_labeled(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...",
                       run_log=GOOD_RUN_LOG + "\nuofa import x.xlsx --sign --key k\n")
    assert _verdict(check_package(pkg), "no signing").skipped


# ── namespace ────────────────────────────────────────────────────────────────

# The whole reserved family, one namespace per branch of the matcher. The first entry is
# the exact namespace an encoder following A-2 chose while clearing the narrower check
# this one replaces.
RESERVED_NAMESPACES = [
    "https://reference-encodings.example.net/pkg/1",
    "https://example.org/pkg/1",
    "https://example.com/pkg/1",
    "https://example.net/pkg/1",
    "https://sub.example.org/pkg/1",
    "https://encodings.test/pkg/1",
    "https://encodings.invalid/pkg/1",
    "https://encodings.localhost/pkg/1",
    "https://encodings.example/pkg/1",
    "https://example.acme.com/pkg/1",
]

# Namespaces that read like the reserved ones without being them. A check that fires on
# these would refuse namespaces an encoder does control, which is the opposite failure.
CONTROLLED_NAMESPACES = [
    "https://github.com/cloudronin/uofa/pkg/1",
    "https://uofa.net/pkg/1",
    "https://myexample.com/pkg/1",
    "https://examples.org/pkg/1",
    "https://acme.example-corp.com/pkg/1",
]


@pytest.mark.parametrize("namespace", RESERVED_NAMESPACES)
def test_reserved_namespace_fails(tmp_path, namespace):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...", run_log=GOOD_RUN_LOG,
                       package_id=namespace)
    result = _verdict(check_package(pkg), "namespace")
    assert not result.passed, namespace
    assert not result.skipped


@pytest.mark.parametrize("namespace", CONTROLLED_NAMESPACES)
def test_controlled_namespace_passes(tmp_path, namespace):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...", run_log=GOOD_RUN_LOG,
                       package_id=namespace)
    assert _verdict(check_package(pkg), "namespace").passed, namespace


def test_reserved_base_uri_in_the_run_log_fails_even_when_the_id_is_clean(tmp_path):
    """The run log declares the namespace; a clean id does not excuse a reserved pin."""
    run_log = GOOD_RUN_LOG.replace("https://uofa.net", "https://example.org")
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...", run_log=run_log,
                       package_id="https://github.com/cloudronin/uofa/pkg/1")
    result = _verdict(check_package(pkg), "namespace")
    assert not result.passed
    assert "base_uri" in result.detail


def test_namespace_check_skips_with_nothing_to_read(tmp_path):
    """No minted id and no declared base_uri skips rather than passing vacuously."""
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...",
                       run_log=GOOD_RUN_LOG.replace("base_uri: https://uofa.net\n", ""))
    assert _verdict(check_package(pkg), "namespace").skipped


def test_namespace_check_survives_an_unparseable_package(tmp_path):
    pkg = _package_dir(tmp_path, ambiguity="A-01 ...", run_log=GOOD_RUN_LOG)
    pkg.write_text("{ not json", encoding="utf8")
    assert _verdict(check_package(pkg), "namespace").passed  # falls back to the run log


# ── CLI wiring ───────────────────────────────────────────────────────────────

def _uofa(*args, cwd=REPO):
    return subprocess.run([sys.executable, "-m", "uofa_cli", *args],
                          cwd=cwd, capture_output=True, text=True)


@pytest.mark.skipif(not REVIEWED.exists(), reason="pilot workbook not present")
def test_import_protocol_check_exits_non_zero_on_failure(tmp_path):
    """A package with no ambiguity log beside it must not pass the gate.

    Uses the reviewed pilot workbook because the gate runs after import succeeds; a
    workbook that fails import validation exits non-zero for a different reason and
    would make this test vacuous.
    """
    out = tmp_path / "out.jsonld"
    proc = _uofa("import", str(REVIEWED), "--pack", "nasa-7009b",
                 "--protocol-check", "-o", str(out))
    assert proc.returncode != 0, proc.stdout + proc.stderr
    assert "protocol-check" in proc.stdout
    assert "ambiguity log" in proc.stdout


@pytest.mark.skipif(not REVIEWED.exists(), reason="pilot workbook not present")
def test_import_without_the_flag_does_not_gate(tmp_path):
    """The gate is opt-in; the same import succeeds without the flag."""
    out = tmp_path / "out.jsonld"
    proc = _uofa("import", str(REVIEWED), "--pack", "nasa-7009b", "-o", str(out))
    assert proc.returncode == 0, proc.stdout + proc.stderr
    assert "protocol-check" not in proc.stdout


def test_extract_flag_is_accepted_and_documented():
    proc = _uofa("extract", "--help")
    assert proc.returncode == 0
    assert "--protocol-check" in proc.stdout
    assert "Informational" in proc.stdout or "informational" in proc.stdout


def test_the_run_25_reconstruction_is_refused(tmp_path):
    """The permanent regression fixture: the specimen that caught a bad fix.

    Run 25 anchored all seventeen required levels and weighed none. Anchoring
    routes `set-anchor -> confirm`, so every cell exported as `confirmed` -- and
    the first draft of the evidence check counted `confirmed` as judgment and
    PASSED this package. That would have been strictly worse than the shape
    heuristic it replaced: the heuristic refused it correctly.

    Every future change to this check must survive this fixture.
    """
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _mark_levels(wb, ["confirmed", "confirmed", "confirmed"])
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert not result.passed, (
        "anchored-but-never-weighed levels passed -- `confirmed` is being read "
        "as evidence of a judgment that anchoring produces as a side effect")
    assert "3 of 3" in result.detail


def test_a_workbook_declaring_v0_8_with_no_provenance_column_is_refused(tmp_path):
    """The case shape-inference could not see, and the reason for the marker.

    An encoder that declares the profile and omits its column is broken. Read as
    a shape, that workbook is indistinguishable from a legacy one and gets the
    legacy excuse -- a defect certified by the check meant to catch it. Read as
    a declaration, the sheet contradicts itself and says so.
    """
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _declare_profile(wb, "v0.8")
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert not result.passed
    assert "declares encoding profile v0.8" in result.detail
    assert "has none" in result.detail


def test_an_undeclared_workbook_carrying_the_column_says_so_out_loud(tmp_path):
    """Never silent. The column is not read -- an undeclared sheet cannot vouch
    for what it records, and reading it anyway is the sniffing the declaration
    replaces -- but the advisory names it, so the discharge is obvious."""
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _mark_levels(wb, ["affirmed", "affirmed", "affirmed"], declare=None)
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert result.passed and result.skipped
    assert "declares no encoding profile" in result.detail
    assert "re-export" in result.detail


def test_an_unparseable_declaration_does_not_buy_the_evidence_path(tmp_path):
    """A typo must not be read as a version claim this checker understands."""
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3), (4, 4), (2, 2)])
    _mark_levels(wb, ["extracted", "extracted", "extracted"], declare="latest")
    result = _verdict(check_workbook(wb, TEMPLATE), "required levels")
    assert result.passed, "an unreadable declaration was treated as v0.8+"


def _package(tmp_path, context, factors, name="uofa.jsonld"):
    import json as _json
    p = tmp_path / name
    p.write_text(_json.dumps({
        "@context": context, "id": "urn:uofa:t", "type": "UnitOfAssurance",
        "hasCredibilityFactor": factors}), encoding="utf8")
    return p


_V08 = "https://raw.githubusercontent.com/cloudronin/uofa/main/spec/context/v0.8.jsonld"
_V05 = "https://raw.githubusercontent.com/cloudronin/uofa/main/spec/context/v0.5.jsonld"


def test_a_v0_8_package_of_unjudged_levels_is_refused_and_names_them(tmp_path):
    """The run-25 shape at the package layer: every level located, none weighed."""
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V08, [
        {"factorType": "Software quality assurance", "requiredLevel": 3,
         "requiredLevelProvenance": "extracted"},
        {"factorType": "Discretization error", "requiredLevel": 2,
         "requiredLevelProvenance": "defaulted"},
    ])
    r = _check_package_levels(p)
    assert not r.passed
    assert "Software quality assurance" in r.detail
    assert "2 of 2" in r.detail


def test_a_v0_8_package_that_was_judged_passes(tmp_path):
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V08, [
        {"factorType": "A", "requiredLevel": 3, "requiredLevelProvenance": "affirmed"},
        {"factorType": "B", "requiredLevel": 3, "requiredLevelProvenance": "corrected"},
        {"factorType": "C", "requiredLevel": 3, "requiredLevelProvenance": "waived"},
    ])
    r = _check_package_levels(p)
    assert r.passed and not r.skipped, r.detail


def test_a_legacy_package_is_advised_not_refused(tmp_path):
    """Third parties script against these exit codes; age is not negligence."""
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V05, [
        {"factorType": "A", "requiredLevel": 3},
        {"factorType": "B", "requiredLevel": 3},
    ])
    r = _check_package_levels(p)
    assert r.passed and r.skipped
    assert "v0.5.jsonld" in r.detail
    assert "cannot state whether" in r.detail


def test_an_inlined_context_is_advised_rather_than_guessed_at(tmp_path):
    """A resolved or signed document carries its context as an object, so it
    declares no version -- and inferring one from the terms inside would be the
    sniffing this fork exists to replace."""
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, {"uofa": "https://example.org/"},
                 [{"factorType": "A", "requiredLevel": 3,
                   "requiredLevelProvenance": "extracted"}])
    r = _check_package_levels(p)
    assert r.passed and r.skipped


def test_confirmed_does_not_satisfy_the_package_check(tmp_path):
    """The token that fooled the first fix, refused at this layer too."""
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V08, [
        {"factorType": "A", "requiredLevel": 3, "requiredLevelProvenance": "confirmed"},
    ])
    assert not _check_package_levels(p).passed


# ─────────────────────────────────────────────────────────────────────────────
# Dispositions leave the denominator, and say so.
#
# The wall these come from: `mark-not-recoverable` gave a reviewer a lawful exit
# from A-7, and the levels checks still counted every disposed requirement as
# unjudged -- so the app read ALL RESOLVED and the authority refused the same
# package on nineteen counts. A lawful exit that strands the reviewer at the
# next gate is half a fix.
#
# Two rules, and the second is what keeps the first from becoming the defect it
# replaced: a disposal is excluded from the denominator, and the exclusion is
# always named. A denominator that quietly shrinks passes for the same reason a
# vacuous check does.
# ─────────────────────────────────────────────────────────────────────────────

def _workbook_levels(tmp_path, tokens, declare="v0.9"):
    wb = _blank_template(tmp_path)
    _add_anchor_column(wb)
    _set_levels(wb, [(3, 3)] * len(tokens))
    _mark_levels(wb, tokens, declare=declare)
    return _verdict(check_workbook(wb, TEMPLATE), "required levels were reviewed")


def test_a_disposed_requirement_leaves_the_denominator(tmp_path):
    """Two judged, one disposed: the package is not owed a third judgment."""
    r = _workbook_levels(tmp_path, ["affirmed", "corrected", "not-recoverable"])
    assert r.passed and not r.skipped, r.detail
    assert "1 excluded from the denominator" in r.detail, r.detail
    assert "not-recoverable" in r.detail, "the exclusion went unnamed"


def test_the_denominator_reports_the_smaller_number(tmp_path):
    """The count under test, read off the sentence a reader gets.

    `1 of 2`, never `1 of 3`: the disposed requirement is not something anyone
    still owes a judgment on, and reporting it as owed is the accusation this
    deploy exists to stop making.
    """
    r = _workbook_levels(tmp_path, ["affirmed", "extracted", "not-recoverable"])
    assert not r.passed
    assert "1 of 2 required level(s)" in r.detail, r.detail
    assert "1 excluded from the denominator" in r.detail, r.detail


def test_disposal_is_not_judgment_credit(tmp_path):
    """A package that disposed of everything claims no judgment at all.

    The vacuous pass this whole programme is about, arrived at from the new
    direction: exclusion shrinks the denominator, and a denominator of zero
    would render a green tick over a package where nobody weighed anything.
    It renders as an advisory instead, carrying the sentence.
    """
    r = _workbook_levels(tmp_path, ["not-recoverable", "source-absent"])
    assert r.skipped, "an all-disposed workbook rendered as a judgment pass"
    assert "no required level was judged" in r.detail, r.detail


def test_a_v0_8_workbook_carrying_a_v0_9_term_is_refused_by_name(tmp_path):
    """The declaration and the sheet disagree, and the refusal says which term.

    `not-recoverable` entered at v0.9. A v0.8 sheet writing it is not a v0.8
    sheet, and the two available alternatives are both worse: accept it (the
    declaration stops constraining anything) or report it as unjudged (a false
    accusation against the one reviewer who did the honest thing).
    """
    r = _workbook_levels(tmp_path, ["not-recoverable"], declare="v0.8")
    assert not r.passed and not r.skipped
    assert "not-recoverable" in r.detail and "v0.8" in r.detail, r.detail


_V09 = "https://raw.githubusercontent.com/cloudronin/uofa/main/spec/context/v0.9.jsonld"


def test_the_package_layer_excludes_disposals_too(tmp_path):
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V09, [
        {"factorType": "A", "requiredLevel": 3, "requiredLevelProvenance": "affirmed"},
        {"factorType": "B", "requiredLevel": 3,
         "requiredLevelProvenance": "not-recoverable"},
    ])
    r = _check_package_levels(p)
    assert r.passed and not r.skipped, r.detail
    assert "all 1 required level(s) carry a judgment" in r.detail, r.detail
    assert "1 not-recoverable" in r.detail, "the exclusion went unnamed"


def test_the_package_layer_refuses_a_v0_9_term_under_a_v0_8_context(tmp_path):
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V08, [
        {"factorType": "A", "requiredLevel": 3,
         "requiredLevelProvenance": "not-recoverable"},
    ])
    r = _check_package_levels(p)
    assert not r.passed and not r.skipped
    assert "not-recoverable" in r.detail and "v0.8" in r.detail, r.detail


def test_an_all_disposed_package_claims_no_judgment(tmp_path):
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V09, [
        {"factorType": "A", "requiredLevel": 3,
         "requiredLevelProvenance": "source-absent"},
        {"factorType": "B", "requiredLevel": 3,
         "requiredLevelProvenance": "not-recoverable"},
    ])
    r = _check_package_levels(p)
    assert r.skipped, "an all-disposed package rendered as a judgment pass"
    assert "no required level carries a judgment" in r.detail, r.detail


def test_source_absent_was_always_a_disposition_and_is_now_read_as_one(tmp_path):
    """The bug predates `not-recoverable` and is fixed by the same rule.

    `source-absent` has been in v0.8's vocabulary since it shipped, and both
    checks counted it as a missing judgment -- so a reviewer who recorded that
    the document does not state a requirement was told they had not weighed it.
    Nothing about that needed a new term to be wrong.
    """
    from uofa_cli.protocol_check import _check_package_levels
    p = _package(tmp_path, _V08, [
        {"factorType": "A", "requiredLevel": 3, "requiredLevelProvenance": "affirmed"},
        {"factorType": "B", "requiredLevel": 3,
         "requiredLevelProvenance": "source-absent"},
    ])
    r = _check_package_levels(p)
    assert r.passed and not r.skipped, r.detail
