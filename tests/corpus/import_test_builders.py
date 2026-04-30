"""Build the 8 SHACL-boundary xlsx templates under corpus/import-tests/.

Every template is derived deterministically from the clean-base helpers in
`tests/fixtures/import/generator.py` via targeted mutations, so the expected
SHACL outcomes stay coupled to `excel_constants.py` (factor names, level
ranges) without duplication.

Expected outcomes (enforced in tests/test_corpus_import.py):

    File                          Import  SHACL
    perfect-vv40.xlsx             pass    pass
    perfect-nasa.xlsx             pass    pass
    minimal-7-factors.xlsx        pass    pass
    missing-required-fields.xlsx  fail    --    (import errors before SHACL runs)
    level-out-of-range.xlsx       pass    fail
    typo-in-factor-name.xlsx      fail    --    (import rejects unknown factor)
    empty-rows.xlsx               pass    pass
    extra-columns.xlsx            pass    pass
"""

from __future__ import annotations

import copy
from pathlib import Path

# generator.py sits under tests/fixtures/import/; build.py prepends that to
# sys.path before importing this module.
from generator import (  # type: ignore[import-not-found]
    _clean_base_nasa,
    _clean_base_vv40,
    generate_fixture,
)


def _build_perfect_vv40(out_dir: Path) -> Path:
    path = out_dir / "perfect-vv40.xlsx"
    generate_fixture(_clean_base_vv40(), path)
    return path


def _build_perfect_nasa(out_dir: Path) -> Path:
    path = out_dir / "perfect-nasa.xlsx"
    generate_fixture(_clean_base_nasa(), path)
    return path


def _build_minimal_7_factors(out_dir: Path) -> Path:
    """Minimal profile: keep the first 7 factors, drop the rest, mark Minimal."""
    spec = _clean_base_vv40()
    spec["summary"]["profile"] = "Minimal"
    spec["factors"] = spec["factors"][:7]
    path = out_dir / "minimal-7-factors.xlsx"
    generate_fixture(spec, path)
    return path


def _build_missing_required_fields(out_dir: Path) -> Path:
    """Drop the Decision sheet entirely. excel_reader must surface a named
    'Sheet \\'Decision\\' not found' error — tester can act on it."""
    spec = _clean_base_vv40()
    del spec["decision"]
    path = out_dir / "missing-required-fields.xlsx"
    generate_fixture(spec, path)
    return path


def _build_level_out_of_range(out_dir: Path) -> Path:
    """Push one factor's required_level to 9 (VV40 valid range is 1–5).
    Import will pass structurally; SHACL reports the range violation."""
    spec = _clean_base_vv40()
    spec["factors"][0]["required_level"] = 9
    path = out_dir / "level-out-of-range.xlsx"
    generate_fixture(spec, path)
    return path


def _build_typo_in_factor_name(out_dir: Path) -> Path:
    """Rename 'Model form' → 'Model Form Accuracy'. excel_constants enforces
    the canonical factor-name list at import time, so the error surfaces
    before SHACL."""
    spec = _clean_base_vv40()
    # Find the "Model form" factor row and rename.
    for factor in spec["factors"]:
        if factor["factor_type"] == "Model form":
            factor["factor_type"] = "Model Form Accuracy"
            break
    path = out_dir / "typo-in-factor-name.xlsx"
    generate_fixture(spec, path)
    return path


def _build_empty_rows(out_dir: Path) -> Path:
    """Inject 3 blank rows into the Credibility Factors sheet after the
    first factor row. excel_reader must skip blanks silently."""
    spec = _clean_base_vv40()
    path = out_dir / "empty-rows.xlsx"
    generate_fixture(spec, path)
    _inject_blank_rows(path, "Credibility Factors", after_row=5, count=3)
    return path


def _build_extra_columns(out_dir: Path) -> Path:
    """Append a 'Notes' column to the Credibility Factors sheet. excel_reader
    reads by column label, so extra columns must be ignored."""
    spec = _clean_base_vv40()
    path = out_dir / "extra-columns.xlsx"
    generate_fixture(spec, path)
    _append_notes_column(path, "Credibility Factors")
    return path


# ── xlsx post-processing helpers ────────────────────────────────


def _inject_blank_rows(xlsx: Path, sheet: str, *, after_row: int, count: int) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    ws.insert_rows(after_row + 1, amount=count)
    wb.save(xlsx)


def _append_notes_column(xlsx: Path, sheet: str) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    new_col = ws.max_column + 1
    # Column header sits on the same row as the other factor headers (row 3).
    ws.cell(row=3, column=new_col, value="Notes")
    # Fill a couple of data rows so the column isn't empty.
    for r in range(5, min(ws.max_row, 8) + 1):
        ws.cell(row=r, column=new_col, value="extra note")
    wb.save(xlsx)


def build_all(out_dir: Path) -> list[Path]:
    return [
        _build_perfect_vv40(out_dir),
        _build_perfect_nasa(out_dir),
        _build_minimal_7_factors(out_dir),
        _build_missing_required_fields(out_dir),
        _build_level_out_of_range(out_dir),
        _build_typo_in_factor_name(out_dir),
        _build_empty_rows(out_dir),
        _build_extra_columns(out_dir),
    ]
