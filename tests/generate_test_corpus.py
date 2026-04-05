#!/usr/bin/env python3
"""Generate Excel test corpus for uofa import testing.

Usage: python tests/generate_test_corpus.py [--output-dir tests/fixtures/import/]

Generates TC-XX-*.xlsx files and a tc_manifest.json companion.
Each file tests a specific scenario — happy path, edge case, or expected error.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Protection
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Add src to path so we can import constants
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from uofa_cli.excel_constants import (
    VV40_FACTOR_CATEGORIES, NASA_ONLY_FACTOR_CATEGORIES, ALL_FACTOR_CATEGORIES,
    VV40_FACTOR_NAMES, NASA_ALL_FACTOR_NAMES, EVIDENCE_TYPES,
)


# ── Sheet builders ────────────────────────────────────────────


def _add_summary_sheet(wb, *, project_name="Test Project", cou_name="Test COU",
                       cou_description=None, profile="Complete", device_class="Class II",
                       mrl="MRL 3", assurance_level="Medium", standards_ref="ASME-VV40-2018",
                       assessor_name="Test Assessor", assessment_date="2026-04-15",
                       source_doc="https://example.org/report", has_uq="Yes"):
    """Add Assessment Summary sheet with header + instruction + data rows."""
    ws = wb.create_sheet("Assessment Summary", 0)
    ws.append(["Unit of Assurance — Assessment Summary"])
    ws.append(["Project Name", "COU Name", "COU Description", "Profile",
               "Device Class", "Model Risk Level", "Assurance Level",
               "Standards Reference", "Assessor Name", "Assessment Date",
               "Source Document", "Has UQ?"])
    ws.append([
        project_name, cou_name, cou_description, profile,
        device_class, mrl, assurance_level, standards_ref,
        assessor_name, assessment_date, source_doc, has_uq,
    ])
    return ws


def _add_model_data_sheet(wb, entities=None):
    """Add Model & Data sheet."""
    ws = wb.create_sheet("Model & Data")
    ws.append(["Model & Data"])
    ws.append(["Entity Type", "Name", "Identifier/URI", "Description", "Version", "Source"])
    if entities is None:
        entities = [
            ("Requirement", "Safety requirement", "https://example.org/req/1", "Test req", None, None),
            ("Model", "Simulation model", "https://example.org/model/1", "Test model", "v1.0", None),
            ("Dataset", "Test dataset", "https://example.org/data/1", "Test data", None, None),
        ]
    for e in entities:
        ws.append(list(e))
    return ws


def _add_validation_results_sheet(wb, results=None, include_type_col=True):
    """Add Validation Results sheet.

    If include_type_col=True, uses v2 layout with Type column.
    If False, uses old layout without Type column.
    """
    ws = wb.create_sheet("Validation Results")
    ws.append(["Validation Results"])
    if include_type_col:
        ws.append(["Result Name", "Type", "Identifier/URI", "Description",
                    "Compares To", "Has UQ?", "UQ Method", "Metric Value", "Pass/Fail"])
    else:
        ws.append(["Result Name", "Identifier/URI", "Description",
                    "Compares To", "Has UQ?", "UQ Method", "Metric Value", "Pass/Fail"])
    if results is None:
        results = [
            _vr_row("Mesh convergence", "ValidationResult", has_uq="Yes",
                     uq_method="GCI", metric="1.8%", pass_fail="Pass",
                     compares_to="https://example.org/data/1",
                     include_type=include_type_col),
        ]
    for r in results:
        ws.append(r)
    return ws


def _vr_row(name, evidence_type="ValidationResult", uri=None, desc=None,
            compares_to=None, has_uq="Yes", uq_method=None, metric=None,
            pass_fail="Pass", include_type=True):
    """Build a validation result row."""
    uri = uri or f"https://example.org/validation/{name.lower().replace(' ', '-')}"
    if include_type:
        return [name, evidence_type, uri, desc, compares_to, has_uq, uq_method, metric, pass_fail]
    else:
        return [name, uri, desc, compares_to, has_uq, uq_method, metric, pass_fail]


def _add_factors_sheet(wb, factors=None, factor_list=None):
    """Add Credibility Factors sheet.

    factors: list of (required_level, achieved_level, acceptance, rationale, status) tuples.
    factor_list: which factor (name, category) pairs to use (defaults to VV40 13).
    """
    ws = wb.create_sheet("Credibility Factors")
    ws.append(["Credibility Factors"])
    ws.append([""])
    ws.append(["Factor Type", "Category", "Required Level", "Achieved Level",
               "Acceptance Criteria", "Rationale", "Factor Status"])
    ws.append(["(pre-filled)", "(pre-filled)", "(1-5)", "(1-5)", "(optional)", "(optional)", "(status)"])

    if factor_list is None:
        factor_list = VV40_FACTOR_CATEGORIES

    if factors is None:
        # Default: all assessed at level 3
        factors = [(3, 3, None, None, "assessed")] * len(factor_list)

    for (fname, fcat), (req, ach, acc, rat, status) in zip(factor_list, factors):
        ws.append([fname, fcat, req, ach, acc, rat, status])

    return ws


def _add_decision_sheet(wb, outcome="Accepted", rationale="Test rationale",
                        criteria_set=None, decided_by="Test Board", decision_date="2026-04-15"):
    """Add Decision sheet."""
    ws = wb.create_sheet("Decision")
    ws.append(["Decision"])
    ws.append(["Decision Outcome", "Decision Rationale", "Criteria Set",
               "Decided By", "Decision Date"])
    ws.append([outcome, rationale, criteria_set, decided_by, decision_date])
    return ws


def _add_lists_sheet(wb):
    """Add hidden _Lists sheet."""
    ws = wb.create_sheet("_Lists")
    ws.sheet_state = "hidden"
    return ws


# ── Test case generators ──────────────────────────────────────


def create_tc01(output_dir):
    """TC-01: Minimal profile, VV40. Only required Minimal fields."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC01 Minimal", cou_name="Minimal COU",
                       profile="Minimal", device_class=None, mrl=None,
                       assurance_level=None, standards_ref=None,
                       cou_description=None, source_doc=None, has_uq=None)
    _add_model_data_sheet(wb, entities=[
        ("Requirement", "Safety req", "https://example.org/req/1", None, None, None),
    ])
    _add_validation_results_sheet(wb, results=[
        _vr_row("Basic validation", has_uq="No"),
    ])
    _add_decision_sheet(wb, outcome="Accepted", rationale="Minimal acceptance")
    _add_lists_sheet(wb)

    path = output_dir / "tc01-vv40-minimal.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Minimal",
        "factor_count": 0,
        "validation_result_count": 1,
        "weakeners": [],
        "assertions": ["has_provenance", "has_context_v04", "has_valid_uri"],
    }


def create_tc02(output_dir):
    """TC-02: Complete profile, all 13 factors assessed, required >= achieved."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC02 Complete", cou_name="All Assessed COU")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Mesh convergence", has_uq="Yes", uq_method="GCI",
                compares_to="https://example.org/data/1"),
        _vr_row("Strain comparison", has_uq="Yes", uq_method="Monte Carlo",
                compares_to="https://example.org/data/2"),
        _vr_row("Fatigue life", has_uq="Yes", uq_method="Bayesian"),
    ])
    # All 13 factors assessed at level 3
    factors = [(3, 3, "Criteria text", "Rationale text", "assessed")] * 13
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb, outcome="Accepted", rationale="All factors meet goals")
    _add_lists_sheet(wb)

    path = output_dir / "tc02-vv40-complete-all-assessed.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 13,
        "validation_result_count": 3,
        "weakeners": [],
        "factor_standards": {"1-13": "ASME-VV40-2018"},
        "assertions": ["has_provenance", "has_context_v04", "zero_weakeners"],
    }


def create_tc03(output_dir):
    """TC-03: Complete profile, mixed statuses. 7 assessed, 3 not-assessed, 2 scoped-out, 1 n/a."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC03 Mixed", cou_name="Mixed Status COU")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Result with UQ", has_uq="Yes", uq_method="GCI",
                compares_to="https://example.org/data/1"),
        _vr_row("Result no UQ", has_uq="No"),
    ])
    factors = [
        (3, 3, None, None, "assessed"),      # 1
        (3, 3, None, None, "assessed"),      # 2
        (3, 3, None, None, "assessed"),      # 3
        (3, 3, None, None, "assessed"),      # 4
        (3, 3, None, None, "assessed"),      # 5
        (3, 3, None, None, "assessed"),      # 6
        (3, 3, None, None, "assessed"),      # 7
        (None, None, None, None, "not-assessed"),  # 8
        (None, None, None, None, "not-assessed"),  # 9
        (None, None, None, None, "not-assessed"),  # 10
        (None, None, None, None, "scoped-out"),    # 11
        (None, None, None, None, "scoped-out"),    # 12
        (None, None, None, None, "not-applicable"), # 13
    ]
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb, outcome="Conditional", rationale="Mixed status")
    _add_lists_sheet(wb)

    path = output_dir / "tc03-vv40-complete-mixed-status.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 7,
        "validation_result_count": 2,
        "assertions": ["has_provenance"],
    }


def create_tc04(output_dir):
    """TC-04: Complete profile, intentional gaps for weakener triggers."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC04 Gaps", cou_name="Weakener Gaps COU")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Result no UQ", has_uq="No"),  # W-AL-01
        _vr_row("Result no compare", has_uq="Yes", uq_method="GCI"),  # no comparesTo → W-EP-01
    ])
    factors = [
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
        (4, 2, None, None, "assessed"),  # achieved < required → W-AR-01
        (3, 3, None, None, "assessed"),
        (3, 3, None, None, "assessed"),
    ]
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb, outcome="Not accepted", rationale="Gaps exist")
    _add_lists_sheet(wb)

    path = output_dir / "tc04-vv40-complete-gaps.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 13,
        "validation_result_count": 2,
        "weakener_assertions": ["W-AL-01", "W-EP-01", "W-AR-01"],
        "assertions": ["has_provenance"],
    }


def create_tc06(output_dir):
    """TC-06: NASA Complete, all 19 factors assessed."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC06 NASA", cou_name="NASA Full COU",
                       standards_ref="NASA-STD-7009B")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Validation 1", has_uq="Yes", uq_method="Monte Carlo",
                compares_to="https://example.org/data/1"),
        _vr_row("Tech review", "ReviewActivity",
                compares_to="https://example.org/org/reviewer-1"),
        _vr_row("Process check", "ProcessAttestation",
                compares_to="https://example.org/org/qa-team"),
        _vr_row("Deploy record", "DeploymentRecord",
                compares_to="https://example.org/system/flight-test-1"),
        _vr_row("Data source", "InputPedigreeLink",
                compares_to="https://example.org/data/arc-jet-facility"),
    ])
    factors = [(3, 3, None, None, "assessed")] * 19
    _add_factors_sheet(wb, factors=factors, factor_list=ALL_FACTOR_CATEGORIES)
    _add_decision_sheet(wb, outcome="Accepted", rationale="All 19 NASA factors assessed")
    _add_lists_sheet(wb)

    path = output_dir / "tc06-nasa-complete-19-factors.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["nasa-7009b"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 19,
        "validation_result_count": 5,
        "assertions": ["has_provenance", "has_context_v04", "nasa_factor_standard"],
    }


def create_tc09(output_dir):
    """TC-09: Minimal profile with nasa-7009b pack."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC09 NASA Min", cou_name="NASA Minimal COU",
                       profile="Minimal", device_class=None, mrl=None,
                       assurance_level=None)
    _add_model_data_sheet(wb, entities=[
        ("Requirement", "Req 1", "https://example.org/req/1", None, None, None),
    ])
    _add_validation_results_sheet(wb, results=[
        _vr_row("Validation", has_uq="No"),
    ])
    _add_decision_sheet(wb, outcome="Accepted")
    _add_lists_sheet(wb)

    path = output_dir / "tc09-nasa-minimal.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["nasa-7009b"],
        "expect": "pass",
        "profile": "Minimal",
        "factor_count": 0,
        "validation_result_count": 1,
        "assertions": ["has_provenance"],
    }


def create_tc19(output_dir):
    """TC-19: Type column absent (old template format)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="TC19 No Type Col", cou_name="Old Format COU")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Result 1", include_type=False, has_uq="Yes", uq_method="GCI",
                compares_to="https://example.org/data/1"),
    ], include_type_col=False)
    factors = [(3, 3, None, None, "assessed")] * 13
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc19-type-column-absent.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 13,
        "validation_result_count": 1,
        "assertions": ["has_provenance", "all_validation_result_type"],
    }


# ── Error case generators ────────────────────────────────────


def create_tc30(output_dir):
    """TC-30: Missing Assessment Summary sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Skip Assessment Summary
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb)
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc30-missing-assessment-summary.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "Sheet 'Assessment Summary' not found",
    }


def create_tc31(output_dir):
    """TC-31: Missing Decision sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb)
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb)
    # Skip Decision
    _add_lists_sheet(wb)

    path = output_dir / "tc31-missing-decision-sheet.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "Sheet 'Decision' not found",
    }


def create_tc35(output_dir):
    """TC-35: Invalid decision value 'Approved'."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, profile="Minimal")
    _add_model_data_sheet(wb, entities=[
        ("Requirement", "Req 1", "https://example.org/req/1", None, None, None),
    ])
    _add_validation_results_sheet(wb)
    _add_decision_sheet(wb, outcome="Approved")
    _add_lists_sheet(wb)

    path = output_dir / "tc35-invalid-decision-value.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "'Approved' is not a valid outcome",
    }


def create_tc37(output_dir):
    """TC-37: Invalid factor type 'Mesh quality'."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb)
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb)
    # Create factors sheet with an invalid factor
    ws = wb.create_sheet("Credibility Factors")
    ws.append(["Credibility Factors"])
    ws.append([""])
    ws.append(["Factor Type", "Category", "Required Level", "Achieved Level",
               "Acceptance Criteria", "Rationale", "Factor Status"])
    ws.append(["(pre-filled)", "(pre-filled)", "(1-5)", "(1-5)", "", "", "(status)"])
    ws.append(["Mesh quality", "Custom", 3, 3, None, None, "assessed"])
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc37-invalid-factor-type.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "'Mesh quality' is not a valid V&V 40 factor type",
    }


def create_tc38(output_dir):
    """TC-38: No requirement rows in Model & Data."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, profile="Minimal")
    _add_model_data_sheet(wb, entities=[
        ("Model", "My model", "https://example.org/model/1", None, None, None),
        ("Dataset", "My data", "https://example.org/data/1", None, None, None),
    ])
    _add_validation_results_sheet(wb)
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc38-no-requirement-rows.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "must have at least one row with Entity Type = 'Requirement'",
    }


def create_tc39(output_dir):
    """TC-39: Invalid evidence type 'CustomType'."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb)
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb, results=[
        _vr_row("Bad evidence", "CustomType"),
    ])
    factors = [(3, 3, None, None, "assessed")] * 13
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc39-invalid-evidence-type.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["nasa-7009b"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "'CustomType' is not a valid evidence type",
    }


def create_tc41(output_dir):
    """TC-41: Not a valid xlsx file (text file with .xlsx extension)."""
    path = output_dir / "tc41-not-xlsx.xlsx"
    path.write_text("This is not an Excel file.\n")
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "error",
        "exit_code": 1,
        "error_contains": "Cannot open workbook",
    }


def create_tc62(output_dir):
    """TC-62: URI generation from project/COU names."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _add_summary_sheet(wb, project_name="My Test Project",
                       cou_name="Thermal Analysis COU")
    _add_model_data_sheet(wb)
    _add_validation_results_sheet(wb)
    factors = [(3, 3, None, None, "assessed")] * 13
    _add_factors_sheet(wb, factors=factors)
    _add_decision_sheet(wb)
    _add_lists_sheet(wb)

    path = output_dir / "tc62-uri-generation.xlsx"
    wb.save(path)
    return {
        "file": path.name,
        "pack": ["vv40"],
        "expect": "pass",
        "profile": "Complete",
        "factor_count": 13,
        "validation_result_count": 1,
        "expected_id": "https://uofa.net/instances/my-test-project/thermal-analysis-cou",
        "assertions": ["has_provenance", "uri_slugification"],
    }


# ── Main ──────────────────────────────────────────────────────


ALL_GENERATORS = {
    "TC-01": create_tc01,
    "TC-02": create_tc02,
    "TC-03": create_tc03,
    "TC-04": create_tc04,
    "TC-06": create_tc06,
    "TC-09": create_tc09,
    "TC-19": create_tc19,
    "TC-30": create_tc30,
    "TC-31": create_tc31,
    "TC-35": create_tc35,
    "TC-37": create_tc37,
    "TC-38": create_tc38,
    "TC-39": create_tc39,
    "TC-41": create_tc41,
    "TC-62": create_tc62,
}


def main():
    parser = argparse.ArgumentParser(description="Generate test corpus for uofa import")
    parser.add_argument("--output-dir", type=Path, default=Path("tests/fixtures/import"),
                        help="output directory (default: tests/fixtures/import)")
    args = parser.parse_args()

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest = {}
    for tc_id, gen_fn in sorted(ALL_GENERATORS.items()):
        entry = gen_fn(output_dir)
        manifest[tc_id] = entry
        print(f"  {tc_id}: {entry['file']}")

    manifest_path = output_dir / "tc_manifest.json"
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)
        f.write("\n")

    print(f"\nGenerated {len(manifest)} test files + manifest in {output_dir}")


if __name__ == "__main__":
    main()
