"""Integration tests for uofa extract — full CLI pipeline with mock provider."""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

REPO_ROOT = Path(__file__).parent.parent
MORRISON_DIR = Path(__file__).parent / "fixtures" / "extract" / "morrison-evidence"
AERO_COU1_DIR = Path(__file__).parent / "fixtures" / "extract" / "aero-evidence-cou1"
AERO_COU2_DIR = Path(__file__).parent / "fixtures" / "extract" / "aero-evidence-cou2"


def run_uofa(*args, cwd=None):
    """Run uofa CLI as subprocess and return CompletedProcess."""
    cmd = [sys.executable, "-m", "uofa_cli", *args]
    env = {"PYTHONPATH": str(REPO_ROOT / "src")}

    import os
    full_env = os.environ.copy()
    full_env.update(env)

    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        cwd=cwd or str(REPO_ROOT),
        env=full_env,
    )


pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")


class TestExtractHelp:
    def test_extract_in_help(self):
        result = run_uofa("--help")
        assert "extract" in result.stdout


class TestExtractNoSource:
    def test_no_source_error(self, tmp_path):
        result = run_uofa("extract", cwd=str(tmp_path))
        assert result.returncode == 1


class TestExtractEmptyFolder:
    def test_empty_folder(self, tmp_path):
        empty = tmp_path / "empty"
        empty.mkdir()
        result = run_uofa("extract", str(empty), "--model", "mock")
        assert result.returncode == 1
        assert "No supported files" in result.stderr or "No supported files" in result.stdout


@pytest.mark.skipif(not MORRISON_DIR.exists(), reason="Morrison evidence not available")
class TestExtractMorrisonMock:
    """Full pipeline test: Morrison evidence → mock LLM → Excel output."""

    def test_extract_succeeds(self, tmp_path):
        out = tmp_path / "output.xlsx"
        result = run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
        )
        # Print output for debugging
        if result.returncode != 0:
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
        assert result.returncode == 0

    def test_output_exists(self, tmp_path):
        out = tmp_path / "output.xlsx"
        run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
        )
        assert out.exists()

    def test_output_has_factors(self, tmp_path):
        out = tmp_path / "output.xlsx"
        run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
        )
        from uofa_cli.excel_constants import VV40_FACTOR_NAMES
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        factors_found = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in VV40_FACTOR_NAMES:
                factors_found.append(val)
        assert len(factors_found) == 13

    def test_output_has_decision(self, tmp_path):
        out = tmp_path / "output.xlsx"
        run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Decision"]
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Accepted":
                found = True
        assert found

    def test_glob_filter(self, tmp_path):
        out = tmp_path / "output.xlsx"
        result = run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
            "--glob", "*.csv,*.txt",
        )
        assert result.returncode == 0
        # Output should mention only csv/txt files
        assert out.exists()

    def test_verbose_output(self, tmp_path):
        out = tmp_path / "output.xlsx"
        result = run_uofa(
            "extract", str(MORRISON_DIR),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
            "--verbose",
        )
        assert result.returncode == 0
        assert "tokens" in result.stdout.lower()


def _run_aero_extract(tmp_path, evidence_dir):
    """Helper: run uofa extract for an aero NASA COU with the mock model."""
    out = tmp_path / "output.xlsx"
    result = run_uofa(
        "extract", str(evidence_dir),
        "--model", "mock",
        "--pack", "nasa-7009b",
        "--output", str(out),
    )
    if result.returncode != 0:
        print("STDOUT:", result.stdout)
        print("STDERR:", result.stderr)
    return result, out


@pytest.mark.skipif(not AERO_COU1_DIR.exists(), reason="Aero COU1 evidence not available")
class TestExtractAeroCou1Mock:
    """Full pipeline test: aero-evidence-cou1 → mock LLM → Excel (19 NASA factors)."""

    def test_extract_succeeds(self, tmp_path):
        result, _ = _run_aero_extract(tmp_path, AERO_COU1_DIR)
        assert result.returncode == 0

    def test_output_exists(self, tmp_path):
        _, out = _run_aero_extract(tmp_path, AERO_COU1_DIR)
        assert out.exists()

    def test_output_has_19_factors(self, tmp_path):
        _, out = _run_aero_extract(tmp_path, AERO_COU1_DIR)
        from uofa_cli.excel_constants import NASA_ALL_FACTOR_NAMES
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        factors_found = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in NASA_ALL_FACTOR_NAMES:
                factors_found.append(val)
        assert len(factors_found) == 19


@pytest.mark.skipif(not AERO_COU2_DIR.exists(), reason="Aero COU2 evidence not available")
class TestExtractAeroCou2Mock:
    """Full pipeline test: aero-evidence-cou2 → mock LLM → Excel (19 NASA factors)."""

    def test_extract_succeeds(self, tmp_path):
        result, _ = _run_aero_extract(tmp_path, AERO_COU2_DIR)
        assert result.returncode == 0

    def test_output_exists(self, tmp_path):
        _, out = _run_aero_extract(tmp_path, AERO_COU2_DIR)
        assert out.exists()

    def test_output_has_19_factors(self, tmp_path):
        _, out = _run_aero_extract(tmp_path, AERO_COU2_DIR)
        from uofa_cli.excel_constants import NASA_ALL_FACTOR_NAMES
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        factors_found = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in NASA_ALL_FACTOR_NAMES:
                factors_found.append(val)
        assert len(factors_found) == 19


import shutil as _shutil


AERO_COU1_IMPORTED = Path(__file__).parent / "fixtures" / "extract" / "aero-cou1-imported.jsonld"
AERO_COU2_IMPORTED = Path(__file__).parent / "fixtures" / "extract" / "aero-cou2-imported.jsonld"


def _count_fires(reasoned_jsonld: Path) -> dict:
    """Parse a reasoned jsonld and return pattern_id -> count."""
    import json
    from collections import Counter
    counts: Counter = Counter()
    data = json.loads(reasoned_jsonld.read_text())
    for node in data.get("@graph", []):
        pid = node.get("https://uofa.net/vocab#patternId")
        if pid:
            if isinstance(pid, dict):
                pid = pid.get("@value", "")
            counts[str(pid)] += 1
    return dict(counts)


@pytest.mark.skipif(
    not AERO_COU1_IMPORTED.exists() or _shutil.which("java") is None,
    reason="Aero COU1 fixture or Java missing",
)
class TestAeroWeakenerPipelineFromFixture:
    """Hard gates on the hand-crafted imported fixtures.

    These tests isolate C3 correctness from LLM+import non-determinism. They
    prove that the rule engine produces the NAFEMS divergence pattern on a
    known input:
      COU1 (Accepted)     -> W-AR-02 fires on every level gap
      COU2 (Not Accepted) -> W-AR-02 stays at zero regardless of gaps
    """

    def _reason(self, tmp_path, fixture: Path) -> dict:
        reasoned = tmp_path / "reasoned.jsonld"
        result = run_uofa(
            "rules", str(fixture),
            "--pack", "nasa-7009b",
            "--format", "jsonld",
            "-o", str(reasoned),
            "--build",
        )
        assert result.returncode == 0, f"rules failed: {result.stderr}"
        assert reasoned.exists()
        return _count_fires(reasoned)

    def test_cou1_w_ar_02_fires_on_gap_factors(self, tmp_path):
        counts = self._reason(tmp_path, AERO_COU1_IMPORTED)
        # Fixture has 3 level-gap factors under Accepted -> W-AR-02 x 3
        assert counts.get("W-AR-02", 0) == 3, counts

    def test_cou1_w_ep_04_fires_on_not_assessed(self, tmp_path):
        counts = self._reason(tmp_path, AERO_COU1_IMPORTED)
        # Results uncertainty is not-assessed at MRL 3 -> W-EP-04 x 1
        assert counts.get("W-EP-04", 0) == 1, counts

    def test_cou1_compound_escalation_fires(self, tmp_path):
        counts = self._reason(tmp_path, AERO_COU1_IMPORTED)
        assert counts.get("COMPOUND-01", 0) >= 1, counts

    def test_cou2_w_ep_04_fires_five_times(self, tmp_path):
        counts = self._reason(tmp_path, AERO_COU2_IMPORTED)
        # Five not-assessed factors at MRL 4 -> W-EP-04 x 5
        assert counts.get("W-EP-04", 0) == 5, counts

    def test_cou2_w_ar_02_does_not_fire(self, tmp_path):
        """Hard gate: Not Accepted -> W-AR-02 must stay at zero.

        This is the Morrison-COU2 parity assertion and the NAFEMS divergence
        headline. If this fires, either the decision outcome is being parsed
        as 'Accepted' or the rule engine is matching on a different property.
        """
        counts = self._reason(tmp_path, AERO_COU2_IMPORTED)
        assert counts.get("W-AR-02", 0) == 0, counts

    def test_cou1_no_mass_w_ar_01(self, tmp_path):
        """Fixture populates acceptanceCriteria on every assessed factor, so
        W-AR-01 should not mass-fire. If this breaks, the fixture generator
        has drifted — regenerate with tests/fixtures/extract/_build_aero_fixtures.py.
        """
        counts = self._reason(tmp_path, AERO_COU1_IMPORTED)
        assert counts.get("W-AR-01", 0) == 0, counts


class TestExtractTextOnly:
    """Test extraction with only text/csv files (no pdfplumber/python-docx needed)."""

    def test_extract_text_csv(self, tmp_path):
        # Create simple evidence files
        evidence = tmp_path / "evidence"
        evidence.mkdir()
        (evidence / "report.txt").write_text(
            "V&V Report\n\nThe CFD model was validated against experimental data.\n"
            "Software: ANSYS CFX v21.0 with ISO 9001 certification.\n"
            "Mesh convergence study: 3 levels, GCI = 1.2%.\n"
        )
        (evidence / "data.csv").write_text(
            "mesh_level,elements,gci_pct\nCoarse,100000,5.2\nMedium,400000,1.8\nFine,1600000,0.6\n"
        )

        out = tmp_path / "result.xlsx"
        result = run_uofa(
            "extract", str(evidence),
            "--model", "mock",
            "--pack", "vv40",
            "--output", str(out),
        )
        assert result.returncode == 0
        assert out.exists()
