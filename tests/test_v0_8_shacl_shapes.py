"""SHACL for v0.8: the closed vocabulary, and judgment claims naming their agent.

**These shapes target `uofa:CredibilityFactor` directly rather than hanging off
`uofa:CredibilityFactorShape`, and that is the whole reason they work.** That
shape is referenced from exactly one place in the core file -- the
ProfileComplete body -- so a package declaring ProfileMinimal never reaches it
and none of its constraints run. `test_the_factor_shape_is_unreachable...`
below records that directly: a bogus `factorStatus` validates clean.

The encoding tool emits ProfileMinimal. A judgment constraint placed in the
usual spot would therefore have been enforced on the packages that do not need
checking and skipped on every package this product actually produces -- a gate
correct on the ordinary case and vacuous on the important one, which is the
defect family this whole version exists to close.
"""

from __future__ import annotations

import copy
import json
import shutil
from pathlib import Path

import openpyxl
import pytest
from rdflib import Graph

from uofa_cli import excel_mapper, excel_reader, paths
from uofa_cli.excel_constants import (
    LEVEL_AFFIRMED_AT_HEADER, LEVEL_AFFIRMED_BY_HEADER, LEVEL_PROVENANCE_HEADER,
)
from uofa_cli.shacl_friendly import _load_data_graph

REPO = Path(__file__).resolve().parents[1]
REVIEWED = REPO / "dev" / "build" / "pilot-johnson" / "johnson-extracted.xlsx"
PACKS = ["nasa-7009b"]


@pytest.fixture(scope="module")
def shapes() -> Graph:
    g = Graph()
    for p in paths.all_shacl_schemas():
        g.parse(str(p), format="turtle")
    return g


@pytest.fixture(scope="module")
def attributed(tmp_path_factory) -> dict:
    """A real package carrying one fully attributed affirmation."""
    if not REVIEWED.exists():
        pytest.skip(f"reviewed fixture not present: {REVIEWED}")
    d = tmp_path_factory.mktemp("wb")
    wb = d / "w.xlsx"
    shutil.copy2(REVIEWED, wb)
    book = openpyxl.load_workbook(wb)
    ws = book["Credibility Factors"]
    c = ws.max_column + 1
    ws.cell(row=3, column=c).value = LEVEL_PROVENANCE_HEADER
    ws.cell(row=3, column=c + 1).value = LEVEL_AFFIRMED_BY_HEADER
    ws.cell(row=3, column=c + 2).value = LEVEL_AFFIRMED_AT_HEADER
    ws.cell(row=5, column=c).value = "affirmed"
    ws.cell(row=5, column=c + 1).value = "V. Vettrivel"
    ws.cell(row=5, column=c + 2).value = "2026-08-24T10:00:00Z"
    ws.cell(row=5, column=3).value = 3
    ws.cell(row=5, column=4).value = 3
    book.save(wb)
    return excel_mapper.map_to_jsonld(
        excel_reader.read_workbook(wb, PACKS), PACKS, wb)


def _validate(doc: dict, shapes: Graph, tmp_path: Path):
    """Validate the way the CLI does, jurisdiction included.

    This called pyshacl directly, so it judged a v0.8 document by every shape in
    the file -- including ones introduced in v0.9, which no real `uofa shacl`
    invocation would apply to it. A test that exercises a shapes graph the
    product never assembles is testing a configuration nobody ships.
    """
    import copy as _copy

    from pyshacl import validate

    from uofa_cli.shacl_friendly import _apply_jurisdiction

    p = tmp_path / "p.jsonld"
    p.write_text(json.dumps(doc), encoding="utf-8")
    in_force = _apply_jurisdiction(_copy.deepcopy(shapes), p)
    conforms, _g, text = validate(data_graph=_load_data_graph(p), shacl_graph=in_force)
    return conforms, text


def _mutated(attributed: dict, fn) -> dict:
    doc = copy.deepcopy(attributed)
    factor = [f for f in doc["hasCredibilityFactor"]
              if isinstance(f, dict) and f.get("requiredLevelProvenance")][0]
    fn(factor)
    return doc


def test_the_package_the_tool_emits_declares_the_minimal_profile(attributed):
    """The premise the targeting decision rests on. If this ever changes, the
    reasoning in this module's docstring needs re-checking, not deleting."""
    assert str(attributed.get("conformsToProfile", "")).endswith("ProfileMinimal")


def test_the_factor_shape_is_unreachable_for_this_profile(attributed, shapes, tmp_path):
    """Recorded, not fixed here: a pre-existing gap wider than v0.8.

    Every CredibilityFactor constraint in the core shapes -- factorType,
    factorStatus, the 0-5 level ranges -- hangs off a shape only ProfileComplete
    reaches. On a ProfileMinimal package none of them run, and this proves it
    with a value the enum forbids.

    Left as a finding rather than repaired: reaching those constraints for
    Minimal packages changes what validates for every existing package, which is
    a decision about the profile system and not about v0.8.
    """
    doc = _mutated(attributed, lambda f: f.update(factorStatus="bogus"))
    conforms, text = _validate(doc, shapes, tmp_path)

    # **Assert the specific thing, not conformance.** This used to read
    # `assert conforms and ...`, so ANY unrelated violation made it fail with a
    # message announcing that the factor shape had started running -- a
    # conclusion the assertion could not support. It fired exactly that way when
    # v0.9 shapes were briefly reaching v0.8 fixtures, and the false report went
    # into a status summary as "uofa#109 closed as a side effect". It was not.
    # A failure message must only claim what its assertion actually tested.
    assert "bogus" not in text, (
        "the CredibilityFactor shape now reaches this profile — uofa#109 may be "
        "closed. Verify against a v0.9 ProfileMinimal package before believing "
        "it, then update this test and the issue.")


def test_a_fully_attributed_affirmation_conforms(attributed, shapes, tmp_path):
    conforms, text = _validate(attributed, shapes, tmp_path)
    assert conforms, text[:600]


def test_a_judgment_token_without_its_activity_is_refused(attributed, shapes, tmp_path):
    """The specimen v0.8 exists for, one layer below the CLI."""
    doc = _mutated(attributed, lambda f: f.pop("hasLevelAffirmation"))
    conforms, text = _validate(doc, shapes, tmp_path)
    assert not conforms
    assert "must carry a hasLevelAffirmation" in text


def test_an_activity_that_names_nobody_is_refused(attributed, shapes, tmp_path):
    """The hollow node: an activity satisfying an existence check and nothing else."""
    doc = _mutated(attributed,
                   lambda f: f.update(hasLevelAffirmation={"type": "LevelAffirmation"}))
    conforms, text = _validate(doc, shapes, tmp_path)
    assert not conforms


def test_an_activity_without_a_timestamp_is_refused(attributed, shapes, tmp_path):
    doc = _mutated(attributed, lambda f: f["hasLevelAffirmation"].pop("affirmedAt"))
    assert not _validate(doc, shapes, tmp_path)[0]


def test_confirmed_is_not_in_the_vocabulary(attributed, shapes, tmp_path):
    """Anchoring's token, refused at the shape layer too."""
    doc = _mutated(attributed,
                   lambda f: f.update(requiredLevelProvenance="confirmed"))
    conforms, text = _validate(doc, shapes, tmp_path)
    assert not conforms
    assert "requiredLevelProvenance must be one" in text


def test_a_machine_token_needs_no_agent(attributed, shapes, tmp_path):
    """`extracted` and `defaulted` claim no judgment, so they claim no agent --
    and must not be dragged into a constraint written for claims that do."""
    doc = _mutated(attributed, lambda f: (f.pop("hasLevelAffirmation"),
                                          f.update(requiredLevelProvenance="extracted")))
    assert _validate(doc, shapes, tmp_path)[0]


def test_a_package_with_no_provenance_at_all_still_conforms(attributed, shapes, tmp_path):
    """No `sh:minCount`. These shapes are version-agnostic, and requiring the
    token would refuse every package written before v0.8. Whether the token
    OUGHT to be present is answerable only against the declared context, which
    is not a triple -- so that half lives in the CLI, which can read it."""
    doc = _mutated(attributed, lambda f: (f.pop("hasLevelAffirmation"),
                                          f.pop("requiredLevelProvenance")))
    assert _validate(doc, shapes, tmp_path)[0]
