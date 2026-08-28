"""Workbook -> JSON-LD for v0.8's required-level provenance.

The gap this closes was found from the other side of the wheel. Two undefined
names -- `LEVEL_PROVENANCE_HEADER` in the reader, `LEVEL_TOKENS` in the mapper --
shipped into this package and were caught by *Credenza's* suite calling
`uofa import`, because nothing here drove a workbook that carried the column.
`LEVEL_TOKENS` sits behind `if factor.get("required_level") is not None`, so
only a fixture with populated required levels reaches it, and there was none.

This package ships as a wheel to third parties who do not have Credenza's suite.
It has to catch its own.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from uofa_cli import excel_mapper, excel_reader
from uofa_cli.excel_constants import (
    JUDGMENT_TOKENS, LEVEL_AFFIRMED_AT_HEADER, LEVEL_AFFIRMED_BY_HEADER,
    LEVEL_PROVENANCE_HEADER, LEVEL_TOKENS, WORKBOOK_PROFILE_HEADER,
    level_vocab_for,
)

REPO = Path(__file__).resolve().parents[1]
REVIEWED = REPO / "dev" / "build" / "pilot-johnson" / "johnson-extracted.xlsx"
PACKS = ["nasa-7009b"]
HEAD = 3


def _workbook_with(tmp_path: Path, tokens: list[str],
                   attribution: tuple[str, str] | None = None,
                   declare: str = "v0.9") -> Path:
    """A reviewed workbook carrying the provenance column and required levels.

    The levels matter: the emission is nested under `required_level is not
    None`, so a fixture that writes tokens onto level-less rows exercises
    nothing. A first draft of this file did exactly that and reported the
    emitter broken when it was the fixture.
    """
    if not REVIEWED.exists():
        pytest.skip(f"reviewed fixture not present: {REVIEWED}")
    out = tmp_path / "w.xlsx"
    shutil.copy2(REVIEWED, out)
    book = openpyxl.load_workbook(out)
    ws = book["Credibility Factors"]
    col = ws.max_column + 1
    ws.cell(row=HEAD, column=col).value = LEVEL_PROVENANCE_HEADER
    ws.cell(row=HEAD, column=col + 1).value = LEVEL_AFFIRMED_BY_HEADER
    ws.cell(row=HEAD, column=col + 2).value = LEVEL_AFFIRMED_AT_HEADER
    for i, token in enumerate(tokens):
        ws.cell(row=HEAD + 2 + i, column=col).value = token
        ws.cell(row=HEAD + 2 + i, column=3).value = 3      # required
        ws.cell(row=HEAD + 2 + i, column=4).value = 3      # achieved
        if attribution:
            ws.cell(row=HEAD + 2 + i, column=col + 1).value = attribution[0]
            ws.cell(row=HEAD + 2 + i, column=col + 2).value = attribution[1]
    # **The sheet declares what it is**, because the emitter now gates the
    # vocabulary on that declaration. An undeclared fixture buys the narrowest
    # vocabulary, which is correct behaviour and silently wrong as a fixture:
    # every token introduced after v0.8 would vanish and the test would read
    # that as the emitter dropping them.
    if declare:
        summary = book["Assessment Summary"]
        head = next(r for r in range(1, 12)
                    for c in range(1, 40)
                    if str(summary.cell(row=r, column=c).value or "").strip()
                    == "Project Name")
        pcol = summary.max_column + 1
        summary.cell(row=head, column=pcol).value = WORKBOOK_PROFILE_HEADER
        summary.cell(row=head + 1, column=pcol).value = declare
    book.save(out)
    return out


def _factors(wb: Path) -> list[dict]:
    doc = excel_mapper.map_to_jsonld(
        excel_reader.read_workbook(wb, PACKS), PACKS, wb)
    return [f for f in doc["hasCredibilityFactor"] if isinstance(f, dict)]


def test_the_reader_finds_the_column_by_header_not_by_position(tmp_path):
    """It is appended after the anchor column, so its index is not fixed."""
    wb = _workbook_with(tmp_path, ["affirmed", "corrected", "extracted"])
    got = [f.get("required_level_provenance")
           for f in excel_reader.read_workbook(wb, PACKS)["factors"][:3]]
    assert got == ["affirmed", "corrected", "extracted"]


def test_every_token_in_the_vocabulary_survives_the_round_trip(tmp_path):
    """The whole closed set, not a sample. `LEVEL_TOKENS` is the contract."""
    tokens = sorted(LEVEL_TOKENS)
    wb = _workbook_with(tmp_path, tokens)
    got = [f.get("requiredLevelProvenance") for f in _factors(wb)[:len(tokens)]]
    assert got == tokens


def test_confirmed_is_never_emitted_as_a_provenance_value(tmp_path):
    """The specimen that fooled the first fix.

    `confirmed` is the encoding tool's LOCATION act -- anchoring produces it --
    and a run-25-shaped package carried nothing else. Asserted against the
    provenance FIELD, never against the document's text: `confirmed` occurs
    legitimately in rationale prose, and a substring search over the whole
    document calls that a violation.
    """
    wb = _workbook_with(tmp_path, ["confirmed", "confirmed", "confirmed"])
    values = {f.get("requiredLevelProvenance") for f in _factors(wb)}
    assert "confirmed" not in LEVEL_TOKENS
    assert "confirmed" not in values, (
        "anchoring's token reached the package as a judgment claim")


def test_a_judgment_token_never_gets_an_activity_that_names_nobody(tmp_path):
    """The hollow node, which is worse than no node.

    The workbook carries no attribution columns, so nothing can fill `actor` or
    `affirmedAt`. Emitting the node anyway produced `{"type":
    "LevelAffirmation"}` -- an activity satisfying any check that asks whether
    one EXISTS, while naming nobody and timing nothing. Omitting it leaves the
    judgment token without its activity, which the v0.8 shape refuses: the
    honest signal that this source cannot say who judged.
    """
    wb = _workbook_with(tmp_path, sorted(JUDGMENT_TOKENS))
    for f in _factors(wb)[:len(JUDGMENT_TOKENS)]:
        assert f.get("requiredLevelProvenance") in JUDGMENT_TOKENS
        node = f.get("hasLevelAffirmation")
        assert node is None or set(node) - {"type"}, (
            f"emitted an affirmation naming nobody: {node}")


def test_a_machine_token_carries_no_affirmation_node(tmp_path):
    """`extracted` and `defaulted` claim no judgment, so they claim no agent."""
    machine = sorted(set(LEVEL_TOKENS) - set(JUDGMENT_TOKENS))
    wb = _workbook_with(tmp_path, machine)
    for f in _factors(wb)[:len(machine)]:
        assert "hasLevelAffirmation" not in f


def test_attribution_travels_from_the_sheet_into_the_activity_node(tmp_path):
    """The whole chain: two columns, the reader, the emitter, the node.

    The workbook gained `Affirmed By` and `Affirmed At` rather than the shape
    gaining an exemption for sheet-carried packages. A claim that answers "who
    judged this" in JSON-LD and shrugs in a workbook is one contract in two
    dialects, and the data was never missing -- the encoding tool records the
    actor and timestamp of every affirm act. The sheet simply had no cells.
    """
    wb = _workbook_with(tmp_path, ["affirmed"],
                        attribution=("V. Vettrivel", "2026-08-24T10:00:00Z"))
    factor = excel_reader.read_workbook(wb, PACKS)["factors"][0]
    assert factor["affirmed_by"] == "V. Vettrivel"
    assert factor["affirmed_at"] == "2026-08-24T10:00:00Z"

    f = _factors(wb)[0]
    assert f["requiredLevelProvenance"] == "affirmed"
    node = f["hasLevelAffirmation"]
    assert node["type"] == "LevelAffirmation"
    assert node["affirmedAt"] == "2026-08-24T10:00:00Z"

    # **`actor` is an IRI and `role` is the name**, the split `hasDecisionRecord`
    # already uses. `actor` is declared `"@type": "@id"`, so a bare name there is
    # not stored as a name: JSON-LD resolves it as a relative IRI against the
    # document and "V. Vettrivel" became `file:///.../V.%20Vettrivel`, an
    # identifier that points nowhere and changes with where the file sits.
    assert node["role"] == "V. Vettrivel"
    assert node["actor"].startswith("http"), node["actor"]
    assert node["actor"].endswith("/org/v-vettrivel")


def test_the_affirming_agent_survives_expansion_as_an_iri(tmp_path):
    """Asserted after RDF expansion, because that is where the defect showed.

    The emitted JSON looked correct -- `"actor": "V. Vettrivel"` reads like a
    name in a name-shaped field. Only expanding it revealed a `file://` IRI.
    A test that stops at the JSON cannot see this class of bug at all.
    """
    import json as _json

    from uofa_cli.shacl_friendly import _load_data_graph

    wb = _workbook_with(tmp_path, ["affirmed"],
                        attribution=("V. Vettrivel", "2026-08-24T10:00:00Z"))
    out = tmp_path / "p.jsonld"
    out.write_text(_json.dumps(excel_mapper.map_to_jsonld(
        excel_reader.read_workbook(wb, PACKS), PACKS, wb)))

    actors = [str(o) for _s, p, o in _load_data_graph(out) if str(p).endswith("#actor")]
    assert actors, "no actor survived expansion"
    for a in actors:
        assert not a.startswith("file:"), (
            f"the agent expanded to a filesystem path: {a}")
        assert a.startswith("http") or a.startswith("urn:")


def test_columns_present_but_empty_is_not_an_affirmation(tmp_path):
    """An export writes the columns for every row; most rows have nothing to say.

    So the common case is a judgment token beside two blank cells, and it must
    read as unattributed rather than as an affirmation by the empty string. The
    reader distinguishes a MISSING column (the sheet cannot say) from an empty
    one (it can and did not); neither is an agent.
    """
    wb = _workbook_with(tmp_path, ["affirmed"], attribution=("", ""))
    factor = excel_reader.read_workbook(wb, PACKS)["factors"][0]
    assert factor["affirmed_by"] is None and factor["affirmed_at"] is None
    assert "hasLevelAffirmation" not in _factors(wb)[0]


def test_the_emitted_terms_are_all_defined_by_the_declared_context(tmp_path):
    """A package may not use a term its own declared context does not define."""
    wb = _workbook_with(tmp_path, sorted(JUDGMENT_TOKENS))
    doc = excel_mapper.map_to_jsonld(
        excel_reader.read_workbook(wb, PACKS), PACKS, wb)
    name = str(doc["@context"]).rsplit("/", 1)[-1]
    ctx = json.loads((REPO / "spec" / "context" / name).read_text())["@context"]
    for term in ("requiredLevelProvenance", "LevelAffirmation",
                 "hasLevelAffirmation", "affirmedAt"):
        assert term in ctx, f"{name} does not define {term}"


# ─────────────────────────────────────────────────────────────────────────────
# The vocabulary is closed AT THE VERSION THE SOURCE DECLARES.
# ─────────────────────────────────────────────────────────────────────────────

def test_a_v0_8_workbook_cannot_launder_a_v0_9_term_into_a_package(tmp_path):
    """The disagreement must not vanish at the boundary.

    The emitter writes the current context onto everything it produces. So a
    workbook declaring v0.8 and carrying `not-recoverable` -- a term v0.8 has no
    way to mean -- would import into a package whose context makes the term
    legal, and every check downstream would see a well-formed v0.9 package. The
    only place that disagreement is still visible is here, at the read.

    Dropping the token leaves the level unjudged, which is true of it, and the
    levels check refuses the package by name.
    """
    wb = _workbook_with(tmp_path, ["not-recoverable", "affirmed"], declare="v0.8")
    values = [f.get("requiredLevelProvenance") for f in _factors(wb)[:2]]
    assert values[0] is None, "a v0.8 sheet's out-of-vocabulary term reached the package"
    assert values[1] == "affirmed", "the gate dropped a term v0.8 does define"


def test_the_same_term_survives_when_the_source_declares_v0_9(tmp_path):
    """And the gate is not a blanket refusal: it is the declaration, honoured."""
    wb = _workbook_with(tmp_path, ["not-recoverable"], declare="v0.9")
    assert _factors(wb)[0].get("requiredLevelProvenance") == "not-recoverable"


def test_an_undeclared_workbook_gets_the_narrowest_vocabulary(tmp_path):
    """Silence is not a claim to the widest set.

    A sheet that declares nothing has not earned any vocabulary, and reading its
    tokens anyway is the field-sniffing the declaration exists to replace.
    """
    wb = _workbook_with(tmp_path, ["affirmed"], declare="")
    assert _factors(wb)[0].get("requiredLevelProvenance") is None


def test_every_token_survives_the_round_trip_under_its_own_version(tmp_path):
    """`LEVEL_TOKENS` is the contract, and each term round-trips from the
    version that introduced it -- which is a stronger statement than the flat
    set was making."""
    tokens = sorted(level_vocab_for((0, 9)))
    wb = _workbook_with(tmp_path, tokens, declare="v0.9")
    assert [f.get("requiredLevelProvenance")
            for f in _factors(wb)[:len(tokens)]] == tokens
    assert set(tokens) == set(LEVEL_TOKENS), "the flat set and the versioned set disagree"


def test_the_package_declares_the_version_its_source_declared(tmp_path):
    """Jurisdiction from the emitter's side.

    A v0.8 workbook transcribed today is a v0.8 document: nothing in it is
    newer than what the sheet held. Stamping the current vocabulary on it would
    subject it to five decision-model shapes introduced at v0.9 that its content
    never claimed to satisfy -- refusing it for missing something nobody asked
    it for, which is the exact reasoning `_apply_jurisdiction` exists to reject.
    """
    def _context(declare):
        wb = _workbook_with(tmp_path / declare.replace(".", "") or "x",
                            ["affirmed"], declare=declare)
        doc = excel_mapper.map_to_jsonld(
            excel_reader.read_workbook(wb, PACKS), PACKS, wb)
        return doc["@context"].rsplit("/", 1)[-1]

    (tmp_path / "v08").mkdir()
    (tmp_path / "v09").mkdir()
    assert _context("v0.8") == "v0.8.jsonld"
    assert _context("v0.9") == "v0.9.jsonld"


def test_an_undeclared_workbook_gets_the_oldest_supported_context(tmp_path):
    """Silence is not a claim to the newest vocabulary, here either."""
    wb = _workbook_with(tmp_path, ["affirmed"], declare="")
    doc = excel_mapper.map_to_jsonld(
        excel_reader.read_workbook(wb, PACKS), PACKS, wb)
    assert doc["@context"].endswith("/v0.8.jsonld")
