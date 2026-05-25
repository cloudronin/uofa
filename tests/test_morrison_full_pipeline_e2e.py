"""End-to-end Morrison pipeline test: 2 evidence dirs → 2 jsonld → reports → diff.

Morrison (VV40 reference example: centrifugal blood pump hemolysis CFD,
Morrison et al. 2019) covers two clinical Contexts of Use of the same
underlying CFD model:

  COU1 — Cardiopulmonary Bypass (Class II, MRL 2, Medium assurance, ACCEPTED)
  COU2 — Ventricular Assist Device (Class III, MRL 5, Low assurance, NOT ACCEPTED)

Evidence ships as two parallel folders (``morrison-evidence-cou1/`` and
``morrison-evidence-cou2/``) that duplicate the shared CFD/validation data
and differ only in the COU-specific decision rationale and Monte Carlo UQ
study. This mirrors the NASA aero pattern so the chained e2e
(extract → import → check → rules → diff) is symmetric between the two
reference cases.

Two variants:

* **TestMorrisonFullPipelineE2EMock** — always runs. Uses ``--model mock``;
  asserts plumbing (extract produces 13 VV40 factors per COU, import
  produces valid JSON-LD, rules + check + diff complete without crashing).

* **TestMorrisonFullPipelineE2ERealLLM** — gated by ``UOFA_RUN_REAL_LLM=1``.
  Uses ``ollama/qwen3.5:4b``. Asserts loose plumbing only — the LLM's
  extracted COU profile from shared evidence isn't predictable enough to
  pin a specific weakener firing pattern. Catches the LLM-output → import
  vocabulary gap (today's bug class) and wheel/bundling drift.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# pytest's rootdir-import mode (no tests/__init__.py) puts the tests
# directory on sys.path, so the bare module name resolves. Don't use
# `from tests.test_extract_integration` — that breaks CI collection.
from test_extract_integration import (
    run_uofa, MORRISON_COU1_DIR, MORRISON_COU2_DIR,
)

REAL_LLM_ENABLED = os.environ.get("UOFA_RUN_REAL_LLM") == "1"

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")


# ── Shared chain runner ───────────────────────────────────────


def _extract_and_import(tmp_dir: Path, evidence_dir: Path, cou_label: str, model: str):
    """Run extract → import for one Morrison COU. Returns (xlsx, jsonld)."""
    xlsx = tmp_dir / f"{cou_label}.xlsx"
    jsonld = tmp_dir / f"{cou_label}.jsonld"

    extract = run_uofa(
        "extract", str(evidence_dir),
        "--model", model,
        "--pack", "vv40",
        "--output", str(xlsx),
    )
    assert extract.returncode == 0, (
        f"extract failed for {cou_label}:\n"
        f"STDOUT: {extract.stdout}\nSTDERR: {extract.stderr}"
    )
    assert xlsx.exists(), f"extract did not produce {xlsx}"

    imp = run_uofa(
        "--pack", "vv40", "import", str(xlsx),
        "--output", str(jsonld),
    )
    assert imp.returncode == 0, (
        f"import failed for {cou_label}:\n"
        f"STDOUT: {imp.stdout}\nSTDERR: {imp.stderr}"
    )
    assert jsonld.exists(), f"import did not produce {jsonld}"

    return xlsx, jsonld


# ── Mock variant: always runs in CI ───────────────────────────


@pytest.mark.skipif(not MORRISON_COU1_DIR.exists(), reason="morrison-evidence-cou1 not available")
@pytest.mark.skipif(not MORRISON_COU2_DIR.exists(), reason="morrison-evidence-cou2 not available")
class TestMorrisonFullPipelineE2EMock:
    """Plumbing-level e2e: every step in the chain completes for both COUs.

    Does NOT assert semantic rule firings (mock data is canned; doesn't
    reproduce Morrison's COU1-vs-COU2 weakener divergence — see the
    per-step TestRules tests in test_integration.py using the pre-built
    cou1/cou2 fixtures for that).
    """

    @pytest.fixture(scope="class")
    def chain(self, tmp_path_factory):
        tmp = tmp_path_factory.mktemp("morrison_e2e_mock")
        cou1_xlsx, cou1_jsonld = _extract_and_import(tmp, MORRISON_COU1_DIR, "cou1", "mock")
        cou2_xlsx, cou2_jsonld = _extract_and_import(tmp, MORRISON_COU2_DIR, "cou2", "mock")
        return {
            "tmp": tmp,
            "cou1_xlsx": cou1_xlsx, "cou1_jsonld": cou1_jsonld,
            "cou2_xlsx": cou2_xlsx, "cou2_jsonld": cou2_jsonld,
        }

    def test_extract_cou1_produces_13_vv40_factors(self, chain):
        import openpyxl
        from uofa_cli.excel_constants import VV40_FACTOR_NAMES
        wb = openpyxl.load_workbook(chain["cou1_xlsx"])
        ws = wb["Credibility Factors"]
        found = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        vv40 = [v for v in found if v in VV40_FACTOR_NAMES]
        assert len(vv40) == 13, f"COU1 has {len(vv40)} VV40 factors, expected 13"

    def test_extract_cou2_produces_13_vv40_factors(self, chain):
        import openpyxl
        from uofa_cli.excel_constants import VV40_FACTOR_NAMES
        wb = openpyxl.load_workbook(chain["cou2_xlsx"])
        ws = wb["Credibility Factors"]
        found = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        vv40 = [v for v in found if v in VV40_FACTOR_NAMES]
        assert len(vv40) == 13, f"COU2 has {len(vv40)} VV40 factors, expected 13"

    def test_imports_produce_valid_jsonld(self, chain):
        import json
        for label in ("cou1", "cou2"):
            doc = json.loads(chain[f"{label}_jsonld"].read_text())
            assert "v0.5.jsonld" in str(doc.get("@context", "")), (
                f"{label}.jsonld missing v0.5 context")
            assert "bindsRequirement" in doc, (
                f"{label}.jsonld missing bindsRequirement")

    def test_rules_runs_on_cou1_without_crash(self, chain):
        result = run_uofa(
            "rules", str(chain["cou1_jsonld"]),
            "--pack", "vv40", "--build",
        )
        assert result.returncode == 0, (
            f"rules failed on cou1:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )

    def test_rules_runs_on_cou2_without_crash(self, chain):
        result = run_uofa(
            "rules", str(chain["cou2_jsonld"]),
            "--pack", "vv40", "--build",
        )
        assert result.returncode == 0, (
            f"rules failed on cou2:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )

    def test_check_runs_without_crash(self, chain):
        # Mock data may or may not satisfy every SHACL/C1 constraint. Reject
        # only an unhandled crash (returncode > 1 or a Python traceback in
        # stderr). Specific pass/fail signal is exercised by
        # TestCheck::test_check_morrison_all_pass against the pre-built
        # canonical fixture in test_integration.py.
        for label in ("cou1", "cou2"):
            result = run_uofa(
                "check", str(chain[f"{label}_jsonld"]),
                "--pack", "vv40",
                "--skip-rules",  # skip Java-dependent C3 in plumbing test
            )
            assert result.returncode in (0, 1), (
                f"check {label} crashed unexpectedly (rc={result.returncode}):\n"
                f"STDOUT: {result.stdout}\nSTDERR: {result.stderr}"
            )
            assert "Traceback" not in result.stderr, (
                f"check {label} produced a Python traceback:\n{result.stderr}"
            )

    def test_diff_chains_both_cous_without_crash(self, chain):
        result = run_uofa(
            "diff", str(chain["cou1_jsonld"]), str(chain["cou2_jsonld"]),
            "--pack", "vv40", "--build",
        )
        # diff always exits 0 (see commands/diff.py:396 — exit code is
        # hard-coded regardless of divergence). Plumbing-level: doesn't crash.
        assert result.returncode == 0, (
            f"diff crashed unexpectedly (rc={result.returncode}):\n"
            f"STDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )


# ── Real-LLM variant: gated by env var ────────────────────────


@pytest.mark.skipif(not REAL_LLM_ENABLED,
                    reason="set UOFA_RUN_REAL_LLM=1 to run real-LLM e2e")
@pytest.mark.skipif(not MORRISON_COU1_DIR.exists(), reason="morrison-evidence-cou1 not available")
@pytest.mark.skipif(not MORRISON_COU2_DIR.exists(), reason="morrison-evidence-cou2 not available")
class TestMorrisonFullPipelineE2ERealLLM:
    """Real-LLM e2e for Morrison. Loose assertions only.

    Catches: LLM-output → import vocabulary gaps, wheel bundling drift,
    mock-only blind spots. Does NOT assert the canonical Morrison
    weakener-firing divergence pattern (pre-built-fixture tests in
    test_integration.py cover that on stable inputs).
    """

    @pytest.fixture(scope="class")
    def chain(self, tmp_path_factory):
        tmp = tmp_path_factory.mktemp("morrison_e2e_real")
        cou1_xlsx, cou1_jsonld = _extract_and_import(
            tmp, MORRISON_COU1_DIR, "cou1", "ollama/qwen3.5:4b")
        cou2_xlsx, cou2_jsonld = _extract_and_import(
            tmp, MORRISON_COU2_DIR, "cou2", "ollama/qwen3.5:4b")
        return {"tmp": tmp,
                "cou1_jsonld": cou1_jsonld, "cou2_jsonld": cou2_jsonld}

    def test_rules_runs_on_both_cous(self, chain):
        for label in ("cou1", "cou2"):
            result = run_uofa(
                "rules", str(chain[f"{label}_jsonld"]),
                "--pack", "vv40", "--build",
            )
            assert result.returncode == 0, (
                f"rules failed on {label}:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
            )

    def test_diff_runs_without_crash(self, chain):
        result = run_uofa(
            "diff", str(chain["cou1_jsonld"]), str(chain["cou2_jsonld"]),
            "--pack", "vv40", "--build",
        )
        # diff always exits 0 (commands/diff.py:396). Whether the two COUs
        # diverged or not is signaled in the stdout summary line; the
        # real-LLM variant doesn't pin a specific divergence count because
        # Morrison's shared-evidence structure can extract to identical or
        # different firing profiles depending on LLM run-to-run variance.
        assert result.returncode == 0, (
            f"diff crashed (rc={result.returncode}):\n"
            f"STDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )
