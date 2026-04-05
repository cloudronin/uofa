"""Tests for excel_writer — template writing, confidence coloring, source comments."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from uofa_cli.llm_extractor import (
    ExtractionResult, FieldExtraction, _mock_extract, _json_to_result,
)
from uofa_cli.excel_writer import (
    write_extraction, _fuzzy_match_dropdown, _apply_confidence_color,
)
from uofa_cli.excel_constants import (
    VALID_DECISION_OUTCOMES, VV40_FACTOR_NAMES, NASA_ALL_FACTOR_NAMES,
)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

pytestmark = pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl required")


# ── Fixtures ─────────────────────────────────────────────────


@pytest.fixture
def mock_vv40_result():
    """A mock ExtractionResult for VV40."""
    raw = json.loads(_mock_extract("vv40"))
    return _json_to_result(raw, "vv40")


@pytest.fixture
def mock_nasa_result():
    """A mock ExtractionResult for NASA."""
    raw = json.loads(_mock_extract("nasa-7009b"))
    return _json_to_result(raw, "nasa-7009b")


# ── write_extraction (no template) ───────────────────────────


class TestWriteExtractionNoTemplate:
    def test_creates_xlsx(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        assert out.exists()

    def test_has_all_sheets(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        expected = {"Assessment Summary", "Model & Data", "Validation Results",
                    "Credibility Factors", "Decision"}
        assert expected.issubset(set(wb.sheetnames))

    def test_summary_populated(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Assessment Summary"]
        # Data row should have project name
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Mock Project":
                found = True
                break
        assert found, "Project name not found in Assessment Summary"

    def test_factors_populated(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        factor_names_found = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in VV40_FACTOR_NAMES:
                factor_names_found.append(val)
        assert len(factor_names_found) == 13

    def test_factor_levels_written(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        levels_found = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in VV40_FACTOR_NAMES:
                req = ws.cell(row=row, column=3).value
                ach = ws.cell(row=row, column=4).value
                if req is not None and ach is not None:
                    levels_found += 1
        assert levels_found == 13

    def test_column_h_empty(self, mock_vv40_result, tmp_path):
        """Column H (Linked Evidence) must be empty for all factor rows."""
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in VV40_FACTOR_NAMES:
                linked = ws.cell(row=row, column=8).value
                assert linked is None, f"Column H not empty for {val}: {linked}"

    def test_decision_populated(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Decision"]
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Accepted":
                found = True
                break
        assert found, "Decision outcome not found"

    def test_nasa_19_factors(self, mock_nasa_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_nasa_result, None, out, "nasa-7009b")
        wb = openpyxl.load_workbook(out)
        ws = wb["Credibility Factors"]
        factor_names_found = []
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and val in NASA_ALL_FACTOR_NAMES:
                factor_names_found.append(val)
        assert len(factor_names_found) == 19


# ── Confidence coloring ──────────────────────────────────────


class TestConfidenceColoring:
    def test_green_high_confidence(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Assessment Summary"]
        # Project name has confidence 0.95 → green
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == "Mock Project":
                fill = cell.fill
                assert fill.fgColor and fill.fgColor.rgb
                assert "C6EFCE" in str(fill.fgColor.rgb)
                break

    def test_yellow_medium_confidence(self, tmp_path):
        """Create a result with medium confidence and verify yellow fill."""
        result = ExtractionResult()
        result.assessment_summary = {
            "project_name": FieldExtraction(value="Test", confidence=0.70, source_file="f.pdf"),
        }
        out = tmp_path / "output.xlsx"
        write_extraction(result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Assessment Summary"]
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == "Test":
                assert "FFEB9C" in str(cell.fill.fgColor.rgb)
                break

    def test_red_low_confidence(self, tmp_path):
        result = ExtractionResult()
        result.assessment_summary = {
            "project_name": FieldExtraction(value="Weak", confidence=0.30, source_file="f.pdf"),
        }
        out = tmp_path / "output.xlsx"
        write_extraction(result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Assessment Summary"]
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == "Weak":
                assert "FFC7CE" in str(cell.fill.fgColor.rgb)
                break


# ── Cell comments ────────────────────────────────────────────


class TestCellComments:
    def test_source_in_comment(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, None, out, "vv40")
        wb = openpyxl.load_workbook(out)
        ws = wb["Assessment Summary"]
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == "Mock Project":
                assert cell.comment is not None
                assert "mock-report.pdf" in cell.comment.text
                assert "Confidence: 95%" in cell.comment.text
                break


# ── Fuzzy dropdown matching ──────────────────────────────────


class TestFuzzyDropdown:
    def test_exact_match(self):
        assert _fuzzy_match_dropdown("Accepted", VALID_DECISION_OUTCOMES) == "Accepted"

    def test_fuzzy_match(self):
        assert _fuzzy_match_dropdown("Acepted", VALID_DECISION_OUTCOMES) == "Accepted"

    def test_no_match_returns_original(self):
        result = _fuzzy_match_dropdown("Totally wrong", VALID_DECISION_OUTCOMES)
        assert result == "Totally wrong"


# ── Template-based writing ───────────────────────────────────


VV40_TEMPLATE = (
    Path(__file__).parent.parent / "packs" / "vv40" / "templates" / "vv40-template.xlsx"
)


@pytest.mark.skipif(not VV40_TEMPLATE.exists(), reason="VV40 template not available")
class TestWriteWithTemplate:
    def test_template_based_write(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, VV40_TEMPLATE, out, "vv40")
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        # Should have all standard sheets
        assert "Assessment Summary" in wb.sheetnames
        assert "Credibility Factors" in wb.sheetnames

    def test_template_preserves_instructions(self, mock_vv40_result, tmp_path):
        out = tmp_path / "output.xlsx"
        write_extraction(mock_vv40_result, VV40_TEMPLATE, out, "vv40")
        wb = openpyxl.load_workbook(out)
        # Check that any _Lists or Instructions sheets are preserved
        template_wb = openpyxl.load_workbook(VV40_TEMPLATE)
        for sheet_name in template_wb.sheetnames:
            if sheet_name.startswith("_") or sheet_name == "Instructions":
                assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} not preserved"
