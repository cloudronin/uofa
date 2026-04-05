"""Dict-based fixture generator for import E2E tests.

Each test case is a Python dict describing the Excel content + expected outcomes.
generate_fixture() creates an xlsx that excel_reader.py can parse.

Usage:
    from tests.fixtures.import.generator import SPECS, generate_fixture
    generate_fixture(SPECS["e2e-clean-vv40"]["data"], Path("output.xlsx"))
"""

import copy
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Add src to path for constants
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent / "src"))
from uofa_cli.excel_constants import (
    VV40_FACTOR_CATEGORIES, NASA_ONLY_FACTOR_CATEGORIES, ALL_FACTOR_CATEGORIES,
)


# ── xlsx writer ───────────────────────────────────────────────


def generate_fixture(spec: dict, output_path: Path) -> None:
    """Generate an xlsx fixture from a test spec dict."""
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Special case: not-xlsx error test
    if spec.get("_raw_text"):
        output_path.write_text(spec["_raw_text"])
        return

    # Only create sheets that the spec includes
    if "summary" in spec:
        _write_summary(wb, spec["summary"])
    if "entities" in spec:
        _write_entities(wb, spec["entities"])
    if "validation_results" in spec:
        _write_validation_results(wb, spec["validation_results"])
    if "factors" in spec:
        _write_factors(wb, spec["factors"])
    if "decision" in spec:
        _write_decision(wb, spec["decision"])

    wb.save(output_path)


def _write_summary(wb, summary):
    ws = wb.create_sheet("Assessment Summary", 0)
    ws.append(["Unit of Assurance — Assessment Summary"])
    ws.append(["Project Name", "COU Name", "COU Description", "Profile",
               "Device Class", "Model Risk Level", "Assurance Level",
               "Standards Reference", "Assessor Name", "Assessment Date",
               "Source Document", "Has UQ?"])
    ws.append([
        summary.get("project_name"),
        summary.get("cou_name"),
        summary.get("cou_description"),
        summary.get("profile"),
        summary.get("device_class"),
        summary.get("model_risk_level"),
        summary.get("assurance_level"),
        summary.get("standards_reference"),
        summary.get("assessor_name"),
        summary.get("assessment_date"),
        summary.get("source_document"),
        summary.get("has_uq"),
    ])


def _write_entities(wb, entities):
    ws = wb.create_sheet("Model & Data")
    ws.append(["Model & Data"])
    ws.append(["Entity Type", "Name", "Identifier/URI", "Description", "Version", "Source"])
    for e in entities:
        ws.append([e.get("type"), e.get("name"), e.get("uri"),
                   e.get("description"), e.get("version"), e.get("source")])


def _write_validation_results(wb, results):
    ws = wb.create_sheet("Validation Results")
    ws.append(["Validation Results"])
    # Always use v2 layout with Type column
    ws.append(["Result Name", "Type", "Identifier/URI", "Description",
               "Compares To", "Has UQ?", "UQ Method", "Metric Value", "Pass/Fail"])
    for r in results:
        ws.append([
            r.get("name"),
            r.get("evidence_type", "ValidationResult"),
            r.get("uri"),
            r.get("description"),
            r.get("compares_to"),
            r.get("has_uq"),
            r.get("uq_method"),
            r.get("metric_value"),
            r.get("pass_fail"),
        ])


def _write_factors(wb, factors):
    ws = wb.create_sheet("Credibility Factors")
    ws.append(["Credibility Factors"])
    ws.append([""])
    ws.append(["Factor Type", "Category", "Required Level", "Achieved Level",
               "Acceptance Criteria", "Rationale", "Factor Status", "Linked Evidence"])
    ws.append(["(pre-filled)", "(pre-filled)", "(1-5)", "(1-5)",
               "(optional)", "(optional)", "(status)", "(URI)"])
    for f in factors:
        ws.append([
            f.get("factor_type"),
            f.get("category"),
            f.get("required_level"),
            f.get("achieved_level"),
            f.get("acceptance_criteria"),
            f.get("rationale"),
            f.get("status", "assessed"),
            f.get("linked_evidence"),
        ])


def _write_decision(wb, decision):
    ws = wb.create_sheet("Decision")
    ws.append(["Decision"])
    ws.append(["Decision Outcome", "Decision Rationale", "Criteria Set",
               "Decided By", "Decision Date"])
    ws.append([
        decision.get("outcome"),
        decision.get("rationale"),
        decision.get("criteria_set"),
        decision.get("decided_by"),
        decision.get("decision_date"),
    ])


# ── Clean base helpers ────────────────────────────────────────


def _clean_factor(name, category, level=2):
    """A clean assessed factor that fires no weakeners."""
    return {
        "factor_type": name,
        "category": category,
        "required_level": level,
        "achieved_level": level,
        "acceptance_criteria": f"Acceptance criteria for {name}",
        "rationale": f"Rationale for {name}",
        "status": "assessed",
        "linked_evidence": None,
    }


def _clean_vv40_factors():
    """13 clean VV40 factors — all assessed, levels matching, all with criteria."""
    return [_clean_factor(name, cat) for name, cat in VV40_FACTOR_CATEGORIES]


def _clean_nasa_factors():
    """19 clean NASA factors — includes hasEvidence where rules require it."""
    factors = _clean_vv40_factors()
    for name, cat in NASA_ONLY_FACTOR_CATEGORIES:
        f = _clean_factor(name, cat)
        # NASA rules check hasEvidence on these specific factors
        if name == "Development technical review":
            f["linked_evidence"] = "https://example.org/validation/review"
        elif name == "Development process and product management":
            f["linked_evidence"] = "https://example.org/validation/process"
        elif name == "Results robustness":
            f["linked_evidence"] = "https://example.org/validation/sensitivity"
        factors.append(f)
    return factors


def _clean_base_vv40():
    """A structurally complete VV40 spec that produces 0 L1 weakeners."""
    return {
        "summary": {
            "project_name": "Clean VV40 Project",
            "cou_name": "Clean COU",
            "cou_description": "A fully compliant credibility assessment.",
            "profile": "Complete",
            "device_class": "Class II",
            "model_risk_level": "MRL 1",  # Low MRL → W-EP-04 won't fire on any stray not-assessed
            "assurance_level": "Low",     # Low → COMPOUND-03 won't fire on any stray Critical
            "standards_reference": "ASME-VV40-2018",
            "assessor_name": "Test Assessor",
            "assessment_date": "2026-01-15",
            "source_document": "https://example.org/report/clean",
            "has_uq": "Yes",
        },
        "entities": [
            {"type": "Requirement", "name": "Safety requirement",
             "uri": "https://example.org/req/1", "description": "Test requirement"},
            {"type": "Model", "name": "Test model",
             "uri": "https://example.org/model/1", "description": "Test model"},
            {"type": "Dataset", "name": "Test dataset",
             "uri": "https://example.org/data/1", "description": "Test dataset"},
        ],
        "validation_results": [
            {
                "name": "Validation comparison",
                "evidence_type": "ValidationResult",
                "uri": "https://example.org/validation/1",
                "description": "Comparison with experimental data",
                "compares_to": "https://example.org/data/1",
                "has_uq": "Yes",
                "uq_method": "GCI",
                "metric_value": "1.5%",
                "pass_fail": "Pass",
            },
        ],
        "factors": _clean_vv40_factors(),
        "decision": {
            "outcome": "Accepted",
            "rationale": "All factors meet required levels.",
            "criteria_set": "https://uofa.net/criteria/ASME-VV40-2018",
            "decided_by": "Test Board",
            "decision_date": "2026-01-15",
        },
    }


def _clean_base_nasa():
    """A structurally complete NASA spec that produces 0 L1 weakeners."""
    base = _clean_base_vv40()
    base["summary"]["standards_reference"] = "NASA-STD-7009B"
    base["summary"]["model_risk_level"] = "MRL 1"
    base["factors"] = _clean_nasa_factors()
    # Add a ReviewActivity VR so W-NASA-02 has evidence available
    base["validation_results"].append({
        "name": "Technical review",
        "evidence_type": "ReviewActivity",
        "uri": "https://example.org/validation/review",
        "compares_to": "https://example.org/org/reviewer",
    })
    base["validation_results"].append({
        "name": "Process attestation",
        "evidence_type": "ProcessAttestation",
        "uri": "https://example.org/validation/process",
        "compares_to": "https://example.org/org/attester",
    })
    return base


def _mutate(base, path, value):
    """Deep-copy base and set a nested value. path is a list of keys/indices."""
    spec = copy.deepcopy(base)
    obj = spec
    for key in path[:-1]:
        obj = obj[key]
    obj[path[-1]] = value
    return spec


def _mutate_factor(base, factor_index, field, value):
    """Deep-copy base and mutate one factor field."""
    spec = copy.deepcopy(base)
    spec["factors"][factor_index][field] = value
    return spec


def _mutate_vr(base, vr_index, field, value):
    """Deep-copy base and mutate one validation result field."""
    spec = copy.deepcopy(base)
    spec["validation_results"][vr_index][field] = value
    return spec


# ── Test case specs ───────────────────────────────────────────


_VV40 = _clean_base_vv40()
_NASA = _clean_base_nasa()

# --- Passing structural tests (no weakener assertions needed) ---

_vv40_minimal = {
    "summary": {
        "project_name": "Minimal Project",
        "cou_name": "Minimal COU",
        "profile": "Minimal",
    },
    "entities": [
        {"type": "Requirement", "name": "Safety req", "uri": "https://example.org/req/1"},
    ],
    "validation_results": [
        {"name": "Basic validation", "has_uq": "No",
         "compares_to": "https://example.org/data/1"},
    ],
    "decision": {"outcome": "Accepted", "rationale": "Minimal acceptance"},
}

_nasa_minimal = copy.deepcopy(_vv40_minimal)
_nasa_minimal["summary"]["project_name"] = "NASA Minimal"

_mixed_status = copy.deepcopy(_VV40)
_mixed_status["summary"]["project_name"] = "Mixed Status"
_mixed_status["summary"]["cou_name"] = "Mixed COU"
for i in range(7, 10):  # factors 7-9 not-assessed
    _mixed_status["factors"][i]["status"] = "not-assessed"
    _mixed_status["factors"][i]["required_level"] = None
    _mixed_status["factors"][i]["achieved_level"] = None
for i in range(10, 12):  # factors 10-11 scoped-out
    _mixed_status["factors"][i]["status"] = "scoped-out"
    _mixed_status["factors"][i]["required_level"] = None
    _mixed_status["factors"][i]["achieved_level"] = None
_mixed_status["factors"][12]["status"] = "not-applicable"
_mixed_status["factors"][12]["required_level"] = None
_mixed_status["factors"][12]["achieved_level"] = None

_uri_gen = copy.deepcopy(_VV40)
_uri_gen["summary"]["project_name"] = "My Test Project"
_uri_gen["summary"]["cou_name"] = "Thermal Analysis COU"

# --- Error cases ---

_err_missing_summary = copy.deepcopy(_VV40)
del _err_missing_summary["summary"]

_err_missing_decision = copy.deepcopy(_VV40)
del _err_missing_decision["decision"]

_err_invalid_decision = copy.deepcopy(_vv40_minimal)
_err_invalid_decision["decision"]["outcome"] = "Approved"

_err_invalid_factor = {
    "summary": _VV40["summary"],
    "entities": _VV40["entities"],
    "validation_results": _VV40["validation_results"],
    "factors": [{"factor_type": "Mesh quality", "category": "Custom",
                 "required_level": 3, "achieved_level": 3, "status": "assessed"}],
    "decision": _VV40["decision"],
}

_err_no_requirement = copy.deepcopy(_VV40)
_err_no_requirement["entities"] = [
    {"type": "Model", "name": "model", "uri": "https://example.org/model/1"},
]

_err_invalid_evidence = copy.deepcopy(_VV40)
_err_invalid_evidence["validation_results"] = [
    {"name": "Bad", "evidence_type": "CustomType"},
]

_err_not_xlsx = {"_raw_text": "This is not an Excel file.\n"}


# --- Weakener mutation tests ---

# W-AL-01: remove has_uq entirely (not "No", absent) from one VR
_missing_uq = copy.deepcopy(_VV40)
_missing_uq["validation_results"][0]["has_uq"] = None  # absent → W-AL-01

# W-AR-01: remove acceptance_criteria from one assessed factor with requiredLevel
_no_criteria = _mutate_factor(_VV40, 0, "acceptance_criteria", None)

# W-AR-02: achieved < required + Accepted
_contradictory = copy.deepcopy(_VV40)
_contradictory["factors"][0]["achieved_level"] = 1
_contradictory["factors"][0]["required_level"] = 3

# W-AR-05: remove compares_to from one VR
_no_comparator = _mutate_vr(_VV40, 0, "compares_to", None)

# W-EP-04: one factor not-assessed at MRL 3
_not_assessed = copy.deepcopy(_VV40)
_not_assessed["summary"]["model_risk_level"] = "MRL 3"
_not_assessed["factors"][0]["status"] = "not-assessed"
_not_assessed["factors"][0]["required_level"] = None
_not_assessed["factors"][0]["achieved_level"] = None
_not_assessed["factors"][0]["acceptance_criteria"] = None

# Negative: scoped-out should NOT fire W-EP-04
_scoped_out = copy.deepcopy(_VV40)
_scoped_out["summary"]["model_risk_level"] = "MRL 3"
_scoped_out["factors"][0]["status"] = "scoped-out"
_scoped_out["factors"][0]["required_level"] = None
_scoped_out["factors"][0]["achieved_level"] = None
_scoped_out["factors"][0]["acceptance_criteria"] = None

# Negative: Not accepted should NOT fire W-AR-02
_not_accepted = copy.deepcopy(_VV40)
_not_accepted["factors"][0]["achieved_level"] = 1
_not_accepted["factors"][0]["required_level"] = 3
_not_accepted["decision"]["outcome"] = "Not accepted"

# Multi-gap: W-AL-01 + W-AR-01 + W-AR-02 + compounds
_multi_gap = copy.deepcopy(_VV40)
_multi_gap["summary"]["assurance_level"] = "Medium"  # enables COMPOUND-03
_multi_gap["validation_results"][0]["has_uq"] = None  # W-AL-01
_multi_gap["factors"][0]["acceptance_criteria"] = None  # W-AR-01 (Critical)
_multi_gap["factors"][1]["achieved_level"] = 1
_multi_gap["factors"][1]["required_level"] = 3  # W-AR-02 (Critical)

# NASA-specific mutations
_nasa_no_review = copy.deepcopy(_NASA)
# Find dev technical review and remove evidence
for f in _nasa_no_review["factors"]:
    if f["factor_type"] == "Development technical review":
        f["linked_evidence"] = None

_nasa_no_process = copy.deepcopy(_NASA)
for f in _nasa_no_process["factors"]:
    if f["factor_type"] == "Development process and product management":
        f["linked_evidence"] = None

_nasa_no_robustness = copy.deepcopy(_NASA)
for f in _nasa_no_robustness["factors"]:
    if f["factor_type"] == "Results robustness":
        f["linked_evidence"] = None

_nasa_ru_unassessed = copy.deepcopy(_NASA)
_nasa_ru_unassessed["summary"]["model_risk_level"] = "MRL 3"
for f in _nasa_ru_unassessed["factors"]:
    if f["factor_type"] == "Results uncertainty":
        f["status"] = "not-assessed"
        f["required_level"] = None
        f["achieved_level"] = None
        f["acceptance_criteria"] = None


# ── SPECS registry ────────────────────────────────────────────


SPECS = {
    # --- Structural pass cases ---
    "e2e-clean-vv40": {
        "packs": ["vv40"],
        "data": _VV40,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 0, "patterns": {}},
    },
    "e2e-clean-nasa": {
        "packs": ["nasa-7009b"],
        "data": _NASA,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 19,
        "expected_vr_count": 3,
        "expected_weakeners": {"total": 0, "patterns": {}},
    },
    "e2e-vv40-minimal": {
        "packs": ["vv40"],
        "data": _vv40_minimal,
        "expect_import": "pass",
        "expected_profile": "Minimal",
        "expected_factor_count": 0,
        "expected_vr_count": 1,
        "expected_weakeners": None,  # Minimal profile — skip weakener check
    },
    "e2e-nasa-minimal": {
        "packs": ["nasa-7009b"],
        "data": _nasa_minimal,
        "expect_import": "pass",
        "expected_profile": "Minimal",
        "expected_factor_count": 0,
        "expected_vr_count": 1,
        "expected_weakeners": None,
    },
    "e2e-vv40-mixed-status": {
        "packs": ["vv40"],
        "data": _mixed_status,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": None,  # Mixed statuses, weakener count depends on MRL
    },
    "e2e-uri-generation": {
        "packs": ["vv40"],
        "data": _uri_gen,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_id": "https://uofa.net/instances/my-test-project/thermal-analysis-cou",
        "expected_weakeners": None,
    },

    # --- Error cases ---
    "e2e-err-missing-summary": {
        "packs": ["vv40"],
        "data": _err_missing_summary,
        "expect_import": "error",
        "expect_error": "Sheet 'Assessment Summary' not found",
    },
    "e2e-err-missing-decision": {
        "packs": ["vv40"],
        "data": _err_missing_decision,
        "expect_import": "error",
        "expect_error": "Sheet 'Decision' not found",
    },
    "e2e-err-invalid-decision": {
        "packs": ["vv40"],
        "data": _err_invalid_decision,
        "expect_import": "error",
        "expect_error": "'Approved' is not a valid outcome",
    },
    "e2e-err-invalid-factor": {
        "packs": ["vv40"],
        "data": _err_invalid_factor,
        "expect_import": "error",
        "expect_error": "'Mesh quality' is not a valid V&V 40 factor type",
    },
    "e2e-err-no-requirement": {
        "packs": ["vv40"],
        "data": _err_no_requirement,
        "expect_import": "error",
        "expect_error": "must have at least one row with Entity Type = 'Requirement'",
    },
    "e2e-err-invalid-evidence": {
        "packs": ["nasa-7009b"],
        "data": _err_invalid_evidence,
        "expect_import": "error",
        "expect_error": "'CustomType' is not a valid evidence type",
    },
    "e2e-err-not-xlsx": {
        "packs": ["vv40"],
        "data": _err_not_xlsx,
        "expect_import": "error",
        "expect_error": "Cannot open workbook",
    },

    # --- Weakener mutation tests (VV40) ---
    "e2e-missing-uq": {
        "packs": ["vv40"],
        "data": _missing_uq,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 1, "patterns": {"W-AL-01": 1}},
    },
    "e2e-no-acceptance-criteria": {
        "packs": ["vv40"],
        "data": _no_criteria,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        # W-AR-01 is Critical. assuranceLevel=Low → no COMPOUND-03
        "expected_weakeners": {"total": 1, "patterns": {"W-AR-01": 1}},
    },
    "e2e-contradictory-accepted": {
        "packs": ["vv40"],
        "data": _contradictory,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        # W-AR-02 is Critical. assuranceLevel=Low → no COMPOUND-03
        "expected_weakeners": {"total": 1, "patterns": {"W-AR-02": 1}},
    },
    "e2e-no-comparator": {
        "packs": ["vv40"],
        "data": _no_comparator,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 1, "patterns": {"W-AR-05": 1}},
    },
    "e2e-not-assessed-at-risk": {
        "packs": ["vv40"],
        "data": _not_assessed,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 1, "patterns": {"W-EP-04": 1}},
    },
    "e2e-scoped-out-no-fire": {
        "packs": ["vv40"],
        "data": _scoped_out,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 0, "patterns": {}},
    },
    "e2e-not-accepted-no-ar02": {
        "packs": ["vv40"],
        "data": _not_accepted,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        "expected_weakeners": {"total": 0, "patterns": {}},
    },
    "e2e-multi-gap-vv40": {
        "packs": ["vv40"],
        "data": _multi_gap,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 13,
        "expected_vr_count": 1,
        # L1: W-AL-01(High) + W-AR-01(Critical) + W-AR-02(Critical)
        # Compounds: COMPOUND-01 fires on each (Crit, High) pair:
        #   W-AR-01 × W-AL-01, W-AR-02 × W-AL-01 = 2
        # COMPOUND-03: 2 Critical pids at "Medium" → 2
        # Total: 3 L1 + 2 COMPOUND-01 + 2 COMPOUND-03 = 7
        "expected_weakeners": {
            "total": 7,
            "patterns": {
                "W-AL-01": 1, "W-AR-01": 1, "W-AR-02": 1,
                "COMPOUND-01": 2, "COMPOUND-03": 2,
            },
        },
    },

    # --- NASA weakener mutation tests ---
    "e2e-nasa-no-review-evidence": {
        "packs": ["nasa-7009b"],
        "data": _nasa_no_review,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 19,
        "expected_vr_count": 3,
        "expected_weakeners": {"total": 1, "patterns": {"W-NASA-02": 1}},
    },
    "e2e-nasa-no-process-evidence": {
        "packs": ["nasa-7009b"],
        "data": _nasa_no_process,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 19,
        "expected_vr_count": 3,
        "expected_weakeners": {"total": 1, "patterns": {"W-NASA-03": 1}},
    },
    "e2e-nasa-no-robustness-evidence": {
        "packs": ["nasa-7009b"],
        "data": _nasa_no_robustness,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 19,
        "expected_vr_count": 3,
        "expected_weakeners": {"total": 1, "patterns": {"W-NASA-06": 1}},
    },
    "e2e-nasa-ru-unassessed": {
        "packs": ["nasa-7009b"],
        "data": _nasa_ru_unassessed,
        "expect_import": "pass",
        "expected_profile": "Complete",
        "expected_factor_count": 19,
        "expected_vr_count": 3,
        "expected_weakeners": {"total": 1, "patterns": {"W-EP-04": 1}},
    },
}


# ── CLI for manual regeneration ───────────────────────────────

if __name__ == "__main__":
    out = Path(__file__).parent
    for name, spec in sorted(SPECS.items()):
        p = out / f"{name}.xlsx"
        generate_fixture(spec["data"], p)
        print(f"  {name}: {p.name}")
    print(f"\nGenerated {len(SPECS)} fixtures in {out}")
