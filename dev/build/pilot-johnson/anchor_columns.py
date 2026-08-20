#!/usr/bin/env python3
"""Add a `Source Anchor` column to every sheet of an extracted workbook.

Spec §2.1 requires every populated cell to carry a citation anchor to a page,
section, table or figure of the source, and tells the encoder to add a home for
it if the template has none. The nasa-7009b template has none, so this adds one.

Two properties make this safe to run against a live workbook:

- **It appends.** The anchor lands in the first free column to the right of the
  existing headers. `excel_reader` reads every sheet by fixed column index
  (`_read_factors` takes A-H and ignores the rest), so nothing downstream sees a
  shifted column.
- **It is idempotent.** A sheet that already has the column is left alone.

The consequence is the finding, not a side effect: because the reader ignores the
column, **the anchors do not survive import.** They live in the workbook, which is
where humans review, and they are absent from the JSON-LD, which is what a
reviewer downstream actually receives. That is finding F-2b.

Run:  python dev/build/pilot-johnson/anchor_columns.py <workbook.xlsx>
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

import openpyxl

HEADER = "Source Anchor"
DESCRIPTION = "Page/section/table of the source, plus method where non-textual"

# sheet name -> a header string that identifies the header row on that sheet.
ANCHORS = {
    "Assessment Summary": "Project Name",
    "Model & Data": "Entity Type",
    "Validation Results": "Result Name",
    "Credibility Factors": "Factor Type",
    "Decision": "Decision Outcome",
}


def _header_row(ws, needle: str) -> int:
    for row in range(1, 12):
        for col in range(1, 20):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.strip() == needle:
                return row
    raise SystemExit(f"sheet {ws.title!r}: no header row containing {needle!r}")


def _last_header_col(ws, row: int) -> int:
    last = 0
    for col in range(1, 40):
        if ws.cell(row=row, column=col).value not in (None, ""):
            last = col
    return last


def add_anchor_columns(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path)
    touched = []
    for sheet, needle in ANCHORS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        head = _header_row(ws, needle)
        last = _last_header_col(ws, head)
        existing = [ws.cell(row=head, column=c).value for c in range(1, last + 1)]
        if HEADER in existing:
            continue
        target = last + 1
        cell = ws.cell(row=head, column=target, value=HEADER)
        cell._style = copy(ws.cell(row=head, column=last)._style)
        # The template puts a description row directly under every header row.
        desc = ws.cell(row=head + 1, column=target, value=DESCRIPTION)
        desc._style = copy(ws.cell(row=head + 1, column=last)._style)
        ws.column_dimensions[cell.column_letter].width = 52
        touched.append(f"{sheet} -> column {cell.column_letter}")
    wb.save(path)
    return touched


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit(__doc__.strip().splitlines()[-1])
    for line in add_anchor_columns(Path(sys.argv[1])) or ["nothing to do"]:
        print(line)
