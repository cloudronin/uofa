#!/usr/bin/env python3
"""The review pass, as data.

Spec §2.1-§2.2: every populated cell carries a citation anchor to a page, section,
table or figure of the source, and no cell passes on extractor confidence. A cell
is CONFIRM (extractor value checked against the source and kept), CORRECT (checked
and replaced), or BLANK (the source does not carry it, so it is emptied and
listed). There is no fourth option and nothing is left as-returned.

Encoding the pass as a table rather than performing it by hand in Excel is the
point: the table IS the record of what was reviewed against what, it re-runs
against a fresh extract, and a reader can diff it. `REVIEW_LEDGER.md` is generated
from it.

Run:  python dev/build/pilot-johnson/review_pass.py
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
RAW = HERE / "raw-extract" / "johnson-extracted-RAW.xlsx"
OUT = HERE / "johnson-extracted.xlsx"

CONFIRM, CORRECT, BLANK = "CONFIRM", "CORRECT", "BLANK"

# (sheet, cell, action, new value or None, anchor, note)
# Anchors are PDF page numbers; the document carries no printed folios.
REVIEW: list[tuple] = [
    # ── Assessment Summary ────────────────────────────────────────────────
    ("Assessment Summary", "A3", CONFIRM, None, "p.5, worked-example description", ""),
    ("Assessment Summary", "B3", CONFIRM, None, "p.14 LCW [M&S 8](2)", ""),
    ("Assessment Summary", "C3", CONFIRM, None, "p.5; p.11 [M&S 6]; p.14 [M&S 10]", ""),
    ("Assessment Summary", "D3", CONFIRM, None,
     "n/a — package-level declaration, derived at import from content", ""),
    ("Assessment Summary", "E3", BLANK, None, "source-absent",
     "Extractor wrote 'N/A', which is not a member of the controlled list. "
     "The field is an FDA/domain category with no aerospace meaning here."),
    ("Assessment Summary", "F3", BLANK, None, "source-absent",
     "Extractor wrote 'MRL 3'. No MRL appears anywhere; it is a V&V 40 construct "
     "absent from a 7009A paper. Criticality evidence lives in C3 instead (p.11)."),
    ("Assessment Summary", "G3", BLANK, None, "source-absent",
     "Extractor wrote 'Medium'. Not stated."),
    ("Assessment Summary", "H3", CORRECT, "NASA-STD-7009A", "p.1 title; p.3 §'THE PROBLEM 7009A ADDRESSES'",
     "Extractor wrote NASA-STD-7009B. The paper applies 7009A throughout. "
     "Note resolve_criteria_set('NASA-STD-7009') also folds to the B identifier; "
     "the literal 'A' string is required to avoid that."),
    ("Assessment Summary", "I3", CONFIRM, None, "p.1 byline", ""),
    ("Assessment Summary", "J3", BLANK, None, "source-absent",
     "Extractor left the template placeholder text in the data row. The report "
     "carries no internal date; NTRS catalogue metadata was not admitted as source."),
    ("Assessment Summary", "K3", CORRECT, "NTRS 20200002832 — Johnson, K.L. (2020), NASA NESC",
     "p.1", "Extractor wrote the local filename."),
    ("Assessment Summary", "L3", CONFIRM, None, "p.22 [M&S 29]; p.9 [M&S 33](1)", ""),

    # ── Model & Data ──────────────────────────────────────────────────────
    ("Model & Data", "B3", CONFIRM, None, "p.14 [M&S 10]; p.16 [M&S 8]", ""),
    ("Model & Data", "D3", CONFIRM, None, "p.14 [M&S 10]; p.16 [M&S 8]", ""),
    ("Model & Data", "E3", BLANK, None, "source-absent", "Template placeholder left by the extractor."),
    ("Model & Data", "F3", BLANK, None, "source-absent", "Template placeholder left by the extractor."),
    ("Model & Data", "B4", CONFIRM, None, "p.15 [M&S 12]; p.16 4.2.2 m", ""),
    ("Model & Data", "D4", CONFIRM, None, "p.15 [M&S 12]; p.16 4.2.2 m; p.20 4.1.2 c", ""),
    ("Model & Data", "B5", CONFIRM, None, "p.12 4.1.2 a; p.12 4.1.3 a,b", ""),
    ("Model & Data", "D5", CONFIRM, None, "p.12 4.1.3 a,b; p.14 note; p.15 [M&S 11]", ""),

    # ── Validation Results ────────────────────────────────────────────────
    ("Validation Results", "A3", CONFIRM, None, "p.18 4.1.3 b(3); p.18 [M&S 15]", ""),
    ("Validation Results", "D3", CONFIRM, None, "p.18 4.1.3 b(3); p.18 [M&S 15]", ""),
    ("Validation Results", "C3", BLANK, None, "source-absent", "Template placeholder."),
    ("Validation Results", "E3", CONFIRM, None, "p.18 4.1.3 b(3)", ""),
    ("Validation Results", "G3", BLANK, None, "source-absent", "Template placeholder."),
    ("Validation Results", "H3", BLANK, None, "source-absent", "Template placeholder."),
    ("Validation Results", "A4", CONFIRM, None, "p.21 [M&S 28](3); p.22 M&S Results", ""),
    ("Validation Results", "D4", CONFIRM, None, "p.21 [M&S 28](3); p.25 Validation rationale", ""),
    ("Validation Results", "H4", CONFIRM, None, "p.22 M&S Results; p.9 [M&S 33](1); p.14 [M&S 10]", ""),
    ("Validation Results", "A5", CONFIRM, None, "p.17 [M&S 17]", ""),
    ("Validation Results", "D5", CONFIRM, None, "p.17 [M&S 17]", ""),
    ("Validation Results", "A6", CONFIRM, None, "p.10 [M&S 36]; p.24 [M&S 36]", ""),
    ("Validation Results", "D6", CONFIRM, None, "p.10 [M&S 36]; p.24 [M&S 36]", ""),
    ("Validation Results", "A7", CONFIRM, None, "p.6; p.19 4.1.2 c; p.20 commentary", ""),
    ("Validation Results", "D7", CONFIRM, None, "p.6; p.8 [M&S 32](7); p.19 4.1.2 c",
     "Retained as the extractor wrote it, but see AMBIGUITY_LOG A-17: p.8 both "
     "describes this waiver and says 'No waivers were required', and p.23 answers "
     "the same requirement 'None'."),
]

# ── Credibility Factors ───────────────────────────────────────────────────
# Levels are carried ONLY within the 0-4 NASA block (AMBIGUITY_LOG A-06). No
# 7009A level is rewritten onto a 1-5 V&V 40 factor, so the V&V 40 side carries
# evidence — status, criteria, rationale, anchor — and no level. That is not an
# omission; it is what a 7009A source can honestly say about a V&V 40 factor.
#
# row: (factor, required, achieved, status, anchor, note)
FACTOR_ROWS = {
    5:  ("Software quality assurance", BLANK, BLANK, "not-assessed",
         "p.6; p.18 4.1.2 a",
         "Extractor wrote 4/4. Analysis-code verification was WAIVED (p.6) and the "
         "verification-practices question is answered 'n/a' (p.18). Software quality "
         "is not separately assessed in a 7009A assessment."),
    6:  ("Numerical code verification", BLANK, BLANK, "assessed",
         "p.18 4.1.3 b(3); p.18 [M&S 15]",
         "Directly answered: duplicate analysis in competitive software returned an "
         "identical model. Evidence kept, level dropped — 7009A rates Verification as "
         "one 0-4 factor (achieved 4, p.25) and that level does not transfer to a "
         "1-5 sub-factor (A-06)."),
    7:  ("Discretization error", BLANK, BLANK, "not-applicable",
         "p.15 [M&S 12]; p.16 4.2.2 m",
         "Extractor said not-assessed; not-applicable is the better fit. A linear "
         "regression on test data has no discretization scheme to converge."),
    8:  ("Numerical solver error", BLANK, BLANK, "scoped-out",
         "p.18 4.2.4 a",
         "Directly answered: solution verification 'No. Not required; unlikely a "
         "problem due to simplicity of both model and analysis technique'. "
         "Extractor's status confirmed; its blank levels confirmed."),
    9:  ("Use error", BLANK, BLANK, "not-assessed",
         "p.16 [M&S 8]",
         "Extractor wrote 3/3. Independent inspection of the modelling process is "
         "real evidence but 7009A assesses no separate use-error factor."),
    10: ("Model form", BLANK, BLANK, "assessed",
         "p.17 [M&S 17]",
         "Directly answered by the conceptual-validation block. Level dropped per A-06."),
    11: ("Model inputs", BLANK, BLANK, "not-assessed",
         "p.16 [M&S 19]; p.16 [M&S 28](1)",
         "Extractor wrote 3/3. This is where 7009A's Input Pedigree would have gone, "
         "and it is a different question. See ESCALATION A-07."),
    12: ("Test samples", BLANK, BLANK, "not-assessed",
         "p.12 4.1.2 a; p.14 note; p.15 [M&S 11]",
         "Extractor wrote 3/3. 18 planned / 15 completed is stated, but 7009A rates "
         "no comparator-sample factor."),
    13: ("Test conditions", BLANK, BLANK, "not-assessed",
         "p.14 [M&S 10]; p.16 [M&S 19]",
         "Extractor wrote 3/3."),
    14: ("Equivalency of input parameters", BLANK, BLANK, "not-assessed",
         "p.18 [M&S 16]; p.19 [M&S 18]",
         "Extractor wrote 3/3."),
    15: ("Output comparison", BLANK, BLANK, "assessed",
         "p.19 4.1.2 a/4.1.3 b(3)/4.2.6 a; p.25 Validation rationale",
         "Directly answered by the empirical-validation block. Level dropped per A-06 "
         "— and note 7009A rates Validation at 1, which is the opposite direction "
         "from the extractor's 4. See RAW_EXTRACT_DELTA §2."),
    16: ("Relevance of the quantities of interest", BLANK, BLANK, "not-assessed",
         "p.14 [M&S 10]; p.22 M&S Results",
         "Extractor wrote 4/4."),
    17: ("Relevance of the validation activities to the COU", BLANK, BLANK, "not-assessed",
         "p.19 4.1.2 c; p.20 commentary",
         "Extractor wrote 1/1, which is the right story on the wrong factor: 7009A's "
         "Validation level 1 is a Validation verdict, not an Applicability one."),
    # NASA-only block, 0-4, same scale as 7009A. Levels transfer.
    18: ("Data pedigree", 3, 3, "assessed",
         "req: p.7 Table 3 shading (geometric recovery, TABLE3_RECOVERY.md); ach: p.25",
         "Both levels confirmed against the source; extractor happened to agree."),
    19: ("Development technical review", BLANK, BLANK, "not-assessed",
         "p.10 [M&S 36]; p.24 [M&S 36]",
         "Extractor wrote 3/3. This factor is new in 7009B and has no 7009A "
         "counterpart, so Johnson assesses no level for it. Review content is "
         "abundant; content is not a level (A-09)."),
    20: ("Development process and product management", 2, 4, "assessed",
         "req: p.7 Table 3 shading (geometric recovery, TABLE3_RECOVERY.md); ach: p.25",
         "REQUIRED LEVEL CORRECTED 4 -> 2. The extractor's required=achieved default "
         "erased a two-level exceedance. This is the cleanest 1:1 mapping in the "
         "table and the clearest demonstration of the §3b failure."),
    21: ("Results uncertainty", 4, 4, "assessed",
         "req: p.7 Table 3 shading (Uncertainty Characterization column, geometric "
         "recovery); ach: p.25",
         "7009A calls this Uncertainty Characterization (A-01)."),
    22: ("Results robustness", 4, 4, "assessed",
         "req: p.7 Table 3 shading (geometric recovery); ach: p.25", ""),
    23: ("Use history", 3, 3, "assessed",
         "req: p.7 Table 3 shading (M&S History column, geometric recovery); ach: p.25",
         "7009A calls this M&S History (A-02). Predeclaration was negotiated; see p.7-8."),
}

DECISION = [
    ("Decision", "A3", CONFIRM, None, "p.19 [M&S 8](1),(3),(4); p.10 Summary",
     "JUDGMENT-CLASS. 'Model acceptance requirements set by the Project were met' "
     "(p.19) supports the outcome; the paper records no decision ACT."),
    ("Decision", "B3", CORRECT,
     "Source states that 'Model acceptance requirements set by the Project were met' "
     "(p.19) and that results are 'offered with the credibility level required for use "
     "of the results in the investigation's risk model' (p.10). Predicted worst-case "
     "penetration 0.098 cm with a 99%/95% tolerance bound of 0.128 cm, both below the "
     "0.16 cm critical depth (p.22, p.14). DRAFT NOTE: the paper records no decision "
     "act — no deciding party, no decision date, and a '(Signed)' line with nothing "
     "after it (p.10). Validation was waived by the investigation team and approved by "
     "the Technical Authority (p.8), which is an approval of a waiver and not an "
     "acceptance of the model; the source contradicts itself on whether a waiver "
     "existed at all (p.8 vs p.23, see AMBIGUITY_LOG A-17).",
     "p.19; p.10; p.22; p.8; p.23",
     "Extractor's rationale asserted a decision by 'Investigation team / Technical "
     "Authority'. Rewritten to state what the source says and what it does not."),
    ("Decision", "C3", BLANK, None, "n/a — auto-filled at import", "Template placeholder."),
    ("Decision", "D3", BLANK, None, "source-absent",
     "Extractor wrote 'Investigation team / Technical Authority'. The TA approved the "
     "validation waiver (p.8), not the model."),
    ("Decision", "E3", BLANK, None, "source-absent",
     "Extractor wrote 2019-10-14, which is the MODEL RELEASE date (p.20), not a "
     "decision date."),
]

ANCHOR_HEADER = "Source Anchor"


def _anchor_col(ws) -> int:
    for col in range(1, 40):
        for row in range(1, 12):
            if ws.cell(row=row, column=col).value == ANCHOR_HEADER:
                return col
    raise SystemExit(f"{ws.title}: anchor column missing — run anchor_columns.py first")


def main() -> None:
    import anchor_columns

    OUT.write_bytes(RAW.read_bytes())
    anchor_columns.add_anchor_columns(OUT)
    wb = openpyxl.load_workbook(OUT)

    tally = {CONFIRM: 0, CORRECT: 0, BLANK: 0}
    ledger: list[tuple] = []

    def apply(sheet, cell, action, value, anchor, note):
        ws = wb[sheet]
        before = ws[cell].value
        if action == BLANK:
            ws[cell] = None
        elif action == CORRECT:
            ws[cell] = value
        tally[action] += 1
        ledger.append((sheet, cell, action, before, ws[cell].value, anchor, note))
        if action != BLANK:
            row = int("".join(c for c in cell if c.isdigit()))
            acol = _anchor_col(ws)
            existing = ws.cell(row=row, column=acol).value
            if anchor not in (existing or ""):
                ws.cell(row=row, column=acol,
                        value=f"{existing}; {anchor}" if existing else anchor)

    for entry in REVIEW + DECISION:
        apply(*entry)

    # Credibility Factors: levels, status, and a per-row anchor.
    ws = wb["Credibility Factors"]
    acol = _anchor_col(ws)
    for row, (factor, req, ach, status, anchor, note) in FACTOR_ROWS.items():
        assert ws.cell(row, 1).value == factor, (row, ws.cell(row, 1).value, factor)
        for col, val in ((3, req), (4, ach)):
            before = ws.cell(row, col).value
            if val is BLANK:
                # openpyxl's ws.cell(r, c, value) IGNORES a None value -- it does
                # not clear the cell. Assign through .value or the blank silently
                # does nothing and the extractor's number survives into import.
                ws.cell(row, col).value = None
                action = BLANK if before is not None else CONFIRM
            else:
                action = CONFIRM if before == val else CORRECT
                ws.cell(row, col).value = val
            tally[action] += 1
            ledger.append(("Credibility Factors", f"{'CD'[col-3]}{row}", action,
                           before, ws.cell(row, col).value, anchor, note))
        before = ws.cell(row, 7).value
        action = CONFIRM if before == status else CORRECT
        ws.cell(row, 7).value = status
        tally[action] += 1
        ledger.append(("Credibility Factors", f"G{row}", action, before, status, anchor, note))
        ws.cell(row, acol).value = anchor

    wb.save(OUT)

    lines = ["# Review ledger — every cell decision, generated by review_pass.py",
             "",
             "State: DRAFT. Spec §2.2 — every cell is CONFIRM (checked against the source",
             "and kept), CORRECT (checked and replaced), or BLANK (source-absent, emptied",
             "and listed). No cell passes on extractor confidence.",
             "",
             f"**{tally[CONFIRM]} confirmed · {tally[CORRECT]} corrected · {tally[BLANK]} blanked "
             f"(source-absent) · {sum(tally.values())} decisions total**",
             "",
             "| Sheet | Cell | Action | Extractor value | Reviewed value | Anchor | Note |",
             "|---|---|---|---|---|---|---|"]
    for sheet, cell, action, before, after, anchor, note in ledger:
        def clean(v):
            if v is None:
                return "_(blank)_"
            s = str(v).replace("|", "\\|").replace("\n", " ")
            return s if len(s) <= 90 else s[:87] + "…"
        lines.append(f"| {sheet} | {cell} | **{action}** | {clean(before)} | {clean(after)} | "
                     f"{clean(anchor)} | {clean(note) if note else ''} |")
    (HERE / "REVIEW_LEDGER.md").write_text("\n".join(lines) + "\n")

    print(f"confirmed {tally[CONFIRM]}   corrected {tally[CORRECT]}   "
          f"blanked {tally[BLANK]}   total {sum(tally.values())}")
    print(f"wrote {OUT} and REVIEW_LEDGER.md")


if __name__ == "__main__":
    main()
