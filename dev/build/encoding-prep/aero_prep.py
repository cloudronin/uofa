#!/usr/bin/env python3
"""Aero encoding prep, both COUs. W4 of the prep spec.

Prepares, it does not review. Section 0 of the spec reserves the cell walk for the
author, so this adds the Source Anchor column and populates candidate anchors at row
level, snapshots the raw extractor output, and imports. Every anchor it writes is a
candidate the author confirms during their walk.

Candidate anchors are authored from the bundle manifest rather than derived from
extractor provenance, because the extractor does not record a per-cell source. The
on-ramp page says "Hover a cell for the document it came from"; the cell comments
carry only a confidence percentage. That gap is why anchoring is a review-pass
product and not a byproduct of extraction.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "pilot-johnson"))
import anchor_columns  # noqa: E402

import openpyxl  # noqa: E402

HERE = Path(__file__).parent

# Sheet -> the bundle file(s) whose content that sheet encodes, per EVIDENCE_MANIFEST.
CANDIDATE_ANCHORS = {
    "Assessment Summary": "cou_definition.docx; risk_assessment_memo.pdf",
    "Model & Data": "cou_definition.docx; cfx_solver_settings.txt; EVIDENCE_MANIFEST.txt",
    "Validation Results": "credibility_assessment_narrative.docx; the study CSVs named in EVIDENCE_MANIFEST.txt",
    "Credibility Factors": "credibility_assessment_narrative.docx",
    "Decision": "decision_rationale.pdf; review_board_minutes",
}

CANDIDATE_NOTE = " [CANDIDATE, author confirms at the cell walk]"


def prep(cou: str) -> None:
    d = HERE / f"aero-{cou}"
    wb_path = d / f"aero-{cou}-extracted.xlsx"
    raw_dir = d / "raw-extract"
    raw_dir.mkdir(exist_ok=True)
    raw = raw_dir / f"aero-{cou}-extracted-RAW.xlsx"
    if not raw.exists():
        shutil.copy2(wb_path, raw)

    # Readable dump of the raw output, so the review-pass delta stays measurable.
    dump = raw_dir / "RAW_EXTRACT_DUMP.txt"
    book = openpyxl.load_workbook(raw)
    lines = [f"Raw extractor output for aero-{cou}, verbatim, before any edit.",
             "Model: anthropic/claude-sonnet-5   Pack: nasa-7009b 0.5.0", ""]
    for ws in book.worksheets:
        lines += ["=" * 74, f"SHEET: {ws.title}", "=" * 74]
        for row in ws.iter_rows():
            cells = [(c.coordinate, str(c.value).replace("\n", "\\n"))
                     for c in row if c.value not in (None, "")]
            if cells:
                lines += [f"  {k:>5}  {v}" for k, v in cells] + [""]
    dump.write_text("\n".join(lines), encoding="utf8")

    anchor_columns.add_anchor_columns(wb_path)
    book = openpyxl.load_workbook(wb_path)
    filled = 0
    for sheet, anchor in CANDIDATE_ANCHORS.items():
        if sheet not in book.sheetnames:
            continue
        ws = book[sheet]
        acol = None
        head = None
        for col in range(1, 40):
            for row in range(1, 12):
                if ws.cell(row=row, column=col).value == anchor_columns.HEADER:
                    acol, head = col, row
        if acol is None:
            continue
        for row in range(head + 2, ws.max_row + 1):
            if any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, acol)):
                ws.cell(row=row, column=acol).value = anchor + CANDIDATE_NOTE
                filled += 1
    book.save(wb_path)
    print(f"{cou}: raw snapshot + dump written, {filled} candidate anchors")


if __name__ == "__main__":
    for c in ("cou1", "cou2"):
        prep(c)
