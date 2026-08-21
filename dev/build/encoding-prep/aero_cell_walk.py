"""Aero COU1/COU2 cell walk — mechanical resolution against pre-registered ground truth.

Run from the repo root:

    python dev/build/encoding-prep/aero_cell_walk.py

For each package this reads the raw extractor workbook, applies the corrections the
walk established, resolves every CANDIDATE anchor, and writes the reviewed workbook
plus `REVIEW_LEDGER.md`.

Three things this script is careful about, each of them a finding in its own right:

1. **The ground truth is the map, the source is the territory.** Where the workbook and
   the GT disagree, the correction is taken from the named source document and the
   ledger cites the source. Where the GT is itself wrong against its own named source,
   the row is marked GT-DEFECT and the workbook stands. See COU1 row 8.

2. **Tolerance applies to achieved, never to required.** A GT `level_tolerance` is
   extraction latitude on the achieved level. Comparing required under the same
   tolerance hides a required level that was defaulted to achieved, which is the A-7
   failure mode exactly.

3. **A stated gap may lose its home; it may not lose its trace.** Where un-merging a
   factor evicts a gap the pack cannot express, the row's rationale carries a visible
   declination pointing at the log entry.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[2]
GT_DIR = ROOT / "tests/fixtures/extract/ground_truth"

CONFIRM, CORRECT, BLANK = "CONFIRM", "CORRECT", "BLANK"

# Factor row -> narrative section. Both bundles use the same section structure.
SECTION = {
    5: "1.1", 6: "1.2", 7: "1.3", 8: "1.4", 9: "1.5",
    10: "2.1", 11: "2.2", 12: "2.3", 13: "2.4",
    14: "3.1", 15: "3.2",
    16: "4.1", 17: "4.2",
    18: "5.1", 19: "5.2", 20: "5.3", 21: "5.4", 22: "5.5", 23: "5.6",
}

NARRATIVE = "credibility_assessment_narrative.docx"

# Where the walk established a fuller citation than "the section for this factor",
# keep it. Keyed (package, section).
ANCHOR_EXTRA = {
    ("cou1", "4.2"): "; §6 open item 4",
    ("cou2", "4.2"): "; §6 open item 5",
}
cou_ctx = [""]

# Anchors for the sheets the ground truth says nothing about. The GT carries a
# source_file per *factor*; Model & Data and Validation Results rows have no GT
# counterpart, so each is resolved by opening the bundle and naming the document that
# actually carries the row. Keyed (sheet, row) -> (anchor, why).
OTHER_ANCHORS = {
    "cou1": {
        ("Model & Data", 4): ("cfx_solver_settings.txt; cou_definition.docx",
            "Model row. The solver configuration file carries the model definition -- code, "
            "turbulence model, mesh, materials -- and cou_definition.docx carries the context "
            "it is built for. EVIDENCE_MANIFEST.txt dropped: a manifest lists documents, it "
            "does not carry a value."),
        ("Model & Data", 5): ("cascade_rig_temperature_data.csv",
            "Dataset row, 48-point thermocouple rake. The CSV is the dataset."),
        ("Model & Data", 6): ("mesh_convergence_study.csv",
            "Dataset row, grid convergence at mid-span. The CSV is the dataset."),
        ("Model & Data", 7): ("sensitivity_study_turbulence_intensity.csv",
            "Dataset row, parametric sensitivity results. The CSV is the dataset."),
        ("Validation Results", 4): (
            "mesh_convergence_study.csv (primary evidence); "
            "credibility_assessment_narrative.docx §1.3 (assessment)",
            "Dual: the CSV carries the GCI numbers, §1.3 carries the assessment drawn from them."),
        ("Validation Results", 5): (
            "cascade_rig_temperature_data.csv (primary evidence); "
            "credibility_assessment_narrative.docx §3.2 (assessment)",
            "Dual: the CSV carries the 48-station comparison, §3.2 the output-comparison assessment."),
        ("Validation Results", 6): (
            "credibility_assessment_narrative.docx §5.4",
            "Single anchor, deliberately. COU1's bundle carries no UQ CSV; the Monte Carlo run "
            "is described only in §5.4, which is also why the package records it Inconclusive."),
        ("Validation Results", 7): (
            "sensitivity_study_turbulence_intensity.csv (primary evidence); "
            "credibility_assessment_narrative.docx §5.5 (assessment)",
            "Dual: the CSV carries the parametric sweep, §5.5 the robustness assessment."),
    },
    "cou2": {
        ("Model & Data", 4): ("cfx_solver_settings.txt; cou_definition.docx",
            "Model row, as COU1."),
        ("Model & Data", 5): ("cascade_reuse_traceability.txt",
            "Dataset row. The dataset itself (cascade_rig_temperature_data.csv) lives in COU1's "
            "bundle, not this one. The in-bundle document that carries its use here is the "
            "re-use traceability record TRC-CRUISE-VAL-001, which names the source dataset "
            "explicitly and states the operating-point mismatch it entails."),
        ("Model & Data", 6): ("cruise_uq_study.csv", "Dataset row. The CSV is the dataset."),
        ("Model & Data", 7): ("sensitivity_study_cruise.csv", "Dataset row. The CSV is the dataset."),
        ("Validation Results", 4): (
            "cascade_reuse_traceability.txt; credibility_assessment_narrative.docx §1.3",
            "Inherited from COU1; the traceability record is what makes the inheritance "
            "auditable from inside this bundle."),
        ("Validation Results", 5): (
            "cascade_reuse_traceability.txt (primary evidence); "
            "credibility_assessment_narrative.docx §4.2 (assessment)",
            "Dual, and the most load-bearing anchor in the package: the traceability record "
            "documents the re-use and the regime mismatch, §4.2 assesses it as not relevant "
            "to the cruise COU."),
        ("Validation Results", 6): (
            "cruise_uq_study.csv (primary evidence); "
            "credibility_assessment_narrative.docx §5.4 (assessment)",
            "Dual: the CSV carries the 1500-sample propagation, §5.4 the assessment."),
        ("Validation Results", 7): (
            "sensitivity_study_cruise.csv (primary evidence); "
            "credibility_assessment_narrative.docx §5.5 (assessment)",
            "Dual: the CSV carries the sweep, §5.5 the robustness assessment."),
    },
}

# Corrections established by the walk, per package. Each is (sheet, cell, action,
# value, anchor, note). Anchors cite the source, never the ground truth.
CORRECTIONS = {
    "cou1": [
        ("Validation Results", "C3", BLANK, None, "source-absent",
         "TEMPLATE PLACEHOLDER. The cell held the template's own hint text, 'Stable URI "
         "or local ID', which reached the package as the node's identifier. protocol-check "
         "reported 'no template placeholder text in data rows: clean' because its scan "
         "starts at head+data_offset (row 4) and the extractor wrote its first data row "
         "into the template's hint row (row 3). Filed as a tooling finding."),
        ("Credibility Factors", "D17", CORRECT, 2,
         f"{NARRATIVE} §4.2; §6 open item 4",
         "UN-MERGED. The extractor merged narrative §4.3 Film Cooling Validation into this "
         "factor and took its more severe L1. The source rates THIS factor at Level 2: §4.2 "
         "'Assessment: Level 2', and §6 open item 4 'Factor 4.2 Relevance of Validation: L2 "
         "achieved vs L3 required'. Corrected to the source. GT agrees (expected_level 2); "
         "GT mapping_notes permit either reading, and the source settles it. The evicted "
         "film-cooling gap is logged as G-06 and filed as a schema finding."),
    ],
    "cou2": [
        ("Validation Results", "C3", BLANK, None, "source-absent",
         "TEMPLATE PLACEHOLDER, same defect as COU1. See the tooling finding."),
    ],
}

# Rationale additions -- the visible declination required where a stated gap loses its
# home. Appended rather than replacing what the extractor wrote.
RATIONALE_NOTES = {
    "cou1": [
        ("Credibility Factors", 6, 17,
         " DECLINED MAPPING: the source also states a Film Cooling Validation gap "
         "(narrative §4.3, Level 1 achieved against Level 3 required, board-flagged as a "
         "material applicability gap). NASA-STD-7009B has no factor for it, so it is not "
         "carried as a level here. See AMBIGUITY_LOG G-06.",
         f"{NARRATIVE} §4.3",
         "Fail-loud declination. A stated gap may lose its home; it may not lose its trace."),
    ],
    "cou2": [
        ("Credibility Factors", 6, 17,
         " DECLINED MAPPING: the source states a Film Cooling Validation gap at narrative "
         "§4.3 (Level 1 against Level 3 required), scoped by the source itself as inherited "
         "from COU1 rather than a COU2-specific finding. No pack factor expresses it. "
         "See AMBIGUITY_LOG G-06.",
         f"{NARRATIVE} §4.3",
         "Fail-loud declination, inherited-gap form."),
    ],
}

# Rows where the ground truth is wrong against its own named source. The workbook
# stands; the ledger records why.
GT_DEFECTS = {
    "cou1": {
        8: ("GT expected_required_level 2 against a source that states no required level. "
            "cfx_solver_settings.txt, the GT's own named source_file, carries no required "
            "level at all. Narrative §1.4 states only 'Assessment: Level 1', while §1.3 and "
            "§4.3 -- the two real shortfalls -- both state 'Achieved Level N against Required "
            "Level M' explicitly. Narrative §6: 'Verification factors: 4 of 5 meet required "
            "level; Discretization Error achieves L1 of required L3', which only holds if "
            "solver error meets its required level. The §6 gaps list excludes it. GT's own "
            "W-AR-02 rationale concedes 'Numerical solver (L1 with implicit L2 required)' -- "
            "an inference, not a statement, and an inferred required level is what A-7 exists "
            "to catch. RULED 2026-08-21: source wins, workbook req 1 / ach 1 stands.")
    },
    "cou2": {},
}


def gt_for(cou: str) -> dict:
    return json.loads((GT_DIR / f"aero-{cou}-nasa7009b.json").read_text())


def anchor_col(ws) -> int:
    for col in range(1, 40):
        for row in range(1, 6):
            if ws.cell(row=row, column=col).value == "Source Anchor":
                return col
    raise SystemExit(f"{ws.title}: no Source Anchor column")


def resolve_anchor(existing: str, gt_src: str, section: str) -> tuple[str, str]:
    """Return (new anchor, why). Both halves genuinely carry it, so both are cited."""
    if existing and "[CANDIDATE" not in existing:
        return existing, "already resolved; left as it stands"
    base = (existing or "").split("[CANDIDATE")[0].strip().rstrip(";").strip()
    if not gt_src:
        return f"{NARRATIVE} §{section}", "no GT source_file; anchored to the narrative section"
    if gt_src in base:
        extra = ANCHOR_EXTRA.get((cou_ctx[0], section), "")
        return (f"{gt_src} §{section}{extra}" if gt_src == NARRATIVE
                else f"{gt_src}; {NARRATIVE} §{section}",
                "candidate agreed with the pre-registered source_file; promoted and "
                "narrowed to the section carrying the level")
    return (f"{gt_src} (primary evidence); {NARRATIVE} §{section} (assessed level)",
            f"candidate named {NARRATIVE} only; the GT names {gt_src}. Both carry part of "
            "it -- the data file the measurement, the narrative the assessed level -- so "
            "the anchor is dual, on the Johnson dual-anchor precedent")


def walk(cou: str) -> None:
    cou_ctx[0] = cou
    pkg = HERE / f"aero-{cou}"
    out = pkg / f"aero-{cou}-extracted.xlsx"

    # The walk builds on the PREP workbook, not on raw-extract. raw-extract has no
    # Source Anchor column at all -- the prep session added it and authored the
    # candidate anchors -- so regenerating from raw would discard the very column
    # this walk exists to resolve. The pre-walk state is snapshotted once, so the
    # walk's own changes stay measurable the way raw-extract makes the prep's changes
    # measurable.
    snap_dir = pkg / "pre-walk"
    snap_dir.mkdir(exist_ok=True)
    snap = snap_dir / f"aero-{cou}-extracted-PREWALK.xlsx"
    if not snap.exists():
        snap.write_bytes(out.read_bytes())
    # Restore before applying, every run. Without this the walk edits its own output
    # and a second run reports its own results as the "before" state -- the ledger
    # would then describe the walk against itself rather than against the prep.
    out.write_bytes(snap.read_bytes())

    wb = openpyxl.load_workbook(out)
    gt = gt_for(cou)
    gt_src = {f["factor_type"]: f.get("source_file", "") for f in gt["expected_factors"]}
    gt_lvl = {f["factor_type"]: f for f in gt["expected_factors"]}

    tally = {CONFIRM: 0, CORRECT: 0, BLANK: 0}
    ledger: list[tuple] = []

    def record(sheet, cell, action, before, after, anchor, note):
        tally[action] = tally.get(action, 0) + 1
        ledger.append((sheet, cell, action, before, after, anchor, note))

    # 1. Corrections.
    for sheet, cell, action, value, anchor, note in CORRECTIONS[cou]:
        ws = wb[sheet]
        before = ws[cell].value
        if action == BLANK:
            ws[cell].value = None
        elif action == CORRECT:
            ws[cell].value = value
        record(sheet, cell, action, before, ws[cell].value, anchor, note)

    # 2. Fail-loud rationale declinations.
    for sheet, col, row, text, anchor, note in RATIONALE_NOTES[cou]:
        ws = wb[sheet]
        before = ws.cell(row, col).value
        if text.strip() not in (before or ""):
            ws.cell(row, col).value = (before or "").rstrip() + text
        record(sheet, f"{chr(64 + col)}{row}", CORRECT, before,
               ws.cell(row, col).value, anchor, note)

    # 3. Anchor resolution across every factor row.
    ws = wb["Credibility Factors"]
    acol = anchor_col(ws)
    for row, section in SECTION.items():
        name = str(ws.cell(row, 1).value or "").strip()
        if not name:
            continue
        before = ws.cell(row, acol).value
        new, why = resolve_anchor(before, gt_src.get(name, ""), section)
        action = CONFIRM if new == before else CORRECT
        ws.cell(row, acol).value = new
        record("Credibility Factors", f"{chr(64 + acol)}{row}", action, before, new, new, why)

    # 3b. Anchors on the sheets the ground truth does not cover.
    for (sheet, row), (new, why) in OTHER_ANCHORS[cou].items():
        ws2 = wb[sheet]
        acol2 = anchor_col(ws2)
        before = ws2.cell(row, acol2).value
        if before and "[CANDIDATE" not in before and before == new:
            action = CONFIRM
        else:
            action = CONFIRM if before == new else CORRECT
        ws2.cell(row, acol2).value = new
        record(sheet, f"{chr(64 + acol2)}{row}", action, before, new, new, why)

    ws = wb["Credibility Factors"]

    # 4. Level confirmation, required compared strictly.
    for row, section in SECTION.items():
        name = str(ws.cell(row, 1).value or "").strip()
        f = gt_lvl.get(name)
        if not f:
            continue
        if row in GT_DEFECTS[cou]:
            record("Credibility Factors", f"C{row}/D{row}", CONFIRM,
                   f"{ws.cell(row,3).value}/{ws.cell(row,4).value}",
                   f"{ws.cell(row,3).value}/{ws.cell(row,4).value}",
                   f"{NARRATIVE} §{section}; §6",
                   "GT-DEFECT, workbook stands. " + GT_DEFECTS[cou][row])
            continue
        req, ach = ws.cell(row, 3).value, ws.cell(row, 4).value
        greq, gach = f.get("expected_required_level"), f.get("expected_level")
        tol = f.get("level_tolerance", 0)
        req_ok = req == greq
        ach_ok = (ach == gach) or (
            None not in (ach, gach) and abs(ach - gach) <= tol)
        already = any(l[1] == f"D{row}" for l in ledger)
        if req_ok and ach_ok and not already:
            record("Credibility Factors", f"C{row}/D{row}", CONFIRM, f"{req}/{ach}",
                   f"{req}/{ach}", f"pre-registered GT + {NARRATIVE} §{section}",
                   "Confirmed by pre-registration: required matches GT exactly (compared "
                   "strictly), achieved matches" +
                   (f" within tolerance {tol}" if ach != gach else " exactly") + ".")

    wb.save(out)

    lines = [
        f"# Review ledger — aero {cou}",
        "",
        "Generated by `aero_cell_walk.py`. Every row is CONFIRM (checked and kept),",
        "CORRECT (checked and replaced), or BLANK (source-absent, emptied and listed).",
        "",
        "**Confirmed-by-preregistration** rows were checked against the author-committed",
        f"ground truth `tests/fixtures/extract/ground_truth/aero-{cou}-nasa7009b.json`.",
        "Required levels are compared **strictly**; `level_tolerance` is extraction latitude",
        "on the achieved level only, and applying it to required would mask A-7 by construction.",
        "",
        f"**{tally.get(CONFIRM,0)} confirmed · {tally.get(CORRECT,0)} corrected · "
        f"{tally.get(BLANK,0)} blanked (source-absent) · "
        f"{sum(tally.values())} decisions total**",
        "",
        "| Sheet | Cell | Action | Before | After | Anchor | Note |",
        "|---|---|---|---|---|---|---|",
    ]

    def clean(v, n=88):
        if v is None or v == "":
            return "_(blank)_"
        s = str(v).replace("|", "\\|").replace("\n", " ")
        return s if len(s) <= n else s[: n - 1] + "…"

    for sheet, cell, action, before, after, anchor, note in ledger:
        lines.append(
            f"| {sheet} | {cell} | **{action}** | {clean(before)} | {clean(after)} | "
            f"{clean(anchor)} | {clean(note, 400)} |")
    (pkg / "REVIEW_LEDGER.md").write_text("\n".join(lines) + "\n")

    print(f"{cou}: confirmed {tally.get(CONFIRM,0)}  corrected {tally.get(CORRECT,0)}  "
          f"blanked {tally.get(BLANK,0)}  total {sum(tally.values())}")


if __name__ == "__main__":
    for c in ("cou1", "cou2"):
        walk(c)
