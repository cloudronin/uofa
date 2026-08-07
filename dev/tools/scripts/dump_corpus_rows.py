#!/usr/bin/env python3
"""Freeze the extracted corpus rows the pinning tests read.

`tests/fixtures/extract_corpus/**/extracted.xlsx` is gitignored, and correctly
so: those files are the output of one paid extraction run, tied to a specific
(model, prompt-version, timestamp), and committing 1.2 MB of binary per run
couples a fixture to a run.

But two tests pin exact figures over them -- `rows == 800`,
`rows_below_required == 223`, `claims_grounded == 859` -- and those assertions
ALREADY couple to that run. So the coupling was never avoided; only the data
justifying it was missing, and in CI the files do not exist, every loop body is
skipped, the totals come out zero and both tests fail on an assertion that says
nothing about the cause.

This writes the derived rows as JSON: text, diffable, a fraction of the size, and
exactly what the tests consume. The numbers become reviewable in a diff rather
than regenerating silently.
"""
from __future__ import annotations

import json
import pathlib
import sys

_ROOT = pathlib.Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(_ROOT / "src"))

OUT = _ROOT / "tests" / "fixtures" / "extract_corpus" / "extracted_rows.json"


def collect() -> dict:
    import openpyxl

    from uofa_cli.excel_constants import NASA_ALL_FACTOR_NAMES, VV40_FACTOR_NAMES

    known = set(VV40_FACTOR_NAMES) | set(NASA_ALL_FACTOR_NAMES)
    bundles = {}
    for bd in sorted((_ROOT / "tests" / "fixtures" / "extract_corpus").glob("*/bundle_*")):
        x = bd / "extracted.xlsx"
        if not x.exists():
            continue
        wb = openpyxl.load_workbook(x, data_only=True)
        if "Credibility Factors" not in wb.sheetnames:
            continue
        ws = wb["Credibility Factors"]
        rows = []
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value in known:
                rows.append({
                    "factor_type": ws.cell(r, 1).value,
                    "required_level": ws.cell(r, 3).value,
                    "achieved_level": ws.cell(r, 4).value,
                    "acceptance_criteria": ws.cell(r, 5).value,
                    "rationale": ws.cell(r, 6).value,
                })
        if rows:
            bundles[str(bd.relative_to(_ROOT / "tests" / "fixtures" / "extract_corpus"))] = rows
    return bundles


def main() -> int:
    bundles = collect()
    if not bundles:
        raise SystemExit(
            "no extracted.xlsx found. This must be run where the extraction "
            "outputs exist; it is what makes them reviewable in CI.")
    OUT.write_text(json.dumps(bundles, indent=1, sort_keys=True) + "\n")
    n = sum(len(v) for v in bundles.values())
    print(f"  {OUT.relative_to(_ROOT)}")
    print(f"  {len(bundles)} bundles, {n} rows, {OUT.stat().st_size/1024:.0f} KB")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
