#!/usr/bin/env python3
"""Score extract accuracy against ground truth for a (pack, case) bundle.

Usage:
    # Morrison/vv40 (default case=cou1)
    python scripts/score_extraction.py --pack vv40 --model ollama/qwen3.5:4b --prompt-version v1

    # Aero/nasa-7009b, both COUs
    python scripts/score_extraction.py --pack nasa-7009b --case cou1 --model ollama/qwen3.5:4b --prompt-version v1-nasa-aero
    python scripts/score_extraction.py --pack nasa-7009b --case cou2 --model ollama/qwen3.5:4b --prompt-version v1-nasa-aero

Pipeline: extract -> parse xlsx -> factor scoring. If the ground truth
declares an expected_weakeners block, also: import -> rules --format
jsonld -o reasoned.jsonld -> parse -> weakener scoring. Hard gate on
must_not_fire patterns and structural_invariants. Appends a log entry
per run to extract_accuracy_log.jsonl.
"""

import argparse
import json
import subprocess
import sys
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
_LOG_PATH = _ROOT / "scripts" / "extract_accuracy_log.jsonl"

# (pack, case) -> (evidence_dir_name, ground_truth_filename, factor_names_symbol)
# factor_names_symbol is resolved at runtime via excel_constants import to
# avoid pulling the module in at module load time.
PACK_CASE_FIXTURES: dict[tuple[str, str], tuple[str, str, str]] = {
    ("vv40", "cou1"):      ("morrison-evidence",    "morrison-cou1.json",       "VV40_FACTOR_NAMES"),
    ("nasa-7009b", "cou1"): ("aero-evidence-cou1",  "aero-cou1-nasa7009b.json", "NASA_ALL_FACTOR_NAMES"),
    ("nasa-7009b", "cou2"): ("aero-evidence-cou2",  "aero-cou2-nasa7009b.json", "NASA_ALL_FACTOR_NAMES"),
}


def resolve_fixture(pack: str, case: str) -> tuple[Path, Path, list[str]]:
    """Return (evidence_dir, ground_truth_path, factor_names_list) for a pack/case pair."""
    key = (pack, case)
    if key not in PACK_CASE_FIXTURES:
        raise SystemExit(
            f"Unknown (pack, case): {key!r}. Known: {list(PACK_CASE_FIXTURES.keys())}"
        )
    ev_name, gt_name, fn_symbol = PACK_CASE_FIXTURES[key]
    from uofa_cli import excel_constants  # noqa: WPS433 (deferred import)
    factor_names = getattr(excel_constants, fn_symbol)
    return (
        _ROOT / "tests" / "fixtures" / "extract" / ev_name,
        _ROOT / "tests" / "fixtures" / "extract" / "ground_truth" / gt_name,
        factor_names,
    )


def _uofa_env() -> dict:
    """Build the subprocess env with repo src/ on PYTHONPATH for in-tree runs."""
    import os
    env = os.environ.copy()
    src = str(_ROOT / "src")
    existing = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = f"{src}:{existing}" if existing else src
    return env


def run_extraction(model: str, evidence_dir: Path, output_xlsx: Path, pack: str = "vv40") -> bool:
    """Run uofa extract and return success status."""
    # NOTE: --pack must follow the subcommand — argparse with parents+action=append
    # drops the top-level value otherwise.
    cmd = [
        sys.executable, "-m", "uofa_cli",
        "extract", str(evidence_dir),
        "--pack", pack,
        "--model", model,
        "-o", str(output_xlsx),
        "--verbose",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(_ROOT), env=_uofa_env())
    print(result.stdout)
    if result.returncode != 0:
        print(f"EXTRACT FAILED: {result.stderr}", file=sys.stderr)
        return False
    return True


def parse_extracted_xlsx(xlsx_path: Path, factor_names: list[str]) -> dict:
    """Parse the extracted Excel file into a dict matching ground truth structure.

    Reads:
    - Assessment Summary sheet -> summary fields
    - Model & Data sheet -> entity list
    - Validation Results sheet -> result list
    - Credibility Factors sheet -> factor list
    - Decision sheet -> outcome, rationale, decided_by, date

    factor_names: list of canonical factor name strings for this pack. Used
    to locate factor rows in the Credibility Factors sheet (column A).
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    result = {
        "assessment_summary": {},
        "entities": [],
        "validation_results": [],
        "credibility_factors": [],
        "decision": {},
    }

    # -- Assessment Summary --
    if "Assessment Summary" in wb.sheetnames:
        ws = wb["Assessment Summary"]
        data_row = _find_data_row(ws)
        field_cols = {
            "project_name": 1,
            "cou_name": 2,
            "cou_description": 3,
            "profile": 4,
            "device_class": 5,
            "model_risk_level": 6,
            "assurance_level": 7,
            "standards_reference": 8,
            "assessor_name": 9,
            "has_uq": 12,
        }
        for field, col in field_cols.items():
            val = ws.cell(row=data_row, column=col).value
            result["assessment_summary"][field] = val

    # -- Model & Data --
    if "Model & Data" in wb.sheetnames:
        ws = wb["Model & Data"]
        data_row = _find_data_row(ws)
        for row in range(data_row, ws.max_row + 1):
            entity_type = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            if not entity_type and not name:
                break
            result["entities"].append({
                "entity_type": entity_type,
                "name": name,
                "description": ws.cell(row=row, column=4).value,
            })

    # -- Validation Results --
    if "Validation Results" in wb.sheetnames:
        ws = wb["Validation Results"]
        data_row = _find_data_row(ws)
        for row in range(data_row, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if not name:
                break
            result["validation_results"].append({
                "name": name,
                "evidence_type": ws.cell(row=row, column=2).value,
                "description": ws.cell(row=row, column=4).value,
                "compares_to": ws.cell(row=row, column=5).value,
                "has_uq": ws.cell(row=row, column=6).value,
                "uq_method": ws.cell(row=row, column=7).value,
                "metric_value": ws.cell(row=row, column=8).value,
                "pass_fail": ws.cell(row=row, column=9).value,
            })

    # -- Credibility Factors --
    if "Credibility Factors" in wb.sheetnames:
        ws = wb["Credibility Factors"]
        for row in range(1, ws.max_row + 1):
            factor_type = ws.cell(row=row, column=1).value
            if factor_type and str(factor_type) in factor_names:
                result["credibility_factors"].append({
                    "factor_type": str(factor_type),
                    "required_level": ws.cell(row=row, column=3).value,
                    "achieved_level": ws.cell(row=row, column=4).value,
                    "acceptance_criteria": ws.cell(row=row, column=5).value,
                    "rationale": ws.cell(row=row, column=6).value,
                    "status": ws.cell(row=row, column=7).value,
                })

    # -- Decision --
    if "Decision" in wb.sheetnames:
        ws = wb["Decision"]
        data_row = _find_data_row(ws)
        result["decision"] = {
            "outcome": ws.cell(row=data_row, column=1).value,
            "rationale": ws.cell(row=data_row, column=2).value,
            "decided_by": ws.cell(row=data_row, column=4).value,
            "decision_date": ws.cell(row=data_row, column=5).value,
        }

    wb.close()
    return result


def _find_data_row(ws) -> int:
    """Find the first data row (row after header keywords)."""
    for row in range(1, min(ws.max_row + 1, 10)):
        cell = ws.cell(row=row, column=1).value
        if cell and any(kw in str(cell) for kw in [
            "Project Name", "Entity Type", "Result Name",
            "Decision Outcome", "Factor Type",
        ]):
            return row + 1
    return 3


def score_factors(extracted_factors: list, ground_truth_factors: list) -> dict:
    """Score credibility factor extraction accuracy."""
    results = {
        "total_factors": len(ground_truth_factors),
        "factors_found": 0,
        "factors_correct_type": 0,
        "factors_correct_level": 0,
        "factors_correct_status": 0,
        "factors_correct_source": 0,
        "per_factor": {},
    }

    extracted_by_type = {}
    for f in extracted_factors:
        ft = f.get("factor_type", "")
        extracted_by_type[ft] = f

    for gt in ground_truth_factors:
        gt_type = gt["factor_type"]
        extracted = extracted_by_type.get(gt_type)

        factor_result = {"ground_truth": gt}

        if extracted is None:
            factor_result["status"] = "MISS"
            factor_result["detail"] = "Factor not found in extraction"
        else:
            results["factors_found"] += 1
            results["factors_correct_type"] += 1
            factor_result["status"] = "FOUND"
            factor_result["extracted"] = extracted

            # Check level accuracy
            gt_level = gt.get("expected_level")
            ext_level = extracted.get("achieved_level")
            tolerance = gt.get("level_tolerance", 1)
            if gt_level is not None and ext_level is not None:
                try:
                    ext_level = int(ext_level)
                except (ValueError, TypeError):
                    ext_level = None

            if gt_level is not None and ext_level is not None:
                if abs(gt_level - ext_level) <= tolerance:
                    results["factors_correct_level"] += 1
                    factor_result["level_match"] = True
                else:
                    factor_result["level_match"] = False
                    factor_result["level_detail"] = f"expected ~{gt_level}\u00b1{tolerance}, got {ext_level}"
            else:
                factor_result["level_match"] = False
                factor_result["level_detail"] = f"expected {gt_level}, got {ext_level}"

            # Check status accuracy
            gt_status = gt.get("expected_status")
            ext_status = (extracted.get("status") or "").lower().strip()
            if gt_status == ext_status:
                results["factors_correct_status"] += 1
                factor_result["status_match"] = True
            else:
                factor_result["status_match"] = False

        results["per_factor"][gt_type] = factor_result

    n = results["total_factors"]
    results["detection_rate"] = results["factors_found"] / n if n else 0
    results["level_accuracy"] = results["factors_correct_level"] / results["factors_found"] if results["factors_found"] else 0
    results["status_accuracy"] = results["factors_correct_status"] / n if n else 0
    results["overall_f1"] = _compute_f1(results, extracted_by_type, ground_truth_factors)

    return results


def _compute_f1(results, extracted_by_type, ground_truth_factors):
    """Compute F1 score for factor detection."""
    gt_types = {f["factor_type"] for f in ground_truth_factors if f["expected_status"] == "assessed"}
    ext_types = set(extracted_by_type.keys())

    tp = len(gt_types & ext_types)
    fp = len(ext_types - gt_types)
    fn = len(gt_types - ext_types)

    precision = tp / (tp + fp) if (tp + fp) else 0
    recall = tp / (tp + fn) if (tp + fn) else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0

    return f1


def score_summary(extracted_summary: dict, ground_truth_summary: dict) -> dict:
    """Score assessment summary field extraction."""
    results = {"total_fields": 0, "correct_fields": 0, "per_field": {}}

    for field, expected in ground_truth_summary.items():
        results["total_fields"] += 1
        extracted_val = extracted_summary.get(field)

        if expected and extracted_val:
            if isinstance(expected, str) and isinstance(extracted_val, str):
                match = expected.lower().strip() == extracted_val.lower().strip()
                if not match and len(expected) > 10:
                    match = expected.lower() in extracted_val.lower() or extracted_val.lower() in expected.lower()
            else:
                match = expected == extracted_val

            if match:
                results["correct_fields"] += 1
                results["per_field"][field] = "MATCH"
            else:
                results["per_field"][field] = f"MISMATCH: expected '{expected}', got '{extracted_val}'"
        elif expected is None and extracted_val is None:
            results["correct_fields"] += 1
            results["per_field"][field] = "MATCH (both null)"
        else:
            results["per_field"][field] = f"MISMATCH: expected '{expected}', got '{extracted_val}'"

    results["accuracy"] = results["correct_fields"] / results["total_fields"] if results["total_fields"] else 0
    return results


def score_decision(extracted_decision: dict, ground_truth_decision: dict) -> dict:
    """Score decision extraction."""
    results = {"outcome_match": False, "rationale_keywords_found": 0}

    ext_outcome = (extracted_decision.get("outcome") or "")
    gt_outcome = ground_truth_decision.get("outcome", "")
    results["outcome_match"] = ext_outcome.lower().strip() == gt_outcome.lower().strip()

    ext_rationale = (extracted_decision.get("rationale") or "").lower()
    for kw in ground_truth_decision.get("rationale_keywords", []):
        if kw.lower() in ext_rationale:
            results["rationale_keywords_found"] += 1

    total_kw = len(ground_truth_decision.get("rationale_keywords", []))
    results["rationale_coverage"] = results["rationale_keywords_found"] / total_kw if total_kw else 0

    return results


# ── Weakener pipeline (C3 scoring) ───────────────────────────────────


def run_import(xlsx_path: Path, jsonld_path: Path, pack: str) -> bool:
    """Run uofa import to produce an (unsigned) jsonld from the extracted xlsx."""
    cmd = [
        sys.executable, "-m", "uofa_cli",
        "import", str(xlsx_path),
        "--pack", pack,
        "-o", str(jsonld_path),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(_ROOT), env=_uofa_env())
    if result.returncode != 0:
        print(f"IMPORT FAILED:\nstdout: {result.stdout}\nstderr: {result.stderr}", file=sys.stderr)
        return False
    return True


def run_rules_jsonld(jsonld_path: Path, reasoned_path: Path, pack: str) -> bool:
    """Run uofa rules with --format jsonld, writing reasoned output to reasoned_path."""
    cmd = [
        sys.executable, "-m", "uofa_cli",
        "rules", str(jsonld_path),
        "--pack", pack,
        "--format", "jsonld", "-o", str(reasoned_path), "--build",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(_ROOT), env=_uofa_env())
    if result.returncode != 0:
        print(f"RULES FAILED:\nstdout: {result.stdout}\nstderr: {result.stderr}", file=sys.stderr)
        return False
    return True


def parse_reasoned_weakeners(reasoned_path: Path) -> Counter:
    """Parse a reasoned jsonld and return a Counter of pattern_id -> count."""
    data = json.loads(reasoned_path.read_text())
    counts: Counter = Counter()
    PATTERN_KEY = "https://uofa.net/vocab#patternId"
    for node in data.get("@graph", []):
        pid = node.get(PATTERN_KEY)
        if pid:
            # pid may be a string or an {"@value": ...} wrapper
            if isinstance(pid, dict):
                pid = pid.get("@value", "")
            counts[str(pid)] += 1
    return counts


def score_weakeners(actual: Counter, expected: dict) -> dict:
    """Score weakener fires against expected_weakeners block.

    Returns dict with:
      - deterministic_fires_verified: bool
      - must_not_fire_verified: bool
      - structural_invariants_passed: bool
      - per_check: list of {"check": str, "passed": bool, "detail": str}
      - weakener_totals: dict(pattern -> count) (plus '__total__' sum)
    """
    per_check: list[dict] = []

    # 1. Deterministic core fires
    det_ok = True
    for pattern, spec in expected.get("deterministic_core_fires", {}).items():
        actual_count = actual.get(pattern, 0)
        if "count" in spec:
            want = spec["count"]
            ok = actual_count == want
            per_check.append({
                "check": f"{pattern} count == {want}",
                "passed": ok,
                "detail": f"actual={actual_count}",
            })
            det_ok = det_ok and ok
        elif "count_min" in spec and "count_max" in spec:
            lo, hi = spec["count_min"], spec["count_max"]
            ok = lo <= actual_count <= hi
            per_check.append({
                "check": f"{pattern} count in [{lo}, {hi}]",
                "passed": ok,
                "detail": f"actual={actual_count}",
            })
            det_ok = det_ok and ok
        elif "count_min" in spec:
            lo = spec["count_min"]
            ok = actual_count >= lo
            per_check.append({
                "check": f"{pattern} count >= {lo}",
                "passed": ok,
                "detail": f"actual={actual_count}",
            })
            det_ok = det_ok and ok

    # 2. Must-not-fire
    must_not_ok = True
    for pattern in expected.get("must_not_fire", []):
        actual_count = actual.get(pattern, 0)
        ok = actual_count == 0
        per_check.append({
            "check": f"{pattern} must NOT fire",
            "passed": ok,
            "detail": f"actual={actual_count}",
        })
        must_not_ok = must_not_ok and ok

    # 3. Structural invariants
    inv = expected.get("structural_invariants", {})
    inv_ok = True
    total = sum(actual.values())
    total_min = inv.get("total_count_min")
    total_max = inv.get("total_count_max")
    if total_min is not None:
        ok = total >= total_min
        per_check.append({"check": f"total >= {total_min}", "passed": ok, "detail": f"actual={total}"})
        inv_ok = inv_ok and ok
    if total_max is not None:
        ok = total <= total_max
        per_check.append({"check": f"total <= {total_max}", "passed": ok, "detail": f"actual={total}"})
        inv_ok = inv_ok and ok
    # Skip the special total_* keys we already processed above.
    _skip = {"total_count_min", "total_count_max"}
    for key, value in inv.items():
        if key in _skip or key.endswith("_rationale"):
            continue
        if key.endswith("_count_exact"):
            pattern = _normalize_pattern_name(key[:-len("_count_exact")])
            actual_count = actual.get(pattern, 0)
            ok = actual_count == value
            per_check.append({
                "check": f"{pattern} count == {value}",
                "passed": ok,
                "detail": f"actual={actual_count}",
            })
            inv_ok = inv_ok and ok
        elif key.endswith("_count_min"):
            pattern = _normalize_pattern_name(key[:-len("_count_min")])
            actual_count = actual.get(pattern, 0)
            ok = actual_count >= value
            per_check.append({
                "check": f"{pattern} count >= {value}",
                "passed": ok,
                "detail": f"actual={actual_count}",
            })
            inv_ok = inv_ok and ok

    totals = dict(actual)
    totals["__total__"] = total

    return {
        "deterministic_fires_verified": det_ok,
        "must_not_fire_verified": must_not_ok,
        "structural_invariants_passed": inv_ok,
        "per_check": per_check,
        "weakener_totals": totals,
    }


def _normalize_pattern_name(key: str) -> str:
    """Convert w_ep_04 / w_ar_02 / compound_01 snake_case to W-EP-04 / COMPOUND-01."""
    parts = key.split("_")
    # Uppercase all parts, join with '-'
    return "-".join(p.upper() for p in parts)


def print_weakener_report(w_scores: dict, expected: dict) -> None:
    print(f"\n  WEAKENERS")
    print(f"  {'─' * 50}")
    totals = w_scores["weakener_totals"]
    total = totals.get("__total__", 0)
    print(f"  Total fires: {total}")
    for pid in sorted(k for k in totals if k != "__total__"):
        print(f"    {pid}: {totals[pid]}")
    print(f"\n  Checks:")
    for c in w_scores["per_check"]:
        icon = "+" if c["passed"] else "x"
        print(f"    {icon} {c['check']} ({c['detail']})")
    gates = [
        ("deterministic fires", w_scores["deterministic_fires_verified"]),
        ("must-not-fire",       w_scores["must_not_fire_verified"]),
        ("structural invariants", w_scores["structural_invariants_passed"]),
    ]
    print(f"\n  Weakener gates:")
    for label, ok in gates:
        print(f"    {'+' if ok else 'x'} {label}")


def print_report(factor_scores, summary_scores, decision_scores, model, prompt_version):
    """Print a formatted accuracy report."""
    print("\n" + "=" * 70)
    print(f"  EXTRACTION ACCURACY REPORT")
    print(f"  Model: {model} | Prompt: {prompt_version}")
    print(f"  Date: {datetime.now().isoformat()}")
    print("=" * 70)

    print(f"\n  CREDIBILITY FACTORS ({factor_scores['total_factors']} total)")
    print(f"  {'─' * 50}")
    print(f"  Detection rate:    {factor_scores['detection_rate']:.0%} ({factor_scores['factors_found']}/{factor_scores['total_factors']})")
    print(f"  Level accuracy:    {factor_scores['level_accuracy']:.0%} ({factor_scores['factors_correct_level']}/{factor_scores['factors_found']})")
    print(f"  Status accuracy:   {factor_scores['status_accuracy']:.0%}")
    print(f"  Factor F1:         {factor_scores['overall_f1']:.2f}")

    print(f"\n  {'Factor':<45s} {'Status':<8s} {'Level':<12s}")
    print(f"  {'─' * 65}")
    for factor_type, result in factor_scores["per_factor"].items():
        status = result["status"]
        if status == "FOUND":
            level_str = "ok" if result.get("level_match") else result.get("level_detail", "?")
        else:
            level_str = "---"
        status_icon = "+" if status == "FOUND" else "x"
        print(f"  {factor_type:<45s} {status_icon:<8s} {level_str:<12s}")

    print(f"\n  ASSESSMENT SUMMARY ({summary_scores['total_fields']} fields)")
    print(f"  {'─' * 50}")
    print(f"  Accuracy: {summary_scores['accuracy']:.0%} ({summary_scores['correct_fields']}/{summary_scores['total_fields']})")
    for field, result in summary_scores["per_field"].items():
        icon = "+" if "MATCH" in result else "x"
        print(f"    {icon} {field}: {result}")

    print(f"\n  DECISION")
    print(f"  {'─' * 50}")
    print(f"  Outcome match:      {'+ yes' if decision_scores['outcome_match'] else 'x no'}")
    print(f"  Rationale coverage: {decision_scores['rationale_coverage']:.0%}")

    gate_f1 = factor_scores["overall_f1"]
    gate_pass = gate_f1 >= 0.70
    print(f"\n  {'=' * 50}")
    print(f"  GATE: F1 = {gate_f1:.2f} {'+ PASS (>=0.70)' if gate_pass else 'x FAIL (<0.70)'}")
    print(f"  {'=' * 50}")

    return gate_pass


def append_to_log(
    model, prompt_version, pack, case,
    factor_scores, summary_scores, decision_scores,
    weakener_scores=None,
):
    """Append results to a tracking log for prompt iteration history."""
    entry = {
        "timestamp": datetime.now().isoformat(),
        "model": model,
        "prompt_version": prompt_version,
        "pack": pack,
        "case": case,
        "factor_f1": round(factor_scores["overall_f1"], 3),
        "factor_detection": round(factor_scores["detection_rate"], 3),
        "factor_level_accuracy": round(factor_scores["level_accuracy"], 3),
        "summary_accuracy": round(summary_scores["accuracy"], 3),
        "decision_outcome_match": decision_scores["outcome_match"],
    }
    if weakener_scores is not None:
        entry["weakener_totals"] = weakener_scores["weakener_totals"]
        entry["deterministic_fires_verified"] = weakener_scores["deterministic_fires_verified"]
        entry["must_not_fire_verified"] = weakener_scores["must_not_fire_verified"]
        entry["structural_invariants_passed"] = weakener_scores["structural_invariants_passed"]
    _LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_LOG_PATH, "a") as f:
        f.write(json.dumps(entry) + "\n")
    print(f"\n  Results appended to {_LOG_PATH}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", default="ollama/qwen3.5:4b",
                        help="model string (default: ollama/qwen3.5:4b)")
    parser.add_argument("--prompt-version", default="v1-baseline",
                        help="prompt version tag for tracking")
    parser.add_argument("--pack", default="vv40",
                        help="pack name (default: vv40)")
    parser.add_argument("--case", default="cou1",
                        help="case within the pack (default: cou1)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="skip extraction and score an existing xlsx file")
    parser.add_argument("--reasoned", type=Path, default=None,
                        help="path to write reasoned jsonld (default: temp file; committed path for slide assets)")
    parser.add_argument("--skip-weakeners", action="store_true",
                        help="skip the import+rules weakener pipeline even if ground truth declares expected_weakeners")
    args = parser.parse_args()

    evidence_dir, ground_truth_path, factor_names = resolve_fixture(args.pack, args.case)

    if not ground_truth_path.exists():
        print(f"Ground truth not found: {ground_truth_path}", file=sys.stderr)
        return 1

    gt = json.loads(ground_truth_path.read_text())

    if args.xlsx:
        output_xlsx = args.xlsx
    else:
        if not evidence_dir.exists():
            print(f"Evidence directory not found: {evidence_dir}", file=sys.stderr)
            return 1

        output_xlsx = Path(tempfile.mktemp(suffix=".xlsx"))
        print(f"Output: {output_xlsx}")

        if not run_extraction(args.model, evidence_dir, output_xlsx, args.pack):
            return 1

    if not output_xlsx.exists():
        print(f"Output file not found: {output_xlsx}", file=sys.stderr)
        return 1

    extracted = parse_extracted_xlsx(output_xlsx, factor_names)

    factor_scores = score_factors(
        extracted["credibility_factors"],
        gt["expected_factors"],
    )
    summary_scores = score_summary(
        extracted["assessment_summary"],
        gt["assessment_summary"],
    )
    decision_scores = score_decision(
        extracted["decision"],
        gt["expected_decision"],
    )

    gate_pass = print_report(
        factor_scores, summary_scores, decision_scores,
        args.model, args.prompt_version,
    )

    # ── Weakener pipeline (if ground truth declares expected_weakeners) ──
    weakener_scores = None
    if "expected_weakeners" in gt and not args.skip_weakeners:
        jsonld_path = Path(tempfile.mktemp(suffix=".jsonld"))
        reasoned_path = args.reasoned or Path(tempfile.mktemp(suffix="-reasoned.jsonld"))
        print(f"\n  Import -> {jsonld_path}")
        if not run_import(output_xlsx, jsonld_path, args.pack):
            print("IMPORT failed; skipping weakener scoring", file=sys.stderr)
        else:
            print(f"  Rules  -> {reasoned_path}")
            if not run_rules_jsonld(jsonld_path, reasoned_path, args.pack):
                print("RULES failed; skipping weakener scoring", file=sys.stderr)
            else:
                counts = parse_reasoned_weakeners(reasoned_path)
                weakener_scores = score_weakeners(counts, gt["expected_weakeners"])
                print_weakener_report(weakener_scores, gt["expected_weakeners"])
                # Any weakener gate failing makes the overall run fail
                w_gate_pass = (
                    weakener_scores["deterministic_fires_verified"]
                    and weakener_scores["must_not_fire_verified"]
                    and weakener_scores["structural_invariants_passed"]
                )
                gate_pass = gate_pass and w_gate_pass
                if not w_gate_pass:
                    print(f"\n  x WEAKENER GATE FAIL", file=sys.stderr)

    append_to_log(
        args.model, args.prompt_version, args.pack, args.case,
        factor_scores, summary_scores, decision_scores,
        weakener_scores,
    )

    return 0 if gate_pass else 1


if __name__ == "__main__":
    sys.exit(main())
