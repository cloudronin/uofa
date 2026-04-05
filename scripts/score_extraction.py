#!/usr/bin/env python3
"""Score extract accuracy against Morrison ground truth.

Usage:
    python scripts/score_extraction.py [--model ollama/qwen3.5:4b] [--prompt-version v1]

Runs uofa extract on morrison-evidence/, parses the output Excel,
and scores against the ground truth. Prints detailed results and
appends to a log file for tracking prompt iterations.
"""

import argparse
import json
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
_EVIDENCE_DIR = _ROOT / "tests" / "fixtures" / "extract" / "morrison-evidence"
_GROUND_TRUTH = _ROOT / "tests" / "fixtures" / "extract" / "ground_truth" / "morrison-cou1.json"
_LOG_PATH = _ROOT / "scripts" / "extract_accuracy_log.jsonl"


def run_extraction(model: str, evidence_dir: Path, output_xlsx: Path, pack: str = "vv40") -> bool:
    """Run uofa extract and return success status."""
    cmd = [
        sys.executable, "-m", "uofa_cli", "extract", str(evidence_dir),
        "--model", model,
        "--pack", pack,
        "-o", str(output_xlsx),
        "--verbose",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(_ROOT))
    print(result.stdout)
    if result.returncode != 0:
        print(f"EXTRACT FAILED: {result.stderr}", file=sys.stderr)
        return False
    return True


def parse_extracted_xlsx(xlsx_path: Path) -> dict:
    """Parse the extracted Excel file into a dict matching ground truth structure.

    Reads:
    - Assessment Summary sheet -> summary fields
    - Model & Data sheet -> entity list
    - Validation Results sheet -> result list
    - Credibility Factors sheet -> factor list
    - Decision sheet -> outcome, rationale, decided_by, date
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    result = {
        "assessment_summary": {},
        "entities": [],
        "validation_results": [],
        "credibility_factors": [],
        "decision": {},
    }

    # -- Assessment Summary --
    if "Assessment Summary" in wb.sheetnames:
        ws = wb["Assessment Summary"]
        data_row = _find_data_row(ws)
        field_cols = {
            "project_name": 1,
            "cou_name": 2,
            "cou_description": 3,
            "profile": 4,
            "device_class": 5,
            "model_risk_level": 6,
            "assurance_level": 7,
            "standards_reference": 8,
            "assessor_name": 9,
            "has_uq": 12,
        }
        for field, col in field_cols.items():
            val = ws.cell(row=data_row, column=col).value
            result["assessment_summary"][field] = val

    # -- Model & Data --
    if "Model & Data" in wb.sheetnames:
        ws = wb["Model & Data"]
        data_row = _find_data_row(ws)
        for row in range(data_row, ws.max_row + 1):
            entity_type = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            if not entity_type and not name:
                break
            result["entities"].append({
                "entity_type": entity_type,
                "name": name,
                "description": ws.cell(row=row, column=4).value,
            })

    # -- Validation Results --
    if "Validation Results" in wb.sheetnames:
        ws = wb["Validation Results"]
        data_row = _find_data_row(ws)
        for row in range(data_row, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if not name:
                break
            result["validation_results"].append({
                "name": name,
                "evidence_type": ws.cell(row=row, column=2).value,
                "description": ws.cell(row=row, column=4).value,
                "compares_to": ws.cell(row=row, column=5).value,
                "has_uq": ws.cell(row=row, column=6).value,
                "uq_method": ws.cell(row=row, column=7).value,
                "metric_value": ws.cell(row=row, column=8).value,
                "pass_fail": ws.cell(row=row, column=9).value,
            })

    # -- Credibility Factors --
    if "Credibility Factors" in wb.sheetnames:
        ws = wb["Credibility Factors"]
        from uofa_cli.excel_constants import VV40_FACTOR_NAMES
        for row in range(1, ws.max_row + 1):
            factor_type = ws.cell(row=row, column=1).value
            if factor_type and str(factor_type) in VV40_FACTOR_NAMES:
                result["credibility_factors"].append({
                    "factor_type": str(factor_type),
                    "required_level": ws.cell(row=row, column=3).value,
                    "achieved_level": ws.cell(row=row, column=4).value,
                    "acceptance_criteria": ws.cell(row=row, column=5).value,
                    "rationale": ws.cell(row=row, column=6).value,
                    "status": ws.cell(row=row, column=7).value,
                })

    # -- Decision --
    if "Decision" in wb.sheetnames:
        ws = wb["Decision"]
        data_row = _find_data_row(ws)
        result["decision"] = {
            "outcome": ws.cell(row=data_row, column=1).value,
            "rationale": ws.cell(row=data_row, column=2).value,
            "decided_by": ws.cell(row=data_row, column=4).value,
            "decision_date": ws.cell(row=data_row, column=5).value,
        }

    wb.close()
    return result


def _find_data_row(ws) -> int:
    """Find the first data row (row after header keywords)."""
    for row in range(1, min(ws.max_row + 1, 10)):
        cell = ws.cell(row=row, column=1).value
        if cell and any(kw in str(cell) for kw in [
            "Project Name", "Entity Type", "Result Name",
            "Decision Outcome", "Factor Type",
        ]):
            return row + 1
    return 3


def score_factors(extracted_factors: list, ground_truth_factors: list) -> dict:
    """Score credibility factor extraction accuracy."""
    results = {
        "total_factors": len(ground_truth_factors),
        "factors_found": 0,
        "factors_correct_type": 0,
        "factors_correct_level": 0,
        "factors_correct_status": 0,
        "factors_correct_source": 0,
        "per_factor": {},
    }

    extracted_by_type = {}
    for f in extracted_factors:
        ft = f.get("factor_type", "")
        extracted_by_type[ft] = f

    for gt in ground_truth_factors:
        gt_type = gt["factor_type"]
        extracted = extracted_by_type.get(gt_type)

        factor_result = {"ground_truth": gt}

        if extracted is None:
            factor_result["status"] = "MISS"
            factor_result["detail"] = "Factor not found in extraction"
        else:
            results["factors_found"] += 1
            results["factors_correct_type"] += 1
            factor_result["status"] = "FOUND"
            factor_result["extracted"] = extracted

            # Check level accuracy
            gt_level = gt.get("expected_level")
            ext_level = extracted.get("achieved_level")
            tolerance = gt.get("level_tolerance", 1)
            if gt_level is not None and ext_level is not None:
                try:
                    ext_level = int(ext_level)
                except (ValueError, TypeError):
                    ext_level = None

            if gt_level is not None and ext_level is not None:
                if abs(gt_level - ext_level) <= tolerance:
                    results["factors_correct_level"] += 1
                    factor_result["level_match"] = True
                else:
                    factor_result["level_match"] = False
                    factor_result["level_detail"] = f"expected ~{gt_level}\u00b1{tolerance}, got {ext_level}"
            else:
                factor_result["level_match"] = False
                factor_result["level_detail"] = f"expected {gt_level}, got {ext_level}"

            # Check status accuracy
            gt_status = gt.get("expected_status")
            ext_status = (extracted.get("status") or "").lower().strip()
            if gt_status == ext_status:
                results["factors_correct_status"] += 1
                factor_result["status_match"] = True
            else:
                factor_result["status_match"] = False

        results["per_factor"][gt_type] = factor_result

    n = results["total_factors"]
    results["detection_rate"] = results["factors_found"] / n if n else 0
    results["level_accuracy"] = results["factors_correct_level"] / results["factors_found"] if results["factors_found"] else 0
    results["status_accuracy"] = results["factors_correct_status"] / n if n else 0
    results["overall_f1"] = _compute_f1(results, extracted_by_type, ground_truth_factors)

    return results


def _compute_f1(results, extracted_by_type, ground_truth_factors):
    """Compute F1 score for factor detection."""
    gt_types = {f["factor_type"] for f in ground_truth_factors if f["expected_status"] == "assessed"}
    ext_types = set(extracted_by_type.keys())

    tp = len(gt_types & ext_types)
    fp = len(ext_types - gt_types)
    fn = len(gt_types - ext_types)

    precision = tp / (tp + fp) if (tp + fp) else 0
    recall = tp / (tp + fn) if (tp + fn) else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0

    return f1


def score_summary(extracted_summary: dict, ground_truth_summary: dict) -> dict:
    """Score assessment summary field extraction."""
    results = {"total_fields": 0, "correct_fields": 0, "per_field": {}}

    for field, expected in ground_truth_summary.items():
        results["total_fields"] += 1
        extracted_val = extracted_summary.get(field)

        if expected and extracted_val:
            if isinstance(expected, str) and isinstance(extracted_val, str):
                match = expected.lower().strip() == extracted_val.lower().strip()
                if not match and len(expected) > 10:
                    match = expected.lower() in extracted_val.lower() or extracted_val.lower() in expected.lower()
            else:
                match = expected == extracted_val

            if match:
                results["correct_fields"] += 1
                results["per_field"][field] = "MATCH"
            else:
                results["per_field"][field] = f"MISMATCH: expected '{expected}', got '{extracted_val}'"
        elif expected is None and extracted_val is None:
            results["correct_fields"] += 1
            results["per_field"][field] = "MATCH (both null)"
        else:
            results["per_field"][field] = f"MISMATCH: expected '{expected}', got '{extracted_val}'"

    results["accuracy"] = results["correct_fields"] / results["total_fields"] if results["total_fields"] else 0
    return results


def score_decision(extracted_decision: dict, ground_truth_decision: dict) -> dict:
    """Score decision extraction."""
    results = {"outcome_match": False, "rationale_keywords_found": 0}

    ext_outcome = (extracted_decision.get("outcome") or "")
    gt_outcome = ground_truth_decision.get("outcome", "")
    results["outcome_match"] = ext_outcome.lower().strip() == gt_outcome.lower().strip()

    ext_rationale = (extracted_decision.get("rationale") or "").lower()
    for kw in ground_truth_decision.get("rationale_keywords", []):
        if kw.lower() in ext_rationale:
            results["rationale_keywords_found"] += 1

    total_kw = len(ground_truth_decision.get("rationale_keywords", []))
    results["rationale_coverage"] = results["rationale_keywords_found"] / total_kw if total_kw else 0

    return results


def print_report(factor_scores, summary_scores, decision_scores, model, prompt_version):
    """Print a formatted accuracy report."""
    print("\n" + "=" * 70)
    print(f"  EXTRACTION ACCURACY REPORT")
    print(f"  Model: {model} | Prompt: {prompt_version}")
    print(f"  Date: {datetime.now().isoformat()}")
    print("=" * 70)

    print(f"\n  CREDIBILITY FACTORS ({factor_scores['total_factors']} total)")
    print(f"  {'─' * 50}")
    print(f"  Detection rate:    {factor_scores['detection_rate']:.0%} ({factor_scores['factors_found']}/{factor_scores['total_factors']})")
    print(f"  Level accuracy:    {factor_scores['level_accuracy']:.0%} ({factor_scores['factors_correct_level']}/{factor_scores['factors_found']})")
    print(f"  Status accuracy:   {factor_scores['status_accuracy']:.0%}")
    print(f"  Factor F1:         {factor_scores['overall_f1']:.2f}")

    print(f"\n  {'Factor':<45s} {'Status':<8s} {'Level':<12s}")
    print(f"  {'─' * 65}")
    for factor_type, result in factor_scores["per_factor"].items():
        status = result["status"]
        if status == "FOUND":
            level_str = "ok" if result.get("level_match") else result.get("level_detail", "?")
        else:
            level_str = "---"
        status_icon = "+" if status == "FOUND" else "x"
        print(f"  {factor_type:<45s} {status_icon:<8s} {level_str:<12s}")

    print(f"\n  ASSESSMENT SUMMARY ({summary_scores['total_fields']} fields)")
    print(f"  {'─' * 50}")
    print(f"  Accuracy: {summary_scores['accuracy']:.0%} ({summary_scores['correct_fields']}/{summary_scores['total_fields']})")
    for field, result in summary_scores["per_field"].items():
        icon = "+" if "MATCH" in result else "x"
        print(f"    {icon} {field}: {result}")

    print(f"\n  DECISION")
    print(f"  {'─' * 50}")
    print(f"  Outcome match:      {'+ yes' if decision_scores['outcome_match'] else 'x no'}")
    print(f"  Rationale coverage: {decision_scores['rationale_coverage']:.0%}")

    gate_f1 = factor_scores["overall_f1"]
    gate_pass = gate_f1 >= 0.70
    print(f"\n  {'=' * 50}")
    print(f"  GATE: F1 = {gate_f1:.2f} {'+ PASS (>=0.70)' if gate_pass else 'x FAIL (<0.70)'}")
    print(f"  {'=' * 50}")

    return gate_pass


def append_to_log(model, prompt_version, factor_scores, summary_scores, decision_scores):
    """Append results to a tracking log for prompt iteration history."""
    entry = {
        "timestamp": datetime.now().isoformat(),
        "model": model,
        "prompt_version": prompt_version,
        "factor_f1": round(factor_scores["overall_f1"], 3),
        "factor_detection": round(factor_scores["detection_rate"], 3),
        "factor_level_accuracy": round(factor_scores["level_accuracy"], 3),
        "summary_accuracy": round(summary_scores["accuracy"], 3),
        "decision_outcome_match": decision_scores["outcome_match"],
    }
    _LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(_LOG_PATH, "a") as f:
        f.write(json.dumps(entry) + "\n")
    print(f"\n  Results appended to {_LOG_PATH}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model", default="ollama/qwen3.5:4b",
                        help="model string (default: ollama/qwen3.5:4b)")
    parser.add_argument("--prompt-version", default="v1-baseline",
                        help="prompt version tag for tracking")
    parser.add_argument("--pack", default="vv40",
                        help="pack name (default: vv40)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="skip extraction and score an existing xlsx file")
    args = parser.parse_args()

    if not _GROUND_TRUTH.exists():
        print(f"Ground truth not found: {_GROUND_TRUTH}", file=sys.stderr)
        return 1

    gt = json.loads(_GROUND_TRUTH.read_text())

    if args.xlsx:
        output_xlsx = args.xlsx
    else:
        if not _EVIDENCE_DIR.exists():
            print(f"Evidence directory not found: {_EVIDENCE_DIR}", file=sys.stderr)
            return 1

        output_xlsx = Path(tempfile.mktemp(suffix=".xlsx"))
        print(f"Output: {output_xlsx}")

        if not run_extraction(args.model, _EVIDENCE_DIR, output_xlsx, args.pack):
            return 1

    if not output_xlsx.exists():
        print(f"Output file not found: {output_xlsx}", file=sys.stderr)
        return 1

    extracted = parse_extracted_xlsx(output_xlsx)

    factor_scores = score_factors(
        extracted["credibility_factors"],
        gt["expected_factors"],
    )
    summary_scores = score_summary(
        extracted["assessment_summary"],
        gt["assessment_summary"],
    )
    decision_scores = score_decision(
        extracted["decision"],
        gt["expected_decision"],
    )

    gate_pass = print_report(
        factor_scores, summary_scores, decision_scores,
        args.model, args.prompt_version,
    )

    append_to_log(
        args.model, args.prompt_version,
        factor_scores, summary_scores, decision_scores,
    )

    return 0 if gate_pass else 1


if __name__ == "__main__":
    sys.exit(main())
