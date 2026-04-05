"""XLSX reader — converts spreadsheet sheets to markdown tables."""

from __future__ import annotations

from pathlib import Path

from uofa_cli.document_reader import DocumentChunk

# Sheets that indicate this is a UofA template, not an evidence file
_UOFA_TEMPLATE_SHEETS = {
    "Assessment Summary", "Model & Data", "Validation Results",
    "Credibility Factors", "Decision",
}


def read_xlsx(path: Path, row_budget: int = 50) -> list[DocumentChunk]:
    """Read an XLSX and return one chunk per sheet as a markdown table."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # Check for UofA template
    sheet_names = set(wb.sheetnames)
    is_template = len(sheet_names & _UOFA_TEMPLATE_SHEETS) >= 3

    chunks: list[DocumentChunk] = []
    warnings: list[str] = []

    if is_template:
        warnings.append(
            f"Warning: {path.name} appears to be a UofA template — "
            f"did you mean 'uofa import'? Processing as evidence file."
        )

    for sheet_name in wb.sheetnames:
        # Skip hidden or instruction sheets
        if sheet_name.startswith("_"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find first non-empty row as header
        header_idx = None
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                header_idx = i
                break

        if header_idx is None:
            continue

        header = [str(cell) if cell is not None else "" for cell in rows[header_idx]]
        # Skip sheets with no meaningful headers
        if all(h == "" for h in header):
            continue

        data_rows = rows[header_idx + 1:]
        # Filter out completely empty rows
        data_rows = [r for r in data_rows if any(cell is not None for cell in r)]

        if not data_rows:
            continue

        total_rows = len(data_rows)
        truncated = total_rows > row_budget
        display_rows = data_rows[:row_budget]

        # Build markdown table
        n_cols = len(header)
        md_lines = [
            f"| {' | '.join(header)} |",
            f"| {' | '.join(['---'] * n_cols)} |",
        ]
        for row in display_rows:
            cells = [str(cell) if cell is not None else "" for cell in row[:n_cols]]
            # Pad if row is shorter than header
            while len(cells) < n_cols:
                cells.append("")
            md_lines.append(f"| {' | '.join(cells)} |")

        if truncated:
            md_lines.append(f"(showing {row_budget} of {total_rows} rows — full data in source file)")

        text = f'=== Sheet: "{sheet_name}" ({total_rows} rows x {n_cols} cols) ===\n'
        text += "\n".join(md_lines)

        chunks.append(DocumentChunk(
            text=text,
            source_file=path.name,
            source_path=str(path),
            sheet_name=sheet_name,
            format="xlsx",
        ))

    wb.close()

    # Attach warnings to the first chunk
    if warnings and chunks:
        chunks[0].text = "\n".join(warnings) + "\n\n" + chunks[0].text

    return chunks
