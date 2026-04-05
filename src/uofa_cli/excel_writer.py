"""Excel writer — writes LLM extraction results into a pack template with confidence coloring."""

from __future__ import annotations

import difflib
import shutil
from pathlib import Path

from uofa_cli.llm_extractor import ExtractionResult, FieldExtraction
from uofa_cli.excel_constants import (
    VV40_FACTOR_NAMES, NASA_ALL_FACTOR_NAMES,
    VV40_FACTOR_CATEGORIES, NASA_ONLY_FACTOR_CATEGORIES,
    VALID_DECISION_OUTCOMES, VALID_PROFILES, VALID_ASSURANCE_LEVELS,
    VALID_DEVICE_CLASSES,
)

# Confidence color thresholds
_GREEN = "C6EFCE"   # >= 0.85
_YELLOW = "FFEB9C"  # 0.50 - 0.84
_RED = "FFC7CE"     # < 0.50


def write_extraction(
    result: ExtractionResult,
    template_path: Path | None,
    output_path: Path,
    pack_name: str,
) -> Path:
    """Write extraction results to an Excel template.

    If template_path exists, copies it and fills in extracted values.
    If template_path is None, creates sheets from scratch.
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment

    if template_path and template_path.exists():
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _create_blank_sheets(wb, pack_name)

    # Write each section
    if "Assessment Summary" in wb.sheetnames:
        _write_summary_sheet(wb["Assessment Summary"], result.assessment_summary)

    if "Model & Data" in wb.sheetnames:
        _write_entities_sheet(wb["Model & Data"], result.model_and_data)

    if "Validation Results" in wb.sheetnames:
        _write_validation_sheet(wb["Validation Results"], result.validation_results)

    if "Credibility Factors" in wb.sheetnames:
        _write_factors_sheet(wb["Credibility Factors"], result.credibility_factors, pack_name)

    if "Decision" in wb.sheetnames:
        _write_decision_sheet(wb["Decision"], result.decision)

    wb.save(output_path)
    return output_path


def _create_blank_sheets(wb, pack_name: str) -> None:
    """Create blank template sheets when no template file exists."""
    # Assessment Summary
    ws = wb.create_sheet("Assessment Summary", 0)
    ws.append(["Unit of Assurance — Assessment Summary"])
    ws.append(["Project Name", "COU Name", "COU Description", "Profile",
               "Device Class", "Model Risk Level", "Assurance Level",
               "Standards Reference", "Assessor Name", "Assessment Date",
               "Source Document", "Has UQ?"])
    ws.append([None] * 12)  # Empty data row

    # Model & Data
    ws = wb.create_sheet("Model & Data")
    ws.append(["Model & Data"])
    ws.append(["Entity Type", "Name", "Identifier/URI", "Description", "Version", "Source"])

    # Validation Results
    ws = wb.create_sheet("Validation Results")
    ws.append(["Validation Results"])
    ws.append(["Result Name", "Type", "Identifier/URI", "Description",
               "Compares To", "Has UQ?", "UQ Method", "Metric Value", "Pass/Fail"])

    # Credibility Factors
    ws = wb.create_sheet("Credibility Factors")
    ws.append(["Credibility Factors"])
    ws.append([""])
    ws.append(["Factor Type", "Category", "Required Level", "Achieved Level",
               "Acceptance Criteria", "Rationale", "Factor Status", "Linked Evidence"])
    ws.append(["(pre-filled)", "(pre-filled)", "(1-5)", "(1-5)",
               "(optional)", "(optional)", "(status)", "(URI)"])

    # Pre-fill factor rows
    is_nasa = "nasa" in pack_name.lower()
    categories = list(VV40_FACTOR_CATEGORIES)
    if is_nasa:
        categories.extend(NASA_ONLY_FACTOR_CATEGORIES)
    for name, cat in categories:
        ws.append([name, cat, None, None, None, None, None, None])

    # Decision
    ws = wb.create_sheet("Decision")
    ws.append(["Decision"])
    ws.append(["Decision Outcome", "Decision Rationale", "Criteria Set",
               "Decided By", "Decision Date"])
    ws.append([None] * 5)


# ── Sheet writers ────────────────────────────────────────────


def _write_summary_sheet(ws, summary: dict[str, FieldExtraction]) -> None:
    """Write assessment summary fields to the data row."""
    # Find the data row (row after header row)
    data_row = _find_data_row(ws)

    field_map = {
        "project_name": 1,
        "cou_name": 2,
        "cou_description": 3,
        "profile": 4,
        "device_class": 5,
        "model_risk_level": 6,
        "assurance_level": 7,
        "standards_reference": 8,
        "assessor_name": 9,
        "has_uq": 12,
    }

    dropdown_maps = {
        "profile": VALID_PROFILES,
        "device_class": VALID_DEVICE_CLASSES,
        "assurance_level": VALID_ASSURANCE_LEVELS,
    }

    for field_name, col in field_map.items():
        fe = summary.get(field_name)
        if fe is None or fe.value is None:
            continue

        value = fe.value
        # Fuzzy match dropdowns
        if field_name in dropdown_maps:
            value = _fuzzy_match_dropdown(str(value), dropdown_maps[field_name])

        cell = ws.cell(row=data_row, column=col)
        cell.value = value
        _apply_confidence_color(cell, fe.confidence)
        _add_source_comment(cell, fe)


def _write_entities_sheet(ws, entities: list[dict[str, FieldExtraction]]) -> None:
    """Write model & data entities."""
    data_row = _find_data_row(ws)

    for i, entity in enumerate(entities):
        row = data_row + i
        field_map = {
            "entity_type": 1,
            "name": 2,
            "description": 4,
        }
        for field_name, col in field_map.items():
            fe = entity.get(field_name)
            if fe is None or fe.value is None:
                continue
            cell = ws.cell(row=row, column=col)
            cell.value = fe.value
            _apply_confidence_color(cell, fe.confidence)
            _add_source_comment(cell, fe)


def _write_validation_sheet(ws, results: list[dict[str, FieldExtraction]]) -> None:
    """Write validation result rows."""
    data_row = _find_data_row(ws)

    for i, vr in enumerate(results):
        row = data_row + i
        field_map = {
            "name": 1,
            "evidence_type": 2,
            "description": 4,
            "compares_to": 5,
            "has_uq": 6,
            "uq_method": 7,
            "metric_value": 8,
            "pass_fail": 9,
        }
        for field_name, col in field_map.items():
            fe = vr.get(field_name)
            if fe is None or fe.value is None:
                continue
            cell = ws.cell(row=row, column=col)
            cell.value = fe.value
            _apply_confidence_color(cell, fe.confidence)
            _add_source_comment(cell, fe)


def _write_factors_sheet(
    ws, factors: list[dict[str, FieldExtraction]], pack_name: str,
) -> None:
    """Write credibility factor values to the existing factor rows in the template.

    Matches extracted factors to existing rows by factor_type in column A.
    Column H (Linked Evidence) is explicitly left empty.
    """
    is_nasa = "nasa" in pack_name.lower()
    valid_names = NASA_ALL_FACTOR_NAMES if is_nasa else VV40_FACTOR_NAMES

    # Build a map of factor_type -> row number from the existing sheet
    factor_rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and str(cell_val) in valid_names:
            factor_rows[str(cell_val)] = row

    # If no pre-existing rows (blank template without factors), start after headers
    if not factor_rows:
        start_row = 5  # After header rows
        for i, factor in enumerate(factors):
            ft_fe = factor.get("factor_type")
            if ft_fe is None:
                continue
            row = start_row + i
            ws.cell(row=row, column=1).value = ft_fe.value
            _write_factor_row(ws, row, factor)
        return

    # Match extracted factors to existing rows
    for factor in factors:
        ft_fe = factor.get("factor_type")
        if ft_fe is None or ft_fe.value is None:
            continue

        ft_name = str(ft_fe.value)
        row = factor_rows.get(ft_name)
        if row is None:
            # Try fuzzy match
            matches = difflib.get_close_matches(ft_name, list(factor_rows.keys()), n=1, cutoff=0.6)
            if matches:
                row = factor_rows[matches[0]]

        if row is not None:
            _write_factor_row(ws, row, factor)


def _write_factor_row(ws, row: int, factor: dict[str, FieldExtraction]) -> None:
    """Write a single factor's data to a row. Column H (8) is left empty."""
    field_map = {
        "required_level": 3,     # Column C
        "achieved_level": 4,     # Column D
        "acceptance_criteria": 5, # Column E
        "rationale": 6,          # Column F
        "status": 7,             # Column G
        # Column H (8) = Linked Evidence — intentionally skipped
    }

    for field_name, col in field_map.items():
        fe = factor.get(field_name)
        if fe is None or fe.value is None:
            continue
        cell = ws.cell(row=row, column=col)
        cell.value = fe.value
        _apply_confidence_color(cell, fe.confidence)
        _add_source_comment(cell, fe)


def _write_decision_sheet(ws, decision: dict[str, FieldExtraction]) -> None:
    """Write decision fields."""
    data_row = _find_data_row(ws)

    field_map = {
        "outcome": 1,
        "rationale": 2,
        "decided_by": 4,
        "decision_date": 5,
    }

    for field_name, col in field_map.items():
        fe = decision.get(field_name)
        if fe is None or fe.value is None:
            continue

        value = fe.value
        if field_name == "outcome":
            value = _fuzzy_match_dropdown(str(value), VALID_DECISION_OUTCOMES)

        cell = ws.cell(row=data_row, column=col)
        cell.value = value
        _apply_confidence_color(cell, fe.confidence)
        _add_source_comment(cell, fe)


# ── Helpers ──────────────────────────────────────────────────


def _find_data_row(ws) -> int:
    """Find the first data row (row after the header row).

    Looks for a row with header-like content, returns the row after it.
    Falls back to row 3 (typical layout: row 1 = title, row 2 = headers, row 3 = data).
    """
    for row in range(1, min(ws.max_row + 1, 10)):
        cell = ws.cell(row=row, column=1).value
        if cell and any(kw in str(cell) for kw in [
            "Project Name", "Entity Type", "Result Name",
            "Decision Outcome", "Factor Type",
        ]):
            return row + 1
    return 3


def _apply_confidence_color(cell, confidence: float) -> None:
    """Apply green/yellow/red fill based on confidence score."""
    from openpyxl.styles import PatternFill

    if confidence >= 0.85:
        fill_color = _GREEN
    elif confidence >= 0.50:
        fill_color = _YELLOW
    else:
        fill_color = _RED

    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def _add_source_comment(cell, fe: FieldExtraction) -> None:
    """Add a cell comment with source attribution."""
    from openpyxl.comments import Comment

    parts = ["[Extract]"]
    if fe.source_file:
        source = f"Source: {fe.source_file}"
        if fe.source_page:
            source += f" p.{fe.source_page}"
        parts.append(source)
    parts.append(f"Confidence: {fe.confidence:.0%}")

    cell.comment = Comment(" | ".join(parts), "uofa extract")


def _fuzzy_match_dropdown(value: str, valid_values: list[str]) -> str:
    """Fuzzy match a value against valid dropdown options."""
    if value in valid_values:
        return value

    matches = difflib.get_close_matches(value, valid_values, n=1, cutoff=0.6)
    return matches[0] if matches else value
