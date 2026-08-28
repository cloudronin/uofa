"""Protocol-check — mechanical conformance checks for a reference encoding.

These are the scriptable subset of `docs/Encoding_Protocol_v0_1.md`, seeded by
finding F-6c of the Johnson pilot. They exist so that no workbook reaches the author's
review pass while it still fails a check a machine could have run.

## Why a flag and not a command

The checks have no state of their own and nothing to orchestrate. They are assertions
about an artifact that `extract` has just written or that `import` has just been handed,
so they ride those commands rather than adding a third that would need to be told where
to look.

## Why the two commands behave differently

Every check here describes a *reviewed* workbook. A freshly extracted one has no citation
anchors, because anchors are what the review pass produces, so on `extract` the flag
prints the table and leaves the exit code alone: it is telling the encoder what review has
to produce, not failing them for not having done it yet. On `import` the workbook is the
reviewed one and the checks are gates, so any failure exits non-zero.

## Placeholder detection is pack-derived, not hardcoded

Every pack template carries a description row under each header row, and its strings are
exactly the hint text that leaks into data rows when the extractor writes nothing
(F-3d). The placeholder set is read from the active pack's own template rather than listed
here, so a pack that changes its hints does not silently stop being checked.

## Why the namespace check names a domain family, not a string

The protocol's rule is *mint under a namespace you control*. The importer's own warning fires on
one string, its `example.org` default, so an encoder can satisfy the warning and still miss the
rule: a reserved example domain under a plausible subdomain reads like a real namespace and is
one nobody controls. The check therefore rejects the whole RFC 2606 / RFC 6761 reserved family.
It reads the package's minted `id` because that string sits inside the canonicalised content the
hash and signature cover, which is what makes the mistake permanent rather than cosmetic.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlsplit

from uofa_cli.excel_constants import (
    LEVEL_PROVENANCE_HEADER, SHEET_NAMES, WORKBOOK_PROFILE_HEADER,
    level_terms_after_vocab, level_vocab_for, version_tuple,
)

ANCHOR_HEADER = "Source Anchor"

# Files the protocol requires beside a package. Names are conventions rather than
# schema, so they are matched case-insensitively against a small candidate set.
AMBIGUITY_LOG_NAMES = ("ambiguity_log.md", "ambiguity-log.md", "ambiguitylog.md")
RUN_LOG_NAMES = ("run_log.md", "run-log.md", "runlog.md")

# Section 3 of the protocol names these as the run log's mandatory pins.
RUN_LOG_FIELDS = ("model", "backend", "site commit", "repo head", "base_uri")

# A run log that labels itself a pilot must not record a signing step, because a pilot
# runs before the protocol governs it and nothing it produces may be signed.
PILOT_MARKERS = ("pilot", "PILOT")

# Domains reserved by RFC 2606 §2-3 and RFC 6761 §6 for documentation and testing. Nobody
# controls any of them, so none can satisfy the protocol's "a namespace you control".
RESERVED_SECOND_LEVEL = ("example.com", "example.net", "example.org")
RESERVED_TLDS = ("test", "example", "invalid", "localhost")

# Any host carrying a bare `example` label is refused too — the protocol says "example.*
# generally". Matching whole labels rather than substrings keeps `myexample.com` clear, and
# refusing `example.acme.com` as well is deliberate: this check exists to catch namespaces
# that look real and are not, so it errs toward the false positive an encoder can override
# by choosing a name that does not read as a placeholder.
PLACEHOLDER_LABEL = "example"

_BASE_URI_LINE = re.compile(r"^\s*base_uri\s*[:=]\s*(\S+)", re.IGNORECASE | re.MULTILINE)


@dataclass(frozen=True)
class CheckResult:
    """One check, its verdict, and enough detail to act on a failure."""

    name: str
    passed: bool
    detail: str = ""
    skipped: bool = False


def _sheet_header_row(ws, first_header: str) -> int | None:
    for row in range(1, 12):
        for col in range(1, 40):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.strip() == first_header:
                return row
    return None


def _anchor_col(ws) -> int | None:
    for col in range(1, 40):
        for row in range(1, 12):
            if ws.cell(row=row, column=col).value == ANCHOR_HEADER:
                return col
    return None


# Sheet -> the header string that identifies its header row, and the first data row
# offset below it. The description row sits between them.
_SHEETS = {
    SHEET_NAMES["summary"]: ("Project Name", 2),
    SHEET_NAMES["model_data"]: ("Entity Type", 2),
    SHEET_NAMES["validation"]: ("Result Name", 2),
    SHEET_NAMES["factors"]: ("Factor Type", 2),
    SHEET_NAMES["decision"]: ("Decision Outcome", 2),
}


def _template_placeholders(template_path: Path | None) -> set[str]:
    """The hint strings a pack's own template puts in its description rows."""
    if not template_path or not template_path.exists():
        return set()
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    hints: set[str] = set()
    for sheet, (first_header, _) in _SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        head = _sheet_header_row(ws, first_header)
        if head is None:
            continue
        for col in range(1, 40):
            value = ws.cell(row=head + 1, column=col).value
            if isinstance(value, str) and len(value.strip()) > 3:
                hints.add(value.strip())
    return hints


def check_workbook(path: Path, template_path: Path | None = None) -> list[CheckResult]:
    """Workbook-side checks. See the module docstring for what these assume."""
    import openpyxl

    results: list[CheckResult] = []
    wb = openpyxl.load_workbook(path)
    placeholders = _template_placeholders(template_path)

    missing_col: list[str] = []
    unanchored: list[str] = []
    leaked: list[str] = []

    for sheet, (first_header, data_offset) in _SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        head = _sheet_header_row(ws, first_header)
        if head is None:
            continue
        acol = _anchor_col(ws)
        if acol is None:
            missing_col.append(sheet)
        start = head + data_offset
        for row in range(start, ws.max_row + 1):
            limit = acol if acol else 40
            values = [ws.cell(row=row, column=c).value for c in range(1, limit)]
            populated = [v for v in values if v not in (None, "")]
            if not populated:
                continue
            if acol and not ws.cell(row=row, column=acol).value:
                unanchored.append(f"{sheet} row {row}")
            for v in populated:
                if isinstance(v, str) and v.strip() in placeholders:
                    leaked.append(f"{sheet} row {row}: {v.strip()[:40]}")

    results.append(CheckResult(
        "anchor column present",
        not missing_col,
        "missing on " + ", ".join(missing_col) if missing_col else "on every data sheet",
    ))
    results.append(CheckResult(
        "anchor non-empty per populated row",
        not unanchored,
        f"{len(unanchored)} row(s) unanchored: " + "; ".join(unanchored[:4])
        if unanchored else "every populated row carries an anchor",
    ))
    results.append(CheckResult(
        "no template placeholder text in data rows",
        not leaked,
        f"{len(leaked)} leak(s): " + "; ".join(leaked[:3]) if leaked
        else ("clean" if placeholders else "no template available to compare against"),
        skipped=not placeholders,
    ))
    results.append(_check_levels(wb))
    return results


#: Provenance values that record the required-level SUFFICIENCY judgment having
#: happened: the level was weighed against the achieved one and agreed with, or
#: changed.
#:
#: **`confirmed` is deliberately absent.** It records that a cell was checked
#: against the source -- the LOCATION question -- and the encoding tool produces
#: it as a side effect of anchoring. A first draft of this check counted it, and
#: reconstructing run 25's package proved the consequence: seventeen levels
#: anchored, none weighed, tokens all `confirmed`, and the check PASSED the
#: exact package it was built to refuse. Strictly worse than the shape
#: heuristic it replaced.
#:
#: Two questions, two acts: anchoring answers *was it located*, affirming
#: answers *was it judged*.
_JUDGED = {"affirmed", "corrected"}

#: **A recorded absence is not a missing judgment.**
#:
#: Two acts say "there is no required level here to weigh, and a human
#: established that": `source-absent` -- the document does not state this
#: requirement -- and `not-recoverable` -- it does, and the admitted text
#: cannot carry it. Neither is a judgment. Neither is a silence either, which
#: is the half this checker used to get wrong: both landed in the same bucket
#: as a level nobody opened, so the honest act read identically to negligence
#: and the reviewer was left no lawful exit at all.
#:
#: A disposed requirement is EXCLUDED FROM THE DENOMINATOR. Exclusion is not
#: judgment credit: nothing is counted as weighed that was not weighed. The
#: package says a smaller number of levels were judged, and says why the
#: number is smaller.
#:
#: Which is the other half of the rule. Every message below NAMES the
#: disposals. A denominator that quietly shrinks is precisely the vacuity this
#: checker exists to refuse -- a check correct on the ordinary case and silent
#: on the important one -- so a check that excludes must state what it
#: excluded, how many, and under which term.
_DISPOSED = {"source-absent", "not-recoverable"}

#: Which provenance tokens each encoding profile can state, and from when --
#: **imported, never restated.** The emitter and this checker disagreeing about
#: a closed set is the exact defect shape this repository keeps finding: two
#: copies that agree until one grows. `level_vocab_for` is that one definition;
#: see `excel_constants.LEVEL_VOCAB` for why the versions are keyed.
#:
#: It mirrors `uofa:introducedIn` on the shapes: the same rule, for the same
#: reason, one layer down.
_vocab_for = level_vocab_for


def _split_levels(rows: list[tuple[str, str]], vocab: frozenset[str],
                  judged: set[str] | frozenset[str]):
    """Sort `(factor, token)` pairs into judged / disposed / unjudged / alien.

    `alien` is tested FIRST and on its own: a term whose meaning arrived AFTER
    the version the artifact declares must not be silently sorted into one of
    the other three, because every one of those three is a claim about what the
    reviewer did, and a term the declaration cannot express supports none of
    them.

    It is deliberately the narrow set -- `level_terms_after`, not "everything
    unrecognised". A token this vocabulary has never contained at any version
    is not a version disagreement; it is simply not a judgment, and belongs in
    `unjudged` with the message that says so.
    """
    future = level_terms_after_vocab(vocab)
    weighed: list[str] = []
    disposed: list[tuple[str, str]] = []
    unjudged: list[str] = []
    alien: list[tuple[str, str]] = []
    for factor, raw in rows:
        token = (raw or "").strip().lower()
        if token in future:
            alien.append((factor, token))
        elif token in judged:
            weighed.append(factor)
        elif token in _DISPOSED:
            disposed.append((factor, token))
        else:
            unjudged.append(factor)
    return weighed, disposed, unjudged, alien


def _disposal_note(disposed: list[tuple[str, str]]) -> str:
    """What the denominator dropped, stated whenever it dropped anything."""
    if not disposed:
        return ""
    counts = Counter(token for _, token in disposed)
    kinds = ", ".join(f"{n} {token}" for token, n in sorted(counts.items()))
    return (f"; {len(disposed)} excluded from the denominator as disposed "
            f"({kinds}) -- a recorded absence is not a judgment")


def _alien_refusal(name: str, declared: str, alien: list[tuple[str, str]]) -> CheckResult:
    """A term the artifact's own declaration cannot express."""
    shown = "; ".join(f"{factor} carries '{token}'" for factor, token in alien[:3])
    more = f" (+{len(alien) - 3} more)" if len(alien) > 3 else ""
    return CheckResult(
        name, False,
        f"this package declares {declared}, whose provenance vocabulary does "
        f"not contain the term(s) it uses: {shown}{more}. The declaration and "
        f"the artifact disagree, so the term can be read neither as a judgment "
        f"nor as a disposition -- {declared} cannot say what it means. "
        f"Re-export from an encoding tool that declares the profile it writes.")


def _declared_profile(wb) -> str:
    """What the workbook says its own encoding shape is, or "" if it says nothing.

    Read from a DECLARATION rather than inferred from which columns happen to be
    present. The difference is not cosmetic: an encoder that writes v0.8 and
    omits the provenance column is broken, and shape-inference reads that
    identical to a legacy workbook and excuses it. A declaration lets the check
    tell those apart -- which is the whole reason the marker exists.
    """
    sheet = SHEET_NAMES.get("summary", "Assessment Summary")
    if sheet not in wb.sheetnames:
        return ""
    ws = wb[sheet]
    head = _sheet_header_row(ws, "Project Name")
    if head is None:
        return ""
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=head, column=col).value or "").strip() == WORKBOOK_PROFILE_HEADER:
            return str(ws.cell(row=head + 1, column=col).value or "").strip()
    return ""


#: `"v0.8"` -> `(0, 8)`, from the one definition that also keys the vocabulary.
_version_tuple = version_tuple


#: The first shape that can state whether a required level was judged.
_EVIDENCE_PROFILE = (0, 8)


def _levels_column(ws, head: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=head, column=col).value or "").strip() == LEVEL_PROVENANCE_HEADER:
            return col
    return None


def _check_levels(wb) -> CheckResult:
    """Was the required-level column REVIEWED? Read the evidence, not the shape.

    This asked whether required differs from achieved anywhere, on the reasoning
    that the extract prompt sets them equal by default, so all-equal means
    nobody looked. The reasoning is sound and the test is a proxy: **agreement
    writes nothing**, so a reviewer who read all seventeen levels and agreed
    with every one produces a workbook byte-identical to one nobody opened.

    Shape-inference therefore punishes honest agreement and misses nothing else.
    Run 25 is the case that made it worth fixing rather than tolerating: the
    reviewer really had left the defaults untouched, so the refusal was correct
    -- and the same reading would have refused a diligent reviewer with the same
    values.

    So a package that can SAY whether the judgment happened is asked; one that
    cannot is warned, not refused. A third party scripting against exit codes
    keeps its contract on the legacy path.
    """
    sheet = SHEET_NAMES["factors"]
    if sheet not in wb.sheetnames:
        return CheckResult("required levels were reviewed", True,
                           "no factors sheet", skipped=True)
    ws = wb[sheet]
    head = _sheet_header_row(ws, "Factor Type")
    if head is None:
        return CheckResult("required levels were reviewed", True,
                           "no factor header", skipped=True)

    prov_col = _levels_column(ws, head)
    pairs, differing, waived = 0, 0, False
    rows: list[tuple[str, str]] = []

    for row in range(head + 2, ws.max_row + 1):
        factor = ws.cell(row=row, column=1).value
        if not factor:
            continue
        req, ach = ws.cell(row=row, column=3).value, ws.cell(row=row, column=4).value
        for col in (5, 6):
            text = ws.cell(row=row, column=col).value
            if isinstance(text, str) and "waiv" in text.lower():
                waived = True
        if prov_col is not None and req is not None:
            rows.append((str(factor),
                         str(ws.cell(row=row, column=prov_col).value or "")))
        if req is None or ach is None:
            continue
        pairs += 1
        if req != ach:
            differing += 1

    # ── which path, decided by what the workbook DECLARES ──────────────────
    #
    # Not by whether the column is present. An encoder that declares v0.8 and
    # ships no provenance column is broken, and shape-inference cannot tell that
    # from a legacy workbook -- it excuses both. The declaration separates "this
    # sheet cannot speak about judgment" from "it can and did not".
    declared = _declared_profile(wb)
    if _version_tuple(declared) >= _EVIDENCE_PROFILE and prov_col is None:
        return CheckResult(
            "required levels were reviewed", False,
            f"this workbook declares encoding profile {declared}, which carries "
            f"a '{LEVEL_PROVENANCE_HEADER}' column, and it has none. The "
            f"declaration and the sheet disagree; re-export from the encoding "
            f"tool rather than trusting either.")

    # ── the evidence path: the package can state whether judgment occurred ──
    if _version_tuple(declared) >= _EVIDENCE_PROFILE:
        name = "required levels were reviewed"
        weighed, disposed, unjudged, alien = _split_levels(
            rows, _vocab_for(_version_tuple(declared)), _JUDGED)
        if alien:
            return _alien_refusal(name, f"encoding profile {declared}", alien)

        # The denominator is what was ASKED of a reviewer, which is every rated
        # level minus the ones a reviewer disposed of. `pairs` counted rows
        # carrying both levels and was never the right roster for this
        # question: it answered "how many can be compared", not "how many
        # needed weighing".
        owed = len(weighed) + len(unjudged)
        note = _disposal_note(disposed)
        if unjudged and not waived:
            shown = "; ".join(unjudged[:4])
            more = f" (+{len(unjudged) - 4} more)" if len(unjudged) > 4 else ""
            return CheckResult(
                name, False,
                f"{len(unjudged)} of {owed} required level(s) have not been "
                f"affirmed or corrected -- nobody has weighed them against the "
                f"achieved level: {shown}{more}. Affirm each in the encoding "
                f"tool, correct it, or record a waiver.{note}")

        # Every level disposed and none judged. This is not a judgment pass and
        # must never render as one: the roster was answered, and the answer was
        # that nothing in it could be weighed. Stated as an advisory so a reader
        # sees the sentence rather than a green tick standing in for it.
        if disposed and not weighed:
            counts = Counter(token for _, token in disposed)
            kinds = ", ".join(f"{n} {token}" for token, n in sorted(counts.items()))
            return CheckResult(
                name, True,
                f"no required level was judged: all {len(disposed)} were "
                f"disposed ({kinds}). Nothing here was weighed against the "
                f"achieved level, and this package does not claim otherwise",
                skipped=True)

        return CheckResult(
            name, True,
            f"every required level carries a review act"
            + (f" ({differing} of {pairs} differ)" if differing else
               f" (all {pairs} agree, and the package says so)")
            + (", waiver recorded" if waived else "") + note)

    # ── the legacy path: advisory, never a hard refusal on its own ─────────
    #
    # A workbook carrying the column but declaring nothing lands here, and the
    # column is deliberately NOT read: an undeclared sheet cannot vouch for what
    # its own column means, and reading it anyway is the field-sniffing the
    # declaration exists to replace. What it must never be is silent, so the
    # advisory says the column was seen and why it went unused.
    seen_but_undeclared = (
        f" A '{LEVEL_PROVENANCE_HEADER}' column is present but this workbook "
        f"declares no encoding profile, so nothing vouches for what it records; "
        f"re-export from the encoding tool to have it read."
        if prov_col is not None else "")
    if pairs == 0:
        return CheckResult("required levels were reviewed", True,
                           "no factor carries both levels", skipped=True)
    if differing or waived:
        return CheckResult("required levels were reviewed", True,
                           f"{differing} of {pairs} differ"
                           + (", waiver recorded" if waived else ""))
    return CheckResult(
        "required levels were reviewed", True,
        f"required equals achieved on all {pairs} factor(s), and this package's "
        f"profile cannot state whether the column was reviewed{seen_but_undeclared} -- treat as "
        f"unreviewed pending confirmation", skipped=True)


def _reserved_reason(uri: str | None) -> str | None:
    """Why this namespace is reserved, or None if it is not one of the reserved family."""
    if not isinstance(uri, str) or not uri.strip():
        return None
    host = (urlsplit(uri.strip()).hostname or "").lower().rstrip(".")
    if not host:
        return None
    labels = host.split(".")
    if host in RESERVED_SECOND_LEVEL or any(
        host.endswith("." + d) for d in RESERVED_SECOND_LEVEL
    ):
        return f"{host} is reserved by RFC 2606 for documentation"
    if labels[-1] in RESERVED_TLDS:
        return f".{labels[-1]} is a reserved special-use TLD (RFC 6761)"
    if PLACEHOLDER_LABEL in labels:
        return f"{host} carries a bare '{PLACEHOLDER_LABEL}' label"
    return None


def _package_id(package_path: Path) -> str | None:
    """The package's own minted identifier, or None if it carries no readable one."""
    try:
        doc = json.loads(package_path.read_text(encoding="utf8", errors="replace"))
    except (OSError, ValueError):
        return None
    if not isinstance(doc, dict):
        return None
    for key in ("id", "@id"):
        value = doc.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _check_namespace(package_path: Path, run_log_text: str | None) -> CheckResult:
    """The minted namespace must be one the encoder controls.

    Checks the package's own `id` first, because that is the string the hash and signature
    cover, then the run log's declared `base_uri`. Either being reserved fails the check;
    neither being present skips it rather than passing vacuously.
    """
    name = "namespace is not a reserved example domain"
    candidates: list[tuple[str, str]] = []
    package_id = _package_id(package_path)
    if package_id:
        candidates.append(("package id", package_id))
    if run_log_text:
        match = _BASE_URI_LINE.search(run_log_text)
        if match:
            candidates.append(("run log base_uri", match.group(1)))

    if not candidates:
        return CheckResult(name, True, "no minted identifier or declared base_uri to read",
                           skipped=True)

    offending = [
        f"{label} {value}: {reason}"
        for label, value in candidates
        if (reason := _reserved_reason(value))
    ]
    if offending:
        return CheckResult(name, False, "; ".join(offending))
    return CheckResult(name, True, ", ".join(f"{label} {value}" for label, value in candidates))


#: Judgment tokens, at the package layer. Same set as the workbook's, stated
#: against the exported vocabulary rather than the tool's internal terms.
_PACKAGE_JUDGED = {"affirmed", "corrected", "waived"}


def _context_version(doc: dict) -> tuple[int, ...]:
    """The version the PACKAGE declares, from its own `@context`.

    Keyed on the context URL rather than `conformsToProfile`: that term carries
    the profile (Minimal/Complete/Disposition) and encodes no context version at
    all, so it cannot answer "can this package state whether a level was
    judged?". The context is the term that actually moves when the vocabulary
    does.

    An inlined context (a resolved or signed document) declares no version here;
    it returns () and takes the advisory path rather than being guessed at.
    """
    ref = doc.get("@context")
    if not isinstance(ref, str):
        return ()
    return _version_tuple(ref.rsplit("/", 1)[-1])


def _check_package_levels(package_path: Path) -> CheckResult:
    """Did anyone weigh the required levels? Asked only of packages that can answer.

    The fork is the point. A package whose context predates the vocabulary has
    no way to record the judgment, so refusing it would punish age rather than
    negligence -- and third parties script against these exit codes. A package
    whose context DOES carry the term is asked, and silence there is a real
    answer: nobody went back to the extractor's guesses.
    """
    name = "required levels were judged"
    try:
        doc = json.loads(package_path.read_text(encoding="utf8"))
    except (OSError, json.JSONDecodeError) as exc:
        return CheckResult(name, False, f"could not read the package: {exc}")

    version = _context_version(doc)
    factors = [f for f in doc.get("hasCredibilityFactor", []) if isinstance(f, dict)]
    rated = [f for f in factors if f.get("requiredLevel") is not None]

    if version < _EVIDENCE_PROFILE:
        declared = str(doc.get("@context") or "none").rsplit("/", 1)[-1]
        return CheckResult(
            name, True,
            f"this package declares context {declared}, whose vocabulary cannot "
            f"state whether a required level was judged -- treat as unreviewed "
            f"pending confirmation", skipped=True)

    if not rated:
        return CheckResult(name, True, "no factor carries a required level",
                           skipped=True)

    declared = str(doc.get("@context") or "none").rsplit("/", 1)[-1]
    weighed, disposed, unjudged, alien = _split_levels(
        [(str(f.get("factorType") or "?"),
          str(f.get("requiredLevelProvenance") or "")) for f in rated],
        _vocab_for(version), _PACKAGE_JUDGED)
    if alien:
        return _alien_refusal(name, f"context {declared}", alien)

    owed = len(weighed) + len(unjudged)
    note = _disposal_note(disposed)
    if unjudged:
        shown = "; ".join(unjudged[:4])
        more = f" (+{len(unjudged) - 4} more)" if len(unjudged) > 4 else ""
        return CheckResult(
            name, False,
            f"{len(unjudged)} of {owed} required level(s) carry no "
            f"judgment: {shown}{more}. Anchoring locates a level; it does not "
            f"weigh it. Affirm or correct each in the encoding tool, or record "
            f"a waiver.{note}")

    # All disposed, none judged -- see the twin branch in `_check_levels`. A
    # green tick here would let a package that weighed nothing read exactly like
    # one that weighed everything, which is the whole failure this check exists
    # to prevent, arrived at from the other side.
    if disposed and not weighed:
        counts = Counter(token for _, token in disposed)
        kinds = ", ".join(f"{n} {token}" for token, n in sorted(counts.items()))
        return CheckResult(
            name, True,
            f"no required level carries a judgment: all {len(disposed)} were "
            f"disposed ({kinds}). This package records that nothing could be "
            f"weighed, and claims no judgment",
            skipped=True)

    return CheckResult(name, True,
                       f"all {owed} required level(s) carry a judgment{note}")


def check_package(package_path: Path) -> list[CheckResult]:
    """Package-side checks, run against the artifacts committed beside the package."""
    results: list[CheckResult] = []
    directory = package_path.parent

    def _find(names: tuple[str, ...]) -> Path | None:
        for child in directory.iterdir():
            if child.is_file() and child.name.lower() in names:
                return child
        return None

    log = _find(AMBIGUITY_LOG_NAMES)
    if log is None:
        results.append(CheckResult("ambiguity log present", False,
                                   f"no ambiguity log beside {package_path.name}"))
    else:
        body = log.read_text(encoding="utf8", errors="replace").strip()
        results.append(CheckResult("ambiguity log present and non-empty",
                                   bool(body), f"{log.name}, {len(body)} chars"))

    run_log = _find(RUN_LOG_NAMES)
    if run_log is None:
        results.append(CheckResult("run log present", False,
                                   f"no run log beside {package_path.name}"))
        results.append(CheckResult("run log carries its pins", False, "no run log"))
        results.append(CheckResult("no signing in a pilot run log", True,
                                   "no run log", skipped=True))
        results.append(_check_namespace(package_path, None))
        return results

    text = run_log.read_text(encoding="utf8", errors="replace")
    lowered = text.lower()
    results.append(CheckResult("run log present", True, run_log.name))

    absent = [f for f in RUN_LOG_FIELDS if f.lower() not in lowered]
    results.append(CheckResult(
        "run log carries its pins", not absent,
        "missing " + ", ".join(absent) if absent else ", ".join(RUN_LOG_FIELDS),
    ))

    is_pilot = any(m.lower() in lowered for m in PILOT_MARKERS)
    if not is_pilot:
        results.append(CheckResult("no signing in a pilot run log", True,
                                   "run log is not pilot-labeled", skipped=True))
    else:
        # A mention inside a prohibition ("no --sign", "without signing") is not a use.
        offending = [
            line.strip() for line in text.splitlines()
            if "--sign" in line
            and not any(n in line.lower() for n in ("no --sign", "without", "never", "not "))
        ]
        results.append(CheckResult(
            "no signing in a pilot run log", not offending,
            "; ".join(offending[:2])[:120] if offending else "pilot run log records no signing",
        ))
    results.append(_check_namespace(package_path, text))
    results.append(_check_package_levels(package_path))
    return results


def render(results: list[CheckResult], title: str) -> bool:
    """Print the pass/fail table. Returns True when nothing failed."""
    from uofa_cli.output import step_header, result_line, info

    step_header(f"protocol-check: {title}")
    for r in results:
        if r.skipped:
            info(f"  - {r.name}  ({r.detail})")
        else:
            result_line(r.name, r.passed, r.detail)
    failed = [r for r in results if not r.passed and not r.skipped]
    if failed:
        info(f"  {len(failed)} check(s) failed")
    return not failed
