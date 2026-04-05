"""Read and validate an Excel workbook for UofA import.

Knows about Excel structure (sheets, rows, cells) but nothing about JSON-LD.
Returns a clean intermediate dict that excel_mapper.py transforms into JSON-LD.
"""

from pathlib import Path

from uofa_cli.excel_constants import (
    SHEET_NAMES, FACTOR_START_ROW,
    VV40_FACTOR_NAMES, NASA_ALL_FACTOR_NAMES, NASA_ONLY_FACTOR_NAMES,
    VV40_LEVEL_RANGE, NASA_LEVEL_RANGE,
    VALID_PROFILES, VALID_DECISION_OUTCOMES, VALID_FACTOR_STATUSES,
    VALID_DEVICE_CLASSES, VALID_ASSURANCE_LEVELS,
    EVIDENCE_TYPES,
)


class ImportError(Exception):
    """Raised when an Excel workbook has validation errors."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))


def _cell_ref(col: int, row: int) -> str:
    """Convert 1-based column index to Excel cell reference like A3, B5."""
    return f"{chr(64 + col)}{row}"


def _cell_value(ws, row: int, col: int) -> str | None:
    """Read a cell value, stripping whitespace. Returns None for empty cells."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def _parse_int(val: str | None) -> int | None:
    """Parse a value as integer, returning None if not parseable."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> str | None:
    """Normalize a date value to ISO 8601 string."""
    if val is None:
        return None
    # openpyxl may return datetime objects directly
    import datetime
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    # Try ISO 8601 first
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Excel serial date number
    try:
        serial = float(s)
        if 1 < serial < 100000:
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(serial))).isoformat()
    except (ValueError, OverflowError):
        pass
    return s  # Return as-is if we can't parse


def read_workbook(xlsx_path: Path, packs: list[str]) -> dict:
    """Read and validate an Excel workbook.

    Returns an intermediate dict with keys:
        summary, entities, validation_results, factors, decision

    Raises ImportError with a list of validation errors.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError([
            "openpyxl is required for Excel import. "
            "Install with: pip install uofa-cli[excel]"
        ])

    if not xlsx_path.exists():
        raise ImportError([f"File not found: {xlsx_path}"])

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception as exc:
        raise ImportError([f"Cannot open workbook: {exc}"])

    errors = []

    # ── Validate required sheets ─────────────────────────────
    required_sheets = [
        SHEET_NAMES["summary"],
        SHEET_NAMES["model_data"],
        SHEET_NAMES["validation"],
        SHEET_NAMES["decision"],
    ]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Sheet '{sheet_name}' not found in workbook")

    if errors:
        raise ImportError(errors)

    # ── Read Assessment Summary (row 3) ──────────────────────
    ws = wb[SHEET_NAMES["summary"]]
    summary = _read_summary(ws, errors)
    profile = summary.get("profile", "Minimal")

    # Credibility Factors is only required for Complete profile
    has_factors_sheet = SHEET_NAMES["factors"] in wb.sheetnames
    if profile == "Complete" and not has_factors_sheet:
        errors.append(f"Sheet '{SHEET_NAMES['factors']}' not found (required for Complete profile)")

    # ── Read Model & Data ────────────────────────────────────
    ws = wb[SHEET_NAMES["model_data"]]
    entities = _read_entities(ws, profile, errors)

    # ── Read Validation Results ──────────────────────────────
    ws = wb[SHEET_NAMES["validation"]]
    validation_results = _read_validation_results(ws, errors)

    # ── Read Credibility Factors ─────────────────────────────
    factors = []
    if has_factors_sheet:
        ws = wb[SHEET_NAMES["factors"]]
        factors = _read_factors(ws, packs, errors)

    # ── Read Decision ────────────────────────────────────────
    ws = wb[SHEET_NAMES["decision"]]
    decision = _read_decision(ws, errors)

    if errors:
        raise ImportError(errors)

    return {
        "summary": summary,
        "entities": entities,
        "validation_results": validation_results,
        "factors": factors,
        "decision": decision,
    }


def _find_data_row(ws, header_keyword: str, search_col: int = 1, max_row: int = 10) -> int:
    """Find the first data row after a header row containing the keyword.

    Scans column search_col for a cell matching header_keyword (case-insensitive).
    Returns the row AFTER the last header/instruction row (i.e., the first row
    that doesn't look like a header).

    Heuristic: the data row is the last non-empty row in the first block.
    For templates with instruction rows, the data row follows the instructions.
    """
    for r in range(1, max_row + 1):
        val = _cell_value(ws, r, search_col)
        if val and header_keyword.lower() in val.lower():
            # Found header row. Data is the next non-instruction row.
            # Check if there's an instruction row right after
            for data_r in range(r + 1, max_row + 1):
                val = _cell_value(ws, data_r, search_col)
                if val and val not in ("", None):
                    # Check if this looks like actual data vs instructions
                    # Instructions tend to be long descriptions starting with verbs/articles
                    # Data tends to be short identifiers
                    # But the safest heuristic: the last row before an empty row
                    return data_r
            return r + 1
    # Fallback: assume row 3 has headers, row 4 has data
    return 4


def _find_header_row(ws, header_keyword: str, search_col: int = 1, max_row: int = 10) -> int:
    """Find the row containing column headers."""
    for r in range(1, max_row + 1):
        val = _cell_value(ws, r, search_col)
        if val and header_keyword.lower() == val.lower():
            return r
    return 2  # default


def _read_summary(ws, errors: list) -> dict:
    """Read Assessment Summary sheet.

    Finds the header row (containing "Project Name") and reads data from
    the last populated row in the header block.
    """
    sheet = SHEET_NAMES["summary"]

    # Find header row by looking for "Project Name"
    header_row = _find_header_row(ws, "Project Name")
    # Data row: scan forward from header to find actual data
    # In the starter template: row 2 = headers, row 3 = instructions, row 4 = data
    # In the spec: row 2 = headers, row 3 = data
    # Strategy: find the LAST non-empty row starting from header+1
    row = header_row + 1
    for r in range(header_row + 1, header_row + 4):
        val = _cell_value(ws, r, 1)
        if val:
            row = r  # keep advancing to last non-empty row

    project_name = _cell_value(ws, row, 1)  # A
    cou_name = _cell_value(ws, row, 2)       # B
    cou_description = _cell_value(ws, row, 3) # C
    profile = _cell_value(ws, row, 4)         # D
    device_class = _cell_value(ws, row, 5)    # E
    mrl = _cell_value(ws, row, 6)             # F
    assurance_level = _cell_value(ws, row, 7) # G
    standards_ref = _cell_value(ws, row, 8)   # H
    assessor_name = _cell_value(ws, row, 9)   # I
    assessment_date = ws.cell(row=row, column=10).value  # J — raw for date parsing
    source_doc = _cell_value(ws, row, 11)     # K
    has_uq = _cell_value(ws, row, 12)         # L

    # Required for Minimal
    if not project_name:
        errors.append(f"Sheet '{sheet}', cell {_cell_ref(1, row)} (Project Name) is required for Minimal profile")
    if not cou_name:
        errors.append(f"Sheet '{sheet}', cell {_cell_ref(2, row)} (COU Name) is required")

    # Validate profile
    if profile and profile not in VALID_PROFILES:
        errors.append(f"Sheet '{sheet}', cell {_cell_ref(4, row)}: '{profile}' is not a valid profile. Expected: {', '.join(VALID_PROFILES)}")
    if not profile:
        profile = "Minimal"

    # Validate device class
    if device_class and device_class not in VALID_DEVICE_CLASSES:
        # Allow free text (Category A-E, Other) per spec
        pass

    # Validate assurance level
    if assurance_level and assurance_level not in VALID_ASSURANCE_LEVELS:
        errors.append(f"Sheet '{sheet}', cell {_cell_ref(7, row)}: '{assurance_level}' is not a valid assurance level. Expected: {', '.join(VALID_ASSURANCE_LEVELS)}")

    # Parse MRL
    mrl_int = None
    if mrl:
        # Handle "MRL 3" format
        mrl_str = mrl.replace("MRL", "").strip()
        mrl_int = _parse_int(mrl_str)

    return {
        "project_name": project_name,
        "cou_name": cou_name,
        "cou_description": cou_description,
        "profile": profile,
        "device_class": device_class,
        "model_risk_level": mrl_int,
        "assurance_level": assurance_level,
        "standards_reference": standards_ref,
        "assessor_name": assessor_name,
        "assessment_date": _parse_date(assessment_date),
        "source_document": source_doc,
        "has_uq": has_uq,
    }


def _read_entities(ws, profile: str, errors: list) -> list[dict]:
    """Read Model & Data sheet."""
    sheet = SHEET_NAMES["model_data"]
    entities = []
    has_requirement = False

    # Find header row containing "Entity Type"
    header_row = _find_header_row(ws, "Entity Type")
    # Skip instruction row if present
    data_start = header_row + 1
    for r in range(header_row + 1, header_row + 4):
        val = _cell_value(ws, r, 1)
        if val and val in ("Requirement", "Model", "Dataset"):
            data_start = r
            break

    for row in range(data_start, ws.max_row + 1):
        entity_type = _cell_value(ws, row, 1)  # A
        if not entity_type:
            continue

        if entity_type not in ("Requirement", "Model", "Dataset"):
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(1, row)}: "
                f"'{entity_type}' is not a valid entity type. "
                f"Expected: Requirement, Model, Dataset"
            )
            continue

        if entity_type == "Requirement":
            has_requirement = True

        name = _cell_value(ws, row, 2)     # B
        uri = _cell_value(ws, row, 3)      # C
        desc = _cell_value(ws, row, 4)     # D
        version = _cell_value(ws, row, 5)  # E
        source = _cell_value(ws, row, 6)   # F

        if not name:
            errors.append(f"Sheet '{sheet}', cell {_cell_ref(2, row)}: Name is required")

        entities.append({
            "entity_type": entity_type,
            "name": name,
            "uri": uri,
            "description": desc,
            "version": version,
            "source": source,
        })

    if not has_requirement:
        errors.append(
            f"Sheet '{sheet}' must have at least one row with Entity Type = 'Requirement'"
        )

    return entities


def _read_validation_results(ws, errors: list) -> list[dict]:
    """Read Validation Results sheet.

    Detects whether the Type column (v2) is present by checking headers.
    Old template: Result Name, Identifier/URI, Description, ...
    v2 template:  Result Name, Type, Identifier/URI, Description, ...
    """
    sheet = SHEET_NAMES["validation"]
    results = []

    # Detect column layout by scanning header rows for "Type"
    has_type_col = False
    header_row = 2  # default
    for r in range(1, 5):
        for c in range(1, 10):
            val = _cell_value(ws, r, c)
            if val and val.lower() == "type":
                has_type_col = True
                header_row = r
                break
            if val and val.lower() == "result name":
                header_row = r
        if has_type_col:
            break

    # Find first data row (skip instruction rows)
    data_start = header_row + 1
    for r in range(header_row + 1, header_row + 4):
        val = _cell_value(ws, r, 1)
        if val and not val.lower().startswith("short name"):
            data_start = r
            break

    if has_type_col:
        # v2 layout: A=Name, B=Type, C=URI, D=Desc, E=ComparesTo, F=HasUQ, G=UQMethod, H=Metric, I=PassFail
        col_type, col_uri, col_desc, col_cmp, col_uq, col_uqm, col_met, col_pf = 2, 3, 4, 5, 6, 7, 8, 9
    else:
        # Old layout: A=Name, B=URI, C=Desc, D=ComparesTo, E=HasUQ, F=UQMethod, G=Metric, H=PassFail
        col_type, col_uri, col_desc, col_cmp, col_uq, col_uqm, col_met, col_pf = None, 2, 3, 4, 5, 6, 7, 8

    for row in range(data_start, ws.max_row + 1):
        name = _cell_value(ws, row, 1)  # A — Result Name
        if not name:
            continue

        evidence_type = _cell_value(ws, row, col_type) if col_type else None
        uri = _cell_value(ws, row, col_uri)
        desc = _cell_value(ws, row, col_desc)
        compares_to = _cell_value(ws, row, col_cmp)
        has_uq = _cell_value(ws, row, col_uq)
        uq_method = _cell_value(ws, row, col_uqm)
        metric = _cell_value(ws, row, col_met)
        pass_fail = _cell_value(ws, row, col_pf)

        # Default evidence type to ValidationResult
        if not evidence_type:
            evidence_type = "ValidationResult"
        elif evidence_type not in EVIDENCE_TYPES:
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(col_type, row)}: "
                f"'{evidence_type}' is not a valid evidence type. "
                f"Expected: {', '.join(EVIDENCE_TYPES)}"
            )

        results.append({
            "name": name,
            "evidence_type": evidence_type,
            "uri": uri,
            "description": desc,
            "compares_to": compares_to,
            "has_uq": has_uq,
            "uq_method": uq_method,
            "metric_value": metric,
            "pass_fail": pass_fail,
        })

    return results


def _read_factors(ws, packs: list[str], errors: list) -> list[dict]:
    """Read Credibility Factors sheet.

    Factor Type (col A) and Category (col B) are pre-populated and locked.
    User fills in Required Level (C), Achieved Level (D), Acceptance Criteria (E),
    Rationale (F), Factor Status (G).
    """
    sheet = SHEET_NAMES["factors"]
    factors = []

    # Determine valid factor names based on active packs
    if "nasa-7009b" in packs:
        valid_names = set(NASA_ALL_FACTOR_NAMES)
    else:
        valid_names = set(VV40_FACTOR_NAMES)

    for row in range(FACTOR_START_ROW, ws.max_row + 1):
        factor_type = _cell_value(ws, row, 1)  # A
        if not factor_type:
            continue

        category = _cell_value(ws, row, 2)         # B
        required_level = _cell_value(ws, row, 3)   # C
        achieved_level = _cell_value(ws, row, 4)   # D
        acceptance = _cell_value(ws, row, 5)        # E
        rationale = _cell_value(ws, row, 6)         # F
        status = _cell_value(ws, row, 7)            # G
        linked_evidence = _cell_value(ws, row, 8)   # H — Linked Evidence URI

        # Validate factor type
        if factor_type not in valid_names:
            if "nasa-7009b" in packs:
                errors.append(
                    f"Sheet '{sheet}', row {row}: "
                    f"'{factor_type}' is not a valid NASA-STD-7009B factor type"
                )
            else:
                errors.append(
                    f"Sheet '{sheet}', row {row}: "
                    f"'{factor_type}' is not a valid V&V 40 factor type"
                )
            continue

        # Validate factor status
        if status and status not in VALID_FACTOR_STATUSES:
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(7, row)}: "
                f"'{status}' is not a valid factor status. "
                f"Expected: {', '.join(VALID_FACTOR_STATUSES)}"
            )

        # Parse and validate levels
        req_int = _parse_int(required_level)
        ach_int = _parse_int(achieved_level)

        # Determine level range based on factor standard
        is_nasa_only = factor_type in NASA_ONLY_FACTOR_NAMES
        if is_nasa_only:
            lo, hi = NASA_LEVEL_RANGE
        else:
            lo, hi = VV40_LEVEL_RANGE

        if req_int is not None and (req_int < lo or req_int > hi):
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(3, row)}: "
                f"Required Level {req_int} out of range {lo}-{hi}"
            )
        if ach_int is not None and (ach_int < lo or ach_int > hi):
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(4, row)}: "
                f"Achieved Level {ach_int} out of range {lo}-{hi}"
            )

        factors.append({
            "factor_type": factor_type,
            "category": category,
            "required_level": req_int,
            "achieved_level": ach_int,
            "acceptance_criteria": acceptance,
            "rationale": rationale,
            "status": status or "not-assessed",
            "linked_evidence": linked_evidence,
        })

    return factors


def _read_decision(ws, errors: list) -> dict:
    """Read Decision sheet."""
    sheet = SHEET_NAMES["decision"]

    # Find header row containing "Decision Outcome"
    header_row = _find_header_row(ws, "Decision Outcome")
    # Data row: last non-empty row after header
    row = header_row + 1
    for r in range(header_row + 1, header_row + 4):
        val = _cell_value(ws, r, 1)
        if val:
            row = r

    outcome = _cell_value(ws, row, 1)      # A
    rationale = _cell_value(ws, row, 2)     # B
    criteria_set = _cell_value(ws, row, 3)  # C
    decided_by = _cell_value(ws, row, 4)    # D
    decision_date = ws.cell(row=row, column=5).value  # E — raw for date

    if not outcome:
        errors.append(f"Sheet '{sheet}', cell {_cell_ref(1, row)}: Decision Outcome is required")
    elif outcome not in VALID_DECISION_OUTCOMES:
        # Also accept "Conditional" per spec
        valid = VALID_DECISION_OUTCOMES + ["Conditional"]
        if outcome not in valid:
            errors.append(
                f"Sheet '{sheet}', cell {_cell_ref(1, row)}: "
                f"'{outcome}' is not a valid outcome. "
                f"Expected: {', '.join(valid)}"
            )

    return {
        "outcome": outcome,
        "rationale": rationale,
        "criteria_set": criteria_set,
        "decided_by": decided_by,
        "decision_date": _parse_date(decision_date),
    }
