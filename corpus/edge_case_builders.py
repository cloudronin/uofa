"""Build the 10 format edge-case fixtures under corpus/edge-cases/.

Every fixture exercises a specific file-format quirk that `uofa extract` or
the underlying readers (src/uofa_cli/readers/) must handle without crashing.
The corresponding tests live in tests/test_corpus_edge_cases.py; each
fixture's expected error/warning pattern is asserted there.

Fixtures:
    scanned-report.pdf       image-only PDF (pdf_reader.py:39 branch)
    huge-appendix.pdf        200-page reportlab doc, several MB
    tiny-note.txt            single short line
    non-english-headers.xlsx Japanese column titles
    merged-cells.xlsx        merged header range on Credibility Factors
    password-protected.xlsx  encrypted with password "test123"
    corrupted-file.pdf       truncated PDF header, unreadable body
    multi-sheet-workbook.xlsx 12 extra filler sheets beyond the canonical 5
    csv-with-semicolons.csv  semicolon-delimited values
    mixed-encoding.txt       UTF-8 BOM prefix + Shift-JIS body
"""

from __future__ import annotations

import io
from pathlib import Path

# generator.py comes from tests/fixtures/import/ via build.py's sys.path prefix.
from generator import _clean_base_vv40, generate_fixture  # type: ignore[import-not-found]


# ── PDF fixtures (reportlab) ────────────────────────────────────


def _build_scanned_report(out_dir: Path) -> Path:
    """Image-only PDF: render a JPEG containing text, embed as a page image.
    pdfplumber returns no extractable text → pdf_reader.py emits the
    '(image-only PDF — no extractable text)' chunk rather than crashing."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.utils import ImageReader
    from PIL import Image, ImageDraw, ImageFont

    path = out_dir / "scanned-report.pdf"

    img = Image.new("RGB", (1200, 1600), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    for i, line in enumerate([
        "SCANNED REPORT — image-only PDF",
        "This PDF contains no text layer. OCR would be",
        "required to extract its contents. uofa extract",
        "should detect this condition and emit a warning",
        "rather than attempting to read text.",
    ]):
        draw.text((80, 80 + i * 60), line, fill="black", font=font)

    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=70)
    buf.seek(0)

    c = canvas.Canvas(str(path), pagesize=letter)
    c.drawImage(ImageReader(buf), 36, 36, width=540, height=720)
    c.showPage()
    c.save()
    return path


def _build_huge_appendix(out_dir: Path) -> Path:
    """200-page PDF. Tests that readers don't blow memory on long docs."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    path = out_dir / "huge-appendix.pdf"
    c = canvas.Canvas(str(path), pagesize=letter)
    for page in range(1, 201):
        c.drawString(72, 720, f"Appendix page {page} / 200")
        for line in range(50):
            c.drawString(72, 700 - line * 12,
                         f"Line {line + 1}: auto-generated filler content "
                         f"for appendix stress-test page {page}.")
        c.showPage()
    c.save()
    return path


def _build_corrupted_pdf(out_dir: Path) -> Path:
    """PDF header bytes + random garbage, truncated mid-object. pdfplumber
    raises; pdf_reader.py must translate that into a named user error."""
    path = out_dir / "corrupted-file.pdf"
    with path.open("wb") as f:
        f.write(b"%PDF-1.4\n%GARBAGE DATA FOLLOWING, NOT A REAL PDF\n")
        f.write(b"0 obj <<\n/Type /Broken\n/Length 999\n")
        f.write(b"stream\n\x00\xff\x01\xde\xad\xbe\xef" * 10)
        # No endstream, no xref, no trailer — deliberately malformed.
    return path


# ── Excel fixtures (openpyxl) ───────────────────────────────────


def _build_non_english_headers(out_dir: Path) -> Path:
    """Japanese column headers on the Credibility Factors sheet. Parser
    keys off the sheet name and canonical English header row; this file
    demonstrates that a workbook with non-English text elsewhere does
    not crash the import path."""
    import openpyxl
    path = out_dir / "non-english-headers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ノート"  # "Notes" in Japanese
    ws.append(["プロジェクト", "説明"])  # Project, Description
    ws.append(["テストプロジェクト", "日本語のテストデータ"])
    wb.save(path)
    return path


def _build_merged_cells(out_dir: Path) -> Path:
    """Merged header range on Credibility Factors. Exercises openpyxl's
    merged-cell iteration behavior in excel_reader."""
    spec = _clean_base_vv40()
    path = out_dir / "merged-cells.xlsx"
    generate_fixture(spec, path)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Credibility Factors"]
    ws.merge_cells("A1:D1")  # merge the title banner row
    wb.save(path)
    return path


def _build_password_protected(out_dir: Path) -> Path:
    """Encrypt a structurally-valid VV40 workbook with password 'test123'.
    excel_reader.py:100-103 must catch the openpyxl load error and surface
    a named, actionable message rather than a traceback."""
    import msoffcrypto
    spec = _clean_base_vv40()
    plain = out_dir / "_plain-for-encryption.xlsx"
    generate_fixture(spec, plain)

    path = out_dir / "password-protected.xlsx"
    with plain.open("rb") as src, path.open("wb") as dst:
        office_file = msoffcrypto.OfficeFile(src)
        office_file.load_key(password="test123")
        office_file.encrypt("test123", dst)
    plain.unlink()
    return path


def _build_multi_sheet_workbook(out_dir: Path) -> Path:
    """Valid VV40 workbook plus 12 irrelevant filler sheets. Parser must
    identify the canonical sheets by name and ignore the rest."""
    spec = _clean_base_vv40()
    path = out_dir / "multi-sheet-workbook.xlsx"
    generate_fixture(spec, path)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    for i in range(12):
        ws = wb.create_sheet(f"Filler {i + 1}")
        ws.append([f"Unrelated content #{i + 1}"])
        ws.append(["A", "B", "C"])
    wb.save(path)
    return path


# ── Text + CSV fixtures (plain I/O) ─────────────────────────────


def _build_tiny_note(out_dir: Path) -> Path:
    path = out_dir / "tiny-note.txt"
    path.write_text("Brief note — single line.\n", encoding="utf-8")
    return path


def _build_csv_with_semicolons(out_dir: Path) -> Path:
    """Semicolon-delimited CSV. csv_reader uses csv.Sniffer which should
    pick the correct dialect; this fixture verifies that behavior."""
    path = out_dir / "csv-with-semicolons.csv"
    path.write_text(
        "factor;category;level\n"
        "Model form;Model influence;3\n"
        "Mesh convergence;Numerical;4\n"
        "Material properties;Model inputs;2\n",
        encoding="utf-8",
    )
    return path


def _build_mixed_encoding(out_dir: Path) -> Path:
    """Shift-JIS body with a UTF-8 BOM prefix. text_reader's chardet
    fallback must either detect the encoding or apply errors='replace'
    without crashing."""
    path = out_dir / "mixed-encoding.txt"
    body = "テスト内容です。文字化けしないことを確認します。\n"
    with path.open("wb") as f:
        f.write(b"\xef\xbb\xbf")  # UTF-8 BOM
        f.write(body.encode("shift_jis"))
    return path


# ── Entry point ─────────────────────────────────────────────────


def build_all(out_dir: Path) -> list[Path]:
    return [
        _build_scanned_report(out_dir),
        _build_huge_appendix(out_dir),
        _build_tiny_note(out_dir),
        _build_non_english_headers(out_dir),
        _build_merged_cells(out_dir),
        _build_password_protected(out_dir),
        _build_corrupted_pdf(out_dir),
        _build_multi_sheet_workbook(out_dir),
        _build_csv_with_semicolons(out_dir),
        _build_mixed_encoding(out_dir),
    ]
